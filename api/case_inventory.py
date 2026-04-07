# api/case_inventory.py
from __future__ import annotations

from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

import openpyxl


def _iso(x: Any) -> Any:
    if isinstance(x, datetime):
        if x.tzinfo is None:
            x = x.replace(tzinfo=timezone.utc)
        return x.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")
    if isinstance(x, date):
        return x.isoformat()
    return x


def inspect_case_xlsx(case_path: Path, sample_rows: int = 3) -> Dict[str, Any]:
    """
    Inventario liviano:
    - hojas
    - max_row/max_col
    - headers (fila 1)
    - sample (primeras N filas de datos)
    """
    if not case_path.exists():
        raise FileNotFoundError(f"case.xlsx not found: {case_path}")

    wb = openpyxl.load_workbook(case_path, read_only=True, data_only=True)

    sheets: List[Dict[str, Any]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        # header row = fila 1
        header = []
        if max_row >= 1 and max_col >= 1:
            row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if row1:
                header = [("" if v is None else str(v)).strip() for v in row1]

        # sample rows = filas 2..(2+sample_rows-1)
        samples = []
        if max_row >= 2 and max_col >= 1 and header:
            end_row = min(max_row, 1 + sample_rows)
            for r in ws.iter_rows(min_row=2, max_row=end_row, values_only=True):
                item = {}
                for i, col in enumerate(header):
                    if not col:
                        continue
                    val = r[i] if i < len(r) else None
                    item[col] = _iso(val)
                samples.append(item)

        sheets.append(
            {
                "name": name,
                "max_row": int(max_row),
                "max_col": int(max_col),
                "headers": header,
                "sample": samples,
            }
        )

    return {
        "case_path": str(case_path),
        "sheet_count": len(sheets),
        "sheets": sheets,
    }
