from __future__ import annotations

import uuid
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook, Workbook
from sqlalchemy.orm import Session

from api.masterdata_models import Company, Branch, OrgUnit, Employee
from api.demand_models import DemandUnit, PoolTurno
from api.case_data_models import RestriccionEmpleado, AusentismoEmpleado


def _norm_time(x: Any) -> str:
    s = "" if x is None else str(x).strip()
    if not s:
        return ""
    # acepta 7:30 -> 07:30:00
    parts = s.split(":")
    if len(parts) == 2:
        hh, mm = parts
        ss = "00"
    elif len(parts) == 3:
        hh, mm, ss = parts
    else:
        return s
    try:
        return f"{int(hh):02d}:{int(mm):02d}:{int(ss):02d}"
    except Exception:
        return s


def make_case_data_template(out_path: Path) -> Path:
    """
    Template de onboarding usando MISMAS hojas/headers del case.
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("README_IMPORT")
    ws.append(["Carga masiva de datos operativos (Vigatec)"])
    ws.append(["Rellena estas hojas: Dotacion, DemandaUnidad, PoolTurnos, RestriccionesEmpleado, AusentismoEmpleado"])
    ws.append(["IMPORTANTE: org_unit_id debe ser org_unit_key (ej: PERU 805)"])
    ws.append(["IMPORTANTE: shift_id debe existir en CatalogoTurnos del template de la empresa"])
    ws.append(["IMPORTANTE: jornada_id debe existir en Jornadas del template de la empresa"])

    # Dotacion
    ws_d = wb.create_sheet("Dotacion")
    ws_d.append([
        "employee_id","rut","nombre","empresa_id","sucursal_id",
        "org_unit_id","org_unit_nombre","cargo_id","jornada_id",
        "contrato_max_min_semana","rubro","fecha_ingreso","es_estudiante",
        "restricciones","cargo","expertise"
    ])
    ws_d.append(["17375535-K","17375535-K","Nombre Apellido","","",
                 "OU_001","OU_001","Operador","J_6X1_44",
                 2640,"","","NO","","Operador","ALTA"])

    # DemandaUnidad
    ws_dem = wb.create_sheet("DemandaUnidad")
    ws_dem.append(["org_unit_id","dia_semana","inicio","fin","requeridos"])
    ws_dem.append(["OU_001","LUN","07:30:00","21:30:00",2])

    # PoolTurnos
    ws_p = wb.create_sheet("PoolTurnos")
    ws_p.append(["org_unit_id","cargo_id","cargo","dia_semana","shift_id","habilitado"])
    ws_p.append(["OU_001","Operador","Operador","LUN","S_0730_1500_60",1])

    # RestriccionesEmpleado
    ws_r = wb.create_sheet("RestriccionesEmpleado")
    ws_r.append(["employee_id","tipo","valor1","valor2","dia_semana","fecha","hard","penalizacion","detalle"])
    ws_r.append(["17375535-K","DIA_LIBRE_FIJO","LUN","","","","1","100000000",""])

    # AusentismoEmpleado
    ws_a = wb.create_sheet("AusentismoEmpleado")
    ws_a.append(["employee_id","fecha_inicio","fecha_fin","ausentismo","detalle","hard","penalizacion"])
    ws_a.append(["17375535-K","2026-03-10","2026-03-12","LM","Licencia médica","1","0"])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _org_unit_by_key(db: Session, company_id: uuid.UUID, org_unit_key: str) -> OrgUnit | None:
    q = (
        db.query(OrgUnit)
        .join(Branch, OrgUnit.branch_id == Branch.id)
        .filter(Branch.company_id == company_id, OrgUnit.org_unit_key == org_unit_key)
        .first()
    )
    return q


def import_case_data(db: Session, company_id: uuid.UUID, xlsx_path: Path) -> Dict[str, Any]:
    """
    Importa hojas del case a DB (upsert).
    """
    c = db.get(Company, company_id)
    if not c:
        raise ValueError("Company not found")

    wb = load_workbook(xlsx_path, data_only=True)

    stats = {
        "dotacion_upsert": 0,
        "demanda_upsert": 0,
        "pool_upsert": 0,
        "restricciones_insert": 0,
        "ausentismo_insert": 0,
        "errors": [],
    }

    # ---- Dotacion
    if "Dotacion" in wb.sheetnames:
        ws = wb["Dotacion"]
        headers = [str(x.value).strip() if x.value is not None else "" for x in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or all(v is None or str(v).strip()=="" for v in r):
                continue
            employee_id = str(r[idx["employee_id"]]).strip() if "employee_id" in idx and r[idx["employee_id"]] else ""
            if not employee_id:
                continue
            nombre = str(r[idx["nombre"]]).strip() if "nombre" in idx and r[idx["nombre"]] else ""
            org_unit_key = str(r[idx["org_unit_id"]]).strip() if "org_unit_id" in idx and r[idx["org_unit_id"]] else ""
            cargo_id = str(r[idx["cargo_id"]]).strip() if "cargo_id" in idx and r[idx["cargo_id"]] else ""
            jornada_id = str(r[idx["jornada_id"]]).strip() if "jornada_id" in idx and r[idx["jornada_id"]] else ""
            contrato = int(r[idx["contrato_max_min_semana"]]) if "contrato_max_min_semana" in idx and r[idx["contrato_max_min_semana"]] else 0
            expertise = str(r[idx["expertise"]]).strip() if "expertise" in idx and r[idx["expertise"]] else None

            ou = _org_unit_by_key(db, company_id, org_unit_key)
            if not ou:
                stats["errors"].append(f"Dotacion: org_unit_id no existe en sistema: {org_unit_key}")
                continue

            e = db.query(Employee).filter(Employee.org_unit_id == ou.id, Employee.employee_key == employee_id).first()
            if e:
                e.nombre = nombre or e.nombre
                e.cargo_id = cargo_id or e.cargo_id
                e.jornada_id = jornada_id or e.jornada_id
                e.contrato_max_min_semana = contrato or e.contrato_max_min_semana
                e.expertise = expertise or e.expertise
                e.active = True
            else:
                e = Employee(
                    org_unit_id=ou.id,
                    employee_key=employee_id,
                    rut=employee_id,
                    nombre=nombre,
                    cargo_id=cargo_id,
                    jornada_id=jornada_id,
                    contrato_max_min_semana=contrato,
                    expertise=expertise,
                    active=True,
                )
                db.add(e)
            stats["dotacion_upsert"] += 1

    # ---- DemandaUnidad
    if "DemandaUnidad" in wb.sheetnames:
        ws = wb["DemandaUnidad"]
        headers = [str(x.value).strip() if x.value is not None else "" for x in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or all(v is None or str(v).strip()=="" for v in r):
                continue
            org_unit_key = str(r[idx["org_unit_id"]]).strip() if "org_unit_id" in idx and r[idx["org_unit_id"]] else ""
            dia = str(r[idx["dia_semana"]]).strip().upper() if "dia_semana" in idx and r[idx["dia_semana"]] else ""
            inicio = _norm_time(r[idx["inicio"]]) if "inicio" in idx else ""
            fin = _norm_time(r[idx["fin"]]) if "fin" in idx else ""
            req = int(r[idx["requeridos"]]) if "requeridos" in idx and r[idx["requeridos"]] is not None else 0

            ou = _org_unit_by_key(db, company_id, org_unit_key)
            if not ou:
                stats["errors"].append(f"DemandaUnidad: org_unit_id no existe: {org_unit_key}")
                continue

            row = (
                db.query(DemandUnit)
                .filter(DemandUnit.org_unit_id == ou.id, DemandUnit.dia_semana == dia, DemandUnit.inicio == inicio, DemandUnit.fin == fin)
                .first()
            )
            if row:
                row.requeridos = req
                row.active = True
            else:
                db.add(DemandUnit(org_unit_id=ou.id, dia_semana=dia, inicio=inicio, fin=fin, requeridos=req, active=True))
            stats["demanda_upsert"] += 1

    # ---- PoolTurnos
    if "PoolTurnos" in wb.sheetnames:
        ws = wb["PoolTurnos"]
        headers = [str(x.value).strip() if x.value is not None else "" for x in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or all(v is None or str(v).strip()=="" for v in r):
                continue
            org_unit_key = str(r[idx["org_unit_id"]]).strip() if "org_unit_id" in idx and r[idx["org_unit_id"]] else ""
            cargo_id = str(r[idx["cargo_id"]]).strip() if "cargo_id" in idx and r[idx["cargo_id"]] else ""
            dia = str(r[idx["dia_semana"]]).strip().upper() if "dia_semana" in idx and r[idx["dia_semana"]] else ""
            shift_id = str(r[idx["shift_id"]]).strip() if "shift_id" in idx and r[idx["shift_id"]] else ""
            habil = int(r[idx["habilitado"]]) if "habilitado" in idx and r[idx["habilitado"]] is not None else 1

            ou = _org_unit_by_key(db, company_id, org_unit_key)
            if not ou:
                stats["errors"].append(f"PoolTurnos: org_unit_id no existe: {org_unit_key}")
                continue

            row = (
                db.query(PoolTurno)
                .filter(PoolTurno.org_unit_id == ou.id, PoolTurno.cargo_id == cargo_id, PoolTurno.dia_semana == dia, PoolTurno.shift_id == shift_id)
                .first()
            )
            if row:
                row.habilitado = bool(habil)
            else:
                db.add(PoolTurno(org_unit_id=ou.id, cargo_id=cargo_id, dia_semana=dia, shift_id=shift_id, habilitado=bool(habil)))
            stats["pool_upsert"] += 1

    # ---- RestriccionesEmpleado (insert simple)
    if "RestriccionesEmpleado" in wb.sheetnames:
        ws = wb["RestriccionesEmpleado"]
        headers = [str(x.value).strip() if x.value is not None else "" for x in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or all(v is None or str(v).strip()=="" for v in r):
                continue
            employee_id = str(r[idx["employee_id"]]).strip() if "employee_id" in idx and r[idx["employee_id"]] else None
            tipo = str(r[idx["tipo"]]).strip() if "tipo" in idx and r[idx["tipo"]] else ""
            if not tipo:
                continue
            db.add(
                RestriccionEmpleado(
                    company_id=company_id,
                    employee_id=employee_id,
                    tipo=tipo,
                    valor1=str(r[idx["valor1"]]).strip() if "valor1" in idx and r[idx["valor1"]] else None,
                    valor2=str(r[idx["valor2"]]).strip() if "valor2" in idx and r[idx["valor2"]] else None,
                    dia_semana=str(r[idx["dia_semana"]]).strip().upper() if "dia_semana" in idx and r[idx["dia_semana"]] else None,
                    fecha=str(r[idx["fecha"]]).strip() if "fecha" in idx and r[idx["fecha"]] else None,
                    hard=bool(int(r[idx["hard"]])) if "hard" in idx and r[idx["hard"]] is not None else False,
                    penalizacion=int(r[idx["penalizacion"]]) if "penalizacion" in idx and r[idx["penalizacion"]] is not None else 0,
                    detalle=str(r[idx["detalle"]]).strip() if "detalle" in idx and r[idx["detalle"]] else None,
                )
            )
            stats["restricciones_insert"] += 1

    # ---- AusentismoEmpleado (insert simple)
    if "AusentismoEmpleado" in wb.sheetnames:
        ws = wb["AusentismoEmpleado"]
        headers = [str(x.value).strip() if x.value is not None else "" for x in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or all(v is None or str(v).strip()=="" for v in r):
                continue
            employee_id = str(r[idx["employee_id"]]).strip() if "employee_id" in idx and r[idx["employee_id"]] else ""
            if not employee_id:
                continue
            db.add(
                AusentismoEmpleado(
                    company_id=company_id,
                    employee_id=employee_id,
                    fecha_inicio=str(r[idx["fecha_inicio"]]).strip() if "fecha_inicio" in idx and r[idx["fecha_inicio"]] else None,
                    fecha_fin=str(r[idx["fecha_fin"]]).strip() if "fecha_fin" in idx and r[idx["fecha_fin"]] else None,
                    ausentismo=str(r[idx["ausentismo"]]).strip() if "ausentismo" in idx and r[idx["ausentismo"]] else None,
                    detalle=str(r[idx["detalle"]]).strip() if "detalle" in idx and r[idx["detalle"]] else None,
                    hard=bool(int(r[idx["hard"]])) if "hard" in idx and r[idx["hard"]] is not None else True,
                    penalizacion=int(r[idx["penalizacion"]]) if "penalizacion" in idx and r[idx["penalizacion"]] is not None else 0,
                )
            )
            stats["ausentismo_insert"] += 1

    db.commit()
    return stats