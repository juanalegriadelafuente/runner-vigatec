# api/company_config.py
from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
from fastapi import APIRouter, HTTPException, Query
from pydantic import BaseModel, Field, field_validator

from api.db import get_db
from api.models import Run
from sqlalchemy.orm import Session
from fastapi import Depends

DOW = {"LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"}

router = APIRouter(tags=["company-config"])


def _storage_base() -> Path:
    # Debe apuntar al mismo storage que usas para runs.
    # Ajusta si tu proyecto usa otra env var.
    return Path(os.getenv("STORAGE_DIR", "/app/storage"))


def _slug(s: str) -> str:
    s = s.strip()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-zA-Z0-9_\-\.]", "_", s)
    return s[:200] if s else "UNKNOWN"


def _company_dir(empresa_id: str) -> Path:
    return _storage_base() / "company_config" / _slug(empresa_id)


def _index_path() -> Path:
    return _storage_base() / "company_config" / "_index.json"


def _load_index() -> Dict[str, str]:
    p = _index_path()
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_index(idx: Dict[str, str]) -> None:
    p = _index_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(idx, ensure_ascii=False, indent=2), encoding="utf-8")


def _patterns_path(empresa_id: str) -> Path:
    return _company_dir(empresa_id) / "demand_patterns.json"


def _normalize_time_str(x: str) -> str:
    """
    Acepta "7:30", "07:30", "07:30:00" y devuelve "HH:MM:SS".
    """
    s = str(x).strip()
    if not s:
        return s
    parts = s.split(":")
    if len(parts) == 2:
        hh, mm = parts
        ss = "00"
    elif len(parts) == 3:
        hh, mm, ss = parts
    else:
        raise ValueError(f"Formato de hora inválido: {x}")
    hh = hh.zfill(2)
    mm = mm.zfill(2)
    ss = ss.zfill(2)
    return f"{hh}:{mm}:{ss}"


class DemandPatternRow(BaseModel):
    org_unit_id: str = Field(..., min_length=1)
    dia_semana: str = Field(..., min_length=3, max_length=3)
    inicio: str = Field(..., min_length=1)
    fin: str = Field(..., min_length=1)
    requeridos: int = Field(..., ge=0)

    @field_validator("dia_semana")
    @classmethod
    def _v_dow(cls, v: str) -> str:
        v = v.strip().upper()
        if v not in DOW:
            raise ValueError(f"dia_semana debe ser uno de {sorted(DOW)}")
        return v

    @field_validator("inicio", "fin")
    @classmethod
    def _v_time(cls, v: str) -> str:
        return _normalize_time_str(v)


class DemandPatternPut(BaseModel):
    patterns: List[DemandPatternRow]


class DemandPatternGet(BaseModel):
    empresa_id: str
    patterns: List[DemandPatternRow]


class CompanyList(BaseModel):
    companies: List[str]


def extract_empresa_ids_from_case(case_path: Path) -> List[str]:
    """
    Lee la hoja Dotacion y devuelve empresa_id únicos.
    """
    wb = openpyxl.load_workbook(case_path, read_only=True, data_only=True)
    if "Dotacion" not in wb.sheetnames:
        return []
    ws = wb["Dotacion"]
    rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    header = next(rows, None)
    if not header:
        return []

    header_map = {str(v).strip(): i for i, v in enumerate(header) if v is not None}
    if "empresa_id" not in header_map:
        return []

    idx = header_map["empresa_id"]
    empresas: Set[str] = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or idx >= len(r):
            continue
        v = r[idx]
        if v is None:
            continue
        s = str(v).strip()
        if s:
            empresas.add(s)

    return sorted(empresas)


def extract_org_units_from_case(case_path: Path) -> Set[str]:
    wb = openpyxl.load_workbook(case_path, read_only=True, data_only=True)
    if "Dotacion" not in wb.sheetnames:
        return set()
    ws = wb["Dotacion"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        return set()

    header_map = {str(v).strip(): i for i, v in enumerate(header) if v is not None}
    if "org_unit_id" not in header_map:
        return set()

    idx = header_map["org_unit_id"]
    ous: Set[str] = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or idx >= len(r):
            continue
        v = r[idx]
        if v is None:
            continue
        s = str(v).strip()
        if s:
            ous.add(s)
    return ous


def apply_company_demand_patterns_to_case(case_path: Path, empresa_id: str) -> Dict[str, Any]:
    """
    Sobrescribe la hoja DemandaUnidad del case.xlsx con lo guardado por empresa.
    NO toca solver, solo arma el input correcto.
    """
    p = _patterns_path(empresa_id)
    if not p.exists():
        return {"applied": False, "reason": "no_patterns_for_company", "empresa_id": empresa_id}

    data = json.loads(p.read_text(encoding="utf-8"))
    patterns = [DemandPatternRow(**row) for row in data.get("patterns", [])]

    org_units_in_case = extract_org_units_from_case(case_path)
    filtered = [r for r in patterns if r.org_unit_id in org_units_in_case] if org_units_in_case else patterns

    wb = openpyxl.load_workbook(case_path)
    if "DemandaUnidad" in wb.sheetnames:
        ws_old = wb["DemandaUnidad"]
        wb.remove(ws_old)

    ws = wb.create_sheet("DemandaUnidad")
    ws.append(["org_unit_id", "dia_semana", "inicio", "fin", "requeridos"])

    for r in filtered:
        ws.append([r.org_unit_id, r.dia_semana, r.inicio, r.fin, r.requeridos])

    wb.save(case_path)

    return {
        "applied": True,
        "empresa_id": empresa_id,
        "rows_written": len(filtered),
        "org_units_in_case": sorted(org_units_in_case),
    }


@router.get("/companies", response_model=CompanyList)
def list_companies():
    idx = _load_index()
    # si no hay index, igual mostramos por carpetas existentes
    base = _storage_base() / "company_config"
    companies = sorted(set(idx.values()))
    if base.exists():
        for d in base.iterdir():
            if d.is_dir() and d.name != "_index.json":
                # no sabemos el nombre real, pero al menos no rompemos
                pass
    return CompanyList(companies=companies)


@router.get("/companies/{empresa_id}/demand-patterns", response_model=DemandPatternGet)
def get_company_demand_patterns(empresa_id: str):
    p = _patterns_path(empresa_id)
    if not p.exists():
        return DemandPatternGet(empresa_id=empresa_id, patterns=[])
    data = json.loads(p.read_text(encoding="utf-8"))
    patterns = [DemandPatternRow(**row) for row in data.get("patterns", [])]
    return DemandPatternGet(empresa_id=empresa_id, patterns=patterns)


@router.put("/companies/{empresa_id}/demand-patterns", response_model=DemandPatternGet)
def put_company_demand_patterns(empresa_id: str, body: DemandPatternPut):
    # Persistimos por empresa
    d = _company_dir(empresa_id)
    d.mkdir(parents=True, exist_ok=True)

    payload = {"empresa_id": empresa_id, "patterns": [r.model_dump() for r in body.patterns]}
    _patterns_path(empresa_id).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    # Actualiza index (para listarlas bonito)
    idx = _load_index()
    idx[_slug(empresa_id)] = empresa_id
    _save_index(idx)

    return DemandPatternGet(empresa_id=empresa_id, patterns=body.patterns)


@router.post("/runs/{run_id}/apply-company-demand-patterns")
def apply_company_demand_patterns_to_run(
    run_id: str,
    empresa_id: Optional[str] = Query(default=None, description="Si el case tiene más de una empresa, especifica cuál."),
    db: Session = Depends(get_db),
):
    run: Run | None = db.get(Run, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    case_path = Path(run.case_path)
    if not case_path.exists():
        raise HTTPException(status_code=404, detail="case.xlsx not found for run")

    if empresa_id is None:
        empresas = extract_empresa_ids_from_case(case_path)
        if len(empresas) == 0:
            raise HTTPException(status_code=400, detail="No se pudo determinar empresa_id desde Dotacion")
        if len(empresas) > 1:
            raise HTTPException(status_code=400, detail=f"Case tiene múltiples empresas: {empresas}. Debes indicar empresa_id.")
        empresa_id = empresas[0]

    info = apply_company_demand_patterns_to_case(case_path, empresa_id)
    return info