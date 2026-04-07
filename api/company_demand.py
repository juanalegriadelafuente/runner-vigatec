from __future__ import annotations

import json
import re
from datetime import datetime, timezone, time
from pathlib import Path
from typing import Any, Optional, List, Dict

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field, field_validator
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from api.db import get_db
from api.models import Run
from api.storage import safe_resolve_under

router = APIRouter(tags=["company-config"])


# -------------------------
# Utils
# -------------------------
DOW_ALLOWED = {"LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"}
STORAGE_ROOT = Path("/app/storage")
COMPANY_CFG_DIR = STORAGE_ROOT / "company_configs"


def _iso_utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _as_str(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    # Pandas a veces trae "nan" como string
    if s.lower() == "nan":
        return ""
    return s


def _slug_empresa_id(empresa_id: str) -> str:
    s = _as_str(empresa_id)
    if not s:
        raise ValueError("empresa_id vacío")
    s = re.sub(r"[^a-zA-Z0-9._-]+", "_", s).strip("_")
    if not s:
        raise ValueError("empresa_id inválido")
    return s


def _normalize_dow(x: Any) -> str:
    s = _as_str(x).upper()
    s = s.replace("MIÉ", "MIE").replace("MIÉ.", "MIE")
    s = s.replace("SÁB", "SAB").replace("SÁB.", "SAB")
    s = s.replace("DOM.", "DOM").replace("LUN.", "LUN").replace("MAR.", "MAR").replace("MIE.", "MIE").replace("JUE.", "JUE").replace("VIE.", "VIE").replace("SAB.", "SAB")
    if s not in DOW_ALLOWED:
        raise ValueError(f"dia_semana inválido: {s} (usa LUN..DOM)")
    return s


def _normalize_time(x: Any) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    if isinstance(x, time):
        return x.strftime("%H:%M:%S")
    if isinstance(x, datetime):
        return x.time().strftime("%H:%M:%S")
    if hasattr(x, "to_pydatetime"):
        try:
            dt = x.to_pydatetime()
            if isinstance(dt, datetime):
                return dt.time().strftime("%H:%M:%S")
        except Exception:
            pass

    s = _as_str(x)
    if not s:
        return ""
    if len(s) == 5 and s[2] == ":":
        return f"{s}:00"
    if len(s) == 8 and s[2] == ":" and s[5] == ":":
        return s
    if len(s) > 8 and s[2] == ":":
        return s[:8]
    raise ValueError(f"hora inválida: {s} (esperado HH:MM o HH:MM:SS)")


def _company_dir(empresa_id: str) -> Path:
    COMPANY_CFG_DIR.mkdir(parents=True, exist_ok=True)
    slug = _slug_empresa_id(empresa_id)
    candidate = COMPANY_CFG_DIR / slug
    # asegurar que no haya path traversal
    return safe_resolve_under(COMPANY_CFG_DIR, candidate)


def _company_demand_file(empresa_id: str) -> Path:
    base = _company_dir(empresa_id)
    candidate = base / "demand_patterns.json"
    return safe_resolve_under(base, candidate)


# -------------------------
# Models (API)
# -------------------------
class DemandPatternRow(BaseModel):
    org_unit_id: str
    cargo_id: str
    dia_semana: str
    inicio: str
    fin: str
    requeridos: int = Field(ge=0)
    cargo: Optional[str] = None  # opcional

    @field_validator("org_unit_id", "cargo_id")
    @classmethod
    def _not_empty(cls, v: str) -> str:
        v = _as_str(v)
        if not v:
            raise ValueError("no puede ser vacío")
        return v

    @field_validator("dia_semana")
    @classmethod
    def _valid_dow(cls, v: str) -> str:
        return _normalize_dow(v)

    @field_validator("inicio", "fin")
    @classmethod
    def _valid_time(cls, v: str) -> str:
        return _normalize_time(v)


class CompanyDemandGetResponse(BaseModel):
    empresa_id: str
    updated_at_utc: str
    count: int
    rows: List[DemandPatternRow]


class CompanyDemandPutRequest(BaseModel):
    rows: List[DemandPatternRow]
    note: Optional[str] = None


class CompanyDemandPutResponse(BaseModel):
    empresa_id: str
    saved_rows: int
    updated_at_utc: str
    path: str


class CompanyListItem(BaseModel):
    empresa_id: str
    updated_at_utc: str
    count: int


class CompanyListResponse(BaseModel):
    companies: List[CompanyListItem]


# -------------------------
# Read/Write Config (Company)
# -------------------------
def load_company_demand_patterns(empresa_id: str) -> Optional[dict]:
    p = _company_demand_file(empresa_id)
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Invalid company demand_patterns.json: {e}")


def save_company_demand_patterns(empresa_id: str, rows: List[DemandPatternRow], note: Optional[str] = None) -> dict:
    p = _company_demand_file(empresa_id)
    payload = {
        "empresa_id": empresa_id,
        "updated_at_utc": _iso_utc_now(),
        "note": note,
        "rows": [r.model_dump() for r in rows],
    }
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


# -------------------------
# Apply to case.xlsx
# -------------------------
def extract_empresa_id_from_case(case_path: Path) -> str:
    if not case_path.exists():
        raise HTTPException(status_code=404, detail="case.xlsx not found")

    try:
        df = pd.read_excel(case_path, sheet_name="Dotacion")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cannot read Dotacion: {e}")

    if "empresa_id" not in df.columns:
        raise HTTPException(status_code=400, detail="Dotacion missing column 'empresa_id'")

    vals = (
        df["empresa_id"]
        .dropna()
        .astype(str)
        .map(lambda x: x.strip())
        .replace({"nan": "", "None": ""})
    )
    vals = vals[vals != ""].unique().tolist()

    if len(vals) == 0:
        raise HTTPException(status_code=400, detail="No empresa_id found in Dotacion")
    if len(vals) > 1:
        raise HTTPException(
            status_code=400,
            detail=f"Multiple empresa_id found in Dotacion (expected 1): {vals}",
        )
    return vals[0]


def _ensure_sheet(wb, name: str) -> Worksheet:
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(title=name)


def _clear_worksheet(ws: Worksheet):
    ws.delete_rows(1, ws.max_row if ws.max_row > 0 else 1)


def _write_table(ws: Worksheet, headers: List[str], data_rows: List[List[Any]]):
    _clear_worksheet(ws)
    ws.append(headers)
    for row in data_rows:
        ws.append(row)


def apply_demand_patterns_to_case(case_path: Path, rows: List[DemandPatternRow], rebuild_demanda_unidad: bool = True) -> None:
    wb = load_workbook(case_path)

    # NecesidadMinimos
    ws_nm = _ensure_sheet(wb, "NecesidadMinimos")
    headers_nm = ["org_unit_id", "cargo_id", "cargo", "dia_semana", "inicio", "fin", "requeridos"]
    data_nm: List[List[Any]] = []
    for r in rows:
        data_nm.append([
            r.org_unit_id,
            r.cargo_id,
            r.cargo or "",
            r.dia_semana,
            r.inicio,
            r.fin,
            int(r.requeridos),
        ])
    _write_table(ws_nm, headers_nm, data_nm)

    # DemandaUnidad (opcional)
    if rebuild_demanda_unidad:
        ws_du = _ensure_sheet(wb, "DemandaUnidad")
        headers_du = ["org_unit_id", "dia_semana", "inicio", "fin", "requeridos"]
        tmp = pd.DataFrame([r.model_dump() for r in rows])
        if not tmp.empty:
            grp = (
                tmp.groupby(["org_unit_id", "dia_semana", "inicio", "fin"], as_index=False)["requeridos"]
                .sum()
                .sort_values(["org_unit_id", "dia_semana", "inicio", "fin"])
            )
            data_du = grp[headers_du].values.tolist()
        else:
            data_du = []
        _write_table(ws_du, headers_du, data_du)

    wb.save(case_path)


# -------------------------
# Routes (Company-level)
# -------------------------
@router.get("/companies", response_model=CompanyListResponse)
def list_companies():
    COMPANY_CFG_DIR.mkdir(parents=True, exist_ok=True)
    companies: List[CompanyListItem] = []

    for d in sorted(COMPANY_CFG_DIR.glob("*")):
        if not d.is_dir():
            continue
        f = d / "demand_patterns.json"
        if not f.exists():
            continue
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            empresa_id = data.get("empresa_id", d.name)
            rows = data.get("rows", [])
            updated = data.get("updated_at_utc", "")
            companies.append(CompanyListItem(empresa_id=empresa_id, updated_at_utc=updated, count=len(rows)))
        except Exception:
            # si hay un json malo, lo ignoramos para no romper el listing
            continue

    return CompanyListResponse(companies=companies)


@router.get("/companies/{empresa_id}/demand-patterns", response_model=CompanyDemandGetResponse)
def get_company_demand_patterns(empresa_id: str):
    data = load_company_demand_patterns(empresa_id)
    if data is None:
        raise HTTPException(status_code=404, detail="No demand-patterns configured for this empresa_id yet")

    raw_rows = data.get("rows", [])
    rows = [DemandPatternRow(**r) for r in raw_rows]

    return CompanyDemandGetResponse(
        empresa_id=data.get("empresa_id", empresa_id),
        updated_at_utc=data.get("updated_at_utc", ""),
        count=len(rows),
        rows=rows,
    )


@router.put("/companies/{empresa_id}/demand-patterns", response_model=CompanyDemandPutResponse)
def put_company_demand_patterns(empresa_id: str, req: CompanyDemandPutRequest):
    payload = save_company_demand_patterns(empresa_id, req.rows, note=req.note)
    f = _company_demand_file(empresa_id)
    return CompanyDemandPutResponse(
        empresa_id=payload["empresa_id"],
        saved_rows=len(req.rows),
        updated_at_utc=payload["updated_at_utc"],
        path=str(f),
    )


# -------------------------
# Routes (Run helpers)
# -------------------------
@router.post("/runs/{run_id}/apply-company-demand-patterns")
def apply_company_demand_patterns_to_run(run_id: str, db: Session = Depends(get_db)):
    """
    Re-aplica la config de empresa al case.xlsx de un run ya creado.
    Útil si cambiaste el patrón de demanda y quieres re-correr el motor.
    """
    run: Run | None = db.get(Run, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    case_path = Path(run.case_path)
    empresa_id = extract_empresa_id_from_case(case_path)

    data = load_company_demand_patterns(empresa_id)
    if data is None:
        raise HTTPException(status_code=404, detail=f"No demand-patterns configured for empresa_id='{empresa_id}'")

    rows = [DemandPatternRow(**r) for r in data.get("rows", [])]
    apply_demand_patterns_to_case(case_path, rows, rebuild_demanda_unidad=True)

    return {
        "ok": True,
        "run_id": str(run.id),
        "empresa_id": empresa_id,
        "applied_rows": len(rows),
        "applied_at_utc": _iso_utc_now(),
    }
