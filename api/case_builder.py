from __future__ import annotations

import calendar
import shutil
import uuid
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Dict, List, Tuple, Set

from sqlalchemy.orm import Session

from api.masterdata_models import Company, Branch, OrgUnit, Employee
from api.demand_models import DemandUnit, PoolTurno
from api.case_data_models import RestriccionEmpleado, AusentismoEmpleado
DOW_MAP = {0: "LUN", 1: "MAR", 2: "MIE", 3: "JUE", 4: "VIE", 5: "SAB", 6: "DOM"}


def _parse_month(month_str: str) -> Tuple[int, int]:
    try:
        y_s, m_s = month_str.split("-", 1)
        y = int(y_s)
        m = int(m_s)
        if m < 1 or m > 12:
            raise ValueError
        return y, m
    except Exception:
        raise ValueError("month must be YYYY-MM (e.g. 2026-03)")


def _monday_of_week(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _sundays_in_month(year: int, month: int) -> List[date]:
    last_day = calendar.monthrange(year, month)[1]
    return [date(year, month, day) for day in range(1, last_day + 1) if date(year, month, day).weekday() == 6]


@dataclass
class PlanningCycle:
    month: str
    start_date: date
    weeks: int
    end_date: date


def compute_cycle(month: str) -> PlanningCycle:
    y, m = _parse_month(month)
    first = date(y, m, 1)
    start = _monday_of_week(first)
    sundays = _sundays_in_month(y, m)
    weeks = len(sundays) if sundays else 4
    end = start + timedelta(days=weeks * 7 - 1)
    return PlanningCycle(month=month, start_date=start, weeks=weeks, end_date=end)


def _as_iso_z(d: date) -> str:
    return f"{d.isoformat()}T00:00:00Z"


def _normalize_time_str(x: Any) -> str:
    s = "" if x is None else str(x).strip()
    if not s:
        return ""
    parts = s.split(":")
    if len(parts) == 2:
        hh, mm = parts
        ss = "00"
    elif len(parts) == 3:
        hh, mm, ss = parts
    else:
        return s
    try:
        return f"{int(hh):02d}:{int(mm):02d}:{int(ss):02d}"
    except Exception:
        return s


def _ensure_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(title=name)


def _reset_sheet(ws, headers: List[str]) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.append(headers)


def _set_parametros(case_wb, start_date: date, weeks: int) -> None:
    ws = _ensure_sheet(case_wb, "Parametros")

    # headers si no existen
    if ws.max_row < 1 or ws["A1"].value is None:
        ws["A1"].value = "parametro"
        ws["B1"].value = "valor"
        ws["C1"].value = "detalle"

    def set_param(key: str, value: Any, detail: str = ""):
        found = False
        for row in ws.iter_rows(min_row=2):
            k = row[0].value
            if k is None:
                continue
            if str(k).strip() == key:
                row[1].value = value
                if len(row) >= 3:
                    row[2].value = detail
                found = True
                break
        if not found:
            ws.append([key, value, detail])

    set_param("fecha_inicio_mes", _as_iso_z(start_date), "Lunes de inicio (calculado)")
    set_param("semanas", int(weeks), "Horizonte (= # domingos del mes)")


def _read_set_from_sheet(wb, sheet_name: str, col_name: str, fallback_col: int = 1) -> Set[str]:
    """
    Lee un set de valores desde una hoja dada, buscando la columna por header.
    Si no encuentra header, usa fallback_col (1-indexed).
    """
    if sheet_name not in wb.sheetnames:
        return set()

    ws = wb[sheet_name]
    if ws.max_row < 2:
        return set()

    # encontrar índice de columna por header
    header_row = [c.value for c in ws[1]]
    idx = None
    for i, v in enumerate(header_row, start=1):
        if v is None:
            continue
        if str(v).strip() == col_name:
            idx = i
            break
    if idx is None:
        idx = fallback_col

    out: Set[str] = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=idx).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            out.add(s)
    return out


def build_case_from_db(
    db: Session,
    company_id: uuid.UUID,
    month: str,
    template_path: Path,
    out_case_path: Path,
    branch_id: uuid.UUID | None = None,
    org_unit_id: uuid.UUID | None = None,
) -> Dict[str, Any]:
    """
    Crea un case.xlsx desde template + DB:
      - Parametros (fecha_inicio_mes, semanas)
      - Dotacion
      - DemandaUnidad
      - PoolTurnos
      - PlanPrevio: headers por ahora

    + Preflight duro:
      - shift_id del Pool debe existir en CatalogoTurnos del template
      - jornada_id de Dotacion debe existir en Jornadas del template
      - al menos 1 empleado, demanda y pool (si no, te aviso con error claro)
    """
    from openpyxl import load_workbook

    company: Company | None = db.get(Company, company_id)
    if not company:
        raise ValueError("Company not found")

    if not template_path.exists():
        raise ValueError("Company template not found")

    out_case_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, out_case_path)

    cycle = compute_cycle(month)

    wb = load_workbook(out_case_path)

    # ✅ sets del template (para validar)
    template_shift_ids = _read_set_from_sheet(wb, "CatalogoTurnos", "shift_id", fallback_col=1)
    template_jornadas = _read_set_from_sheet(wb, "Jornadas", "jornada_id", fallback_col=1)

    # --- Parametros
    _set_parametros(wb, cycle.start_date, cycle.weeks)

    # --- Employees (Employee -> OrgUnit -> Branch)
    emp_q = (
        db.query(Employee, OrgUnit, Branch)
        .join(OrgUnit, Employee.org_unit_id == OrgUnit.id)
        .join(Branch, OrgUnit.branch_id == Branch.id)
        .filter(Branch.company_id == company_id)
    )
    if org_unit_id:
        emp_q = emp_q.filter(OrgUnit.id == org_unit_id)
    elif branch_id:
        emp_q = emp_q.filter(Branch.id == branch_id)
    rows_emp = emp_q.all()
    if len(rows_emp) == 0:
        raise ValueError("No hay empleados cargados para esta empresa (no se puede armar Dotación).")

    # --- Demanda
    dem_q = (
        db.query(DemandUnit, OrgUnit, Branch)
        .join(OrgUnit, DemandUnit.org_unit_id == OrgUnit.id)
        .join(Branch, OrgUnit.branch_id == Branch.id)
        .filter(Branch.company_id == company_id, DemandUnit.active == True)  # noqa: E712
    )
    if org_unit_id:
        dem_q = dem_q.filter(OrgUnit.id == org_unit_id)
    elif branch_id:
        dem_q = dem_q.filter(Branch.id == branch_id)
    rows_dem = dem_q.all()
    if len(rows_dem) == 0:
        raise ValueError("No hay DemandaUnidad activa cargada para esta empresa (demanda=0).")

    # --- Pool
    pool_q = (
        db.query(PoolTurno, OrgUnit, Branch)
        .join(OrgUnit, PoolTurno.org_unit_id == OrgUnit.id)
        .join(Branch, OrgUnit.branch_id == Branch.id)
        .filter(Branch.company_id == company_id)
    )
    if org_unit_id:
        pool_q = pool_q.filter(OrgUnit.id == org_unit_id)
    elif branch_id:
        pool_q = pool_q.filter(Branch.id == branch_id)
    rows_pool = pool_q.all()
    if len(rows_pool) == 0:
        raise ValueError("No hay PoolTurnos cargado para esta empresa.")

    # ✅ Preflight: jornadas existen
    missing_jornadas = sorted(
        {e.jornada_id for (e, _ou, _br) in rows_emp if str(e.jornada_id).strip()} - template_jornadas
    )
    if missing_jornadas:
        raise ValueError(
            "Jornadas no encontradas en el template (hoja Jornadas): "
            + ", ".join(missing_jornadas)
        )

    # ✅ Preflight: shift_ids existen
    pool_shift_ids = {p.shift_id for (p, _ou, _br) in rows_pool if str(p.shift_id).strip()}
    missing_shift_ids = sorted([s for s in pool_shift_ids if s not in template_shift_ids and s != "LIBRE"])
    if missing_shift_ids:
        raise ValueError(
            "Shift IDs del Pool NO existen en el template (hoja CatalogoTurnos): "
            + ", ".join(missing_shift_ids)
        )

    # --- Dotacion sheet
    ws_dot = _ensure_sheet(wb, "Dotacion")
    dot_headers = [
        "employee_id", "rut", "nombre", "empresa_id", "sucursal_id",
        "org_unit_id", "org_unit_nombre", "cargo_id", "jornada_id",
        "contrato_max_min_semana", "rubro", "fecha_ingreso", "es_estudiante",
        "restricciones", "cargo", "expertise",
    ]
    _reset_sheet(ws_dot, dot_headers)

    for e, ou, br in rows_emp:
        ws_dot.append(
            [
                e.employee_key,
                e.rut or e.employee_key,
                e.nombre,
                company.name,
                br.code,
                ou.org_unit_key,
                ou.name,
                e.cargo_id,
                e.jornada_id,
                int(e.contrato_max_min_semana),
                "",
                "",
                "NO",
                "",
                e.cargo_id,
                e.expertise or "",
            ]
        )

    # --- DemandaUnidad sheet
    # Dos curvas: requeridos (mínimo operativo) + requeridos_ideal (demanda real)
    ws_dem = _ensure_sheet(wb, "DemandaUnidad")
    dem_headers = ["org_unit_id", "dia_semana", "inicio", "fin", "requeridos", "requeridos_ideal"]
    _reset_sheet(ws_dem, dem_headers)

    for d, ou, _br in rows_dem:
        # Si requeridos_ideal es None, usar requeridos como fallback
        ideal = d.requeridos_ideal if d.requeridos_ideal is not None else d.requeridos
        ws_dem.append(
            [
                ou.org_unit_key,
                d.dia_semana,
                _normalize_time_str(d.inicio),
                _normalize_time_str(d.fin),
                int(d.requeridos),
                int(ideal),
            ]
        )

    # --- PoolTurnos sheet
    ws_pool = _ensure_sheet(wb, "PoolTurnos")
    pool_headers = ["org_unit_id", "cargo_id", "cargo", "dia_semana", "shift_id", "habilitado"]
    _reset_sheet(ws_pool, pool_headers)

    for p, ou, _br in rows_pool:
        ws_pool.append(
            [
                ou.org_unit_key,
                p.cargo_id,
                p.cargo_id,
                p.dia_semana,
                p.shift_id,
                1 if p.habilitado else 0,
            ]
        )

    # --- PlanPrevio sheet (headers solamente)
    ws_prev = _ensure_sheet(wb, "PlanPrevio")
    prev_headers = ["employee_id", "fecha", "dia_semana", "org_unit_id", "cargo", "shift_id", "es_saliente", "nota"]
    _reset_sheet(ws_prev, prev_headers)
        # --- RestriccionesEmpleado
    ws_res = _ensure_sheet(wb, "RestriccionesEmpleado")
    res_headers = ["employee_id","tipo","valor1","valor2","dia_semana","fecha","hard","penalizacion","detalle"]
    _reset_sheet(ws_res, res_headers)

    res_q = db.query(RestriccionEmpleado).filter(RestriccionEmpleado.company_id == company_id)
    if rows_emp:  # si hay scope, solo incluir restricciones de los empleados del scope
        scoped_keys = {e.employee_key for (e, _ou, _br) in rows_emp}
        res_q = res_q.filter(RestriccionEmpleado.employee_id.in_(scoped_keys))
    res_rows = res_q.all()
    for r in res_rows:
        ws_res.append([
            r.employee_id or "",
            r.tipo,
            r.valor1 or "",
            r.valor2 or "",
            r.dia_semana or "",
            r.fecha or "",
            1 if r.hard else 0,
            int(r.penalizacion or 0),
            r.detalle or "",
        ])

    # --- AusentismoEmpleado
    ws_aus = _ensure_sheet(wb, "AusentismoEmpleado")
    aus_headers = ["employee_id","fecha_inicio","fecha_fin","ausentismo","detalle","hard","penalizacion"]
    _reset_sheet(ws_aus, aus_headers)

    aus_q = db.query(AusentismoEmpleado).filter(AusentismoEmpleado.company_id == company_id)
    if rows_emp:  # solo ausentismos de empleados del scope
        scoped_keys = {e.employee_key for (e, _ou, _br) in rows_emp}
        aus_q = aus_q.filter(AusentismoEmpleado.employee_id.in_(scoped_keys))
    aus_rows = aus_q.all()
    for a in aus_rows:
        ws_aus.append([
            a.employee_id,
            a.fecha_inicio or "",
            a.fecha_fin or "",
            a.ausentismo or "",
            a.detalle or "",
            1 if a.hard else 0,
            int(a.penalizacion or 0),
        ])
    wb.save(out_case_path)

    return {
        "month": month,
        "cycle_start": cycle.start_date.isoformat(),
        "cycle_end": cycle.end_date.isoformat(),
        "weeks": cycle.weeks,
        "employees": len(rows_emp),
        "demand_rows": len(rows_dem),
        "pool_rows": len(rows_pool),
        "note": "OK preflight: jornadas y shift_ids consistentes con template. PlanPrevio aún vacío (demo).",
    }