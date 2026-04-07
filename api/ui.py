from __future__ import annotations

import os
import uuid
import glob
from dataclasses import dataclass
from pathlib import Path
import shutil
from urllib.parse import quote_plus
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import api.plan_models  # noqa: F401

import re
from openpyxl import load_workbook

import csv
from datetime import datetime, timezone, date, timedelta
from collections import defaultdict

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, Response
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func

from api.db import get_db
from api.masterdata_models import Company, Branch, OrgUnit, Employee
from api.holiday_models import Holiday, HolidayCl
from api.holiday_models import Holiday, HolidayCl
from api.demand_models import DemandUnit, PoolTurno
from api.models import Run
from api.storage import ensure_dirs, file_mtime_utc, get_run_paths
from api.qa import load_qa, qa_message, qa_status, qa_summary
from api.plan_models import PlanOverride
from api.case_builder import build_case_from_db
from api.vocab import CAT_AUSENTISMO, CAT_JORNADA, CAT_RESTR_TIPO, list_vocab, seed_company_vocab
AUS_CODES_FALLBACK = ["LM", "VAC", "PA"]

_DOW_ES = {0: "Lunes", 1: "Martes", 2: "Miércoles", 3: "Jueves", 4: "Viernes", 5: "Sábado", 6: "Domingo"}

RESTR_TIPOS_FALLBACK = [
    "DIA_LIBRE_FIJO",
    "NO_TRABAJAR_FECHA",
    "PROHIBIR_TURNO",
    "SOLO_TURNOS_TIPO",
    "VENTANA_HORARIA",
    "REGLA_APERTURA_CARGO",
    "REGLA_CIERRE_CARGO",
    "ROL_APERTURA",
    "ROL_CIERRE",
]


from api.case_data_models import RestriccionEmpleado, AusentismoEmpleado


from datetime import datetime, timezone, date
import calendar

from datetime import timedelta

from datetime import datetime, timezone

from api.rbac_models import User, UserScope
from api.rbac import (
    get_current_user,
    filter_companies,
    filter_branches,
    filter_org_units,
    can_see_company,
)

from worker.celery_app import celery_app

TEMPLATES = Jinja2Templates(directory=str(Path(__file__).parent / "templates"))
router = APIRouter(prefix="/ui", tags=["ui"])


# -------------------------
# UI Context (cookies)
# -------------------------
@dataclass
class UiCtx:
    user_id: Optional[uuid.UUID] = None
    company_id: Optional[uuid.UUID] = None
    branch_id: Optional[uuid.UUID] = None
    org_unit_id: Optional[uuid.UUID] = None


def _parse_uuid(v: Optional[str]) -> Optional[uuid.UUID]:
    if not v:
        return None
    try:
        return uuid.UUID(v)
    except Exception:
        return None


def _load_ctx(request: Request) -> UiCtx:
    return UiCtx(
        user_id=_parse_uuid(request.cookies.get("ctx_user_id")),
        company_id=_parse_uuid(request.cookies.get("ctx_company_id")),
        branch_id=_parse_uuid(request.cookies.get("ctx_branch_id")),
        org_unit_id=_parse_uuid(request.cookies.get("ctx_org_unit_id")),
    )


def _fecha_es(dt_obj) -> str:
    dias = {0: "lunes", 1: "martes", 2: "miércoles", 3: "jueves", 4: "viernes", 5: "sábado", 6: "domingo"}
    meses = {1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio",
             7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"}
    return f"{dias[dt_obj.weekday()]} {dt_obj.day} de {meses[dt_obj.month]} de {dt_obj.year}"


def _redirect(url: str) -> RedirectResponse:
    return RedirectResponse(url=url, status_code=303)


def _enrich(db: Session, request: Request, data: Dict[str, Any]) -> Dict[str, Any]:
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)

    companies = filter_companies(db, current_user) if current_user else []
    branches = filter_branches(db, current_user, ctx.company_id) if current_user else []
    org_units = filter_org_units(db, current_user, ctx.branch_id, ctx.company_id) if current_user else []

    data.update(
        {
            "ctx": ctx,
            "current_user": current_user,
            "ctx_companies": companies,
            "ctx_branches": branches,
            "ctx_org_units": org_units,
        }
    )
    return data


# -------------------------
# Context setter (cookies)
# -------------------------
@router.get("/app/set-context")
def ui_set_context(
    request: Request,
    user_id: str = "",
    company_id: str = "",
    branch_id: str = "",
    org_unit_id: str = "",
):
    resp = RedirectResponse(url="/ui", status_code=303)

    if user_id:
        resp.set_cookie("ctx_user_id", user_id)

    if company_id:
        resp.set_cookie("ctx_company_id", company_id)
    else:
        resp.delete_cookie("ctx_company_id")

    if branch_id:
        resp.set_cookie("ctx_branch_id", branch_id)
    else:
        resp.delete_cookie("ctx_branch_id")

    if org_unit_id:
        resp.set_cookie("ctx_org_unit_id", org_unit_id)
    else:
        resp.delete_cookie("ctx_org_unit_id")

    return resp


# ─── KPI stats ───────────────────────────────────────────────────────────────

def _build_stats(db: Session, active_ou: Optional[OrgUnit]) -> dict:
    today      = date.today()
    week_start = today - timedelta(days=today.weekday())

    emp_q = db.query(func.count(Employee.id)).filter(Employee.active == True)
    if active_ou:
        emp_q = emp_q.filter(Employee.org_unit_id == active_ou.id)
    total_employees = emp_q.scalar() or 0

    new_this_week = (
        db.query(func.count(Employee.id))
        .filter(Employee.active == True, Employee.created_at >= week_start)
        .scalar() or 0
    )

    total_ous      = db.query(func.count(OrgUnit.id)).scalar() or 0
    total_branches = db.query(func.count(Branch.id)).scalar() or 0

    # Ausentismos activos (fecha_inicio/fin guardadas como string YYYY-MM-DD)
    today_str = today.isoformat()
    ausencias = (
        db.query(AusentismoEmpleado)
        .filter(
            AusentismoEmpleado.fecha_inicio <= today_str,
            AusentismoEmpleado.fecha_fin    >= today_str,
        ).all()
    )
    active_absences = len(ausencias)
    # "Pendientes" = sin detalle (campo detalle vacío)
    pending_absences = sum(1 for a in ausencias if not a.detalle)

    # Alertas: ausentismos sin detalle + runs fallidos recientes
    restriction_alerts = (
        db.query(func.count(AusentismoEmpleado.id))
        .filter(AusentismoEmpleado.detalle == None)
        .scalar() or 0
    )

    # Cobertura: basada en el último run exitoso disponible
    last_ok_run = (
        db.query(Run)
        .filter(Run.status.in_(["success", "ok", "completed"]))
        .order_by(Run.created_at.desc())
        .first()
    )
    total_runs = db.query(func.count(Run.id)).scalar() or 1
    ok_runs    = db.query(func.count(Run.id)).filter(
        Run.status.in_(["success", "ok", "completed"])
    ).scalar() or 0
    coverage_pct = round((ok_runs / total_runs) * 100) if last_ok_run else 0

    return {
        "total_employees":    total_employees,
        "new_employees_week": new_this_week,
        "total_ous":          total_ous,
        "total_branches":     total_branches,
        "active_absences":    active_absences,
        "justified_absences": active_absences - pending_absences,
        "pending_absences":   pending_absences,
        "restriction_alerts": restriction_alerts,
        "coverage_pct":       coverage_pct,
    }


# ─── Cobertura por OU ────────────────────────────────────────────────────────

def _build_ou_coverage(db: Session) -> list[dict]:
    ous = (
        db.query(OrgUnit)
        .join(Branch)
        .join(Company)
        .all()
    )
    result = []
    for ou in ous:
        emp_count = (
            db.query(func.count(Employee.id))
            .filter(Employee.org_unit_id == ou.id, Employee.active == True)
            .scalar() or 0
        )
        # Cobertura: si hay empleados activos asumimos 100%,
        # baja si hay ausentismos activos hoy
        today_str = date.today().isoformat()
        emp_keys = [
            e.employee_key for e in
            db.query(Employee).filter(Employee.org_unit_id == ou.id, Employee.active == True).all()
        ]
        ausentes = 0
        if emp_keys:
            ausentes = (
                db.query(func.count(AusentismoEmpleado.id))
                .filter(
                    AusentismoEmpleado.employee_id.in_(emp_keys),
                    AusentismoEmpleado.fecha_inicio <= today_str,
                    AusentismoEmpleado.fecha_fin    >= today_str,
                ).scalar() or 0
            )
        coverage = round(((emp_count - ausentes) / emp_count) * 100) if emp_count else 0
        result.append({
            "name":           ou.name,
            "branch_name":    ou.branch.name,
            "company_name":   ou.branch.company.name,
            "employee_count": emp_count,
            "coverage":       max(coverage, 0),
        })
    return sorted(result, key=lambda x: x["coverage"])


# ─── Alertas ─────────────────────────────────────────────────────────────────

def _time_ago(dt) -> str:
    """Humaniza una fecha sin dependencias externas."""
    if dt is None:
        return "—"
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt)
        except Exception:
            return dt
    now  = datetime.now(dt.tzinfo)
    diff = now - dt
    s    = int(diff.total_seconds())
    if s < 60:       return "Ahora"
    if s < 3600:     return f"Hace {s // 60} min"
    if s < 86400:    return f"Hace {s // 3600} h"
    if s < 172800:   return "Ayer"
    return dt.strftime("%-d %b")

def _build_alerts(db: Session) -> list[dict]:
    alerts = []

    # Restricciones hard
    hard = (
        db.query(RestriccionEmpleado)
        .filter(RestriccionEmpleado.hard == True)
        .order_by(RestriccionEmpleado.updated_at.desc())
        .limit(2).all()
    )
    for r in hard:
        who = f"Empleado {r.employee_id}" if r.employee_id else "Global"
        alerts.append({
            "type":        "danger",
            "icon":        "⚠️",
            "title":       f"Restricción hard — {who}",
            "description": r.detalle or f"Tipo: {r.tipo}",
            "time_ago":    _time_ago(r.updated_at),
        })

    # Ausentismos sin detalle (pendientes)
    today_str = date.today().isoformat()
    pending = (
        db.query(AusentismoEmpleado)
        .filter(
            AusentismoEmpleado.fecha_inicio <= today_str,
            AusentismoEmpleado.fecha_fin    >= today_str,
            AusentismoEmpleado.detalle      == None,
        )
        .limit(2).all()
    )
    for a in pending:
        alerts.append({
            "type":        "warning",
            "icon":        "🏥",
            "title":       "Ausentismo sin detalle",
            "description": f"Empleado {a.employee_id} · desde {a.fecha_inicio}",
            "time_ago":    _time_ago(a.created_at),
        })

    # Último run completado
    last_ok = (
        db.query(Run)
        .filter(Run.status == "success")
        .order_by(Run.finished_at.desc())
        .first()
    )
    if last_ok:
        alerts.append({
            "type":        "success",
            "icon":        "✅",
            "title":       f"Run completado",
            "description": f"{last_ok.original_filename or 'sin nombre'} · sin errores",
            "time_ago":    _time_ago(last_ok.finished_at),
        })

    # Runs fallidos recientes
    last_fail = (
        db.query(Run)
        .filter(Run.status == "failed")
        .order_by(Run.finished_at.desc())
        .first()
    )
    if last_fail:
        alerts.append({
            "type":        "danger",
            "icon":        "❌",
            "title":       "Run fallido",
            "description": last_fail.error_message or "Revisa el log del run.",
            "time_ago":    _time_ago(last_fail.finished_at),
        })

    return alerts[:6]


# ─── Vista semanal v2: Cobertura por OU (escalable) ───────────────────────────

def _build_weekly_v2(db: Session, active_company: Optional[Company] = None) -> Tuple[list, list]:
    """
    Construye vista semanal agregada por OU mostrando % de cobertura por día.
    Lee desde el último reporte_brechas.xlsx disponible.
    
    Calcula: % de slots sin faltantes por día (no suma de personas).
    
    Retorna:
    - weekly_data: lista de dicts con {ou_name, branch_name, days: [{coverage_pct, status}]}
    - week_days: lista de {label, date_str, is_today}
    """
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    days = [week_start + timedelta(days=i) for i in range(7)]
    DAY_LABELS = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    week_days = [
        {"label": DAY_LABELS[i], "date_str": d.isoformat(), "is_today": d == today}
        for i, d in enumerate(days)
    ]
    
    # Buscar el reporte de brechas más reciente
    best_path = None
    best_mtime = 0
    for pattern in ["/app/storage/runs/*/out/reporte_brechas.xlsx",
                    "/app/storage/runs/*/out/reporte_brechas.csv"]:
        for path in glob.glob(pattern):
            try:
                mtime = Path(path).stat().st_mtime
                if mtime > best_mtime:
                    best_mtime = mtime
                    best_path = path
            except Exception:
                continue
    
    if not best_path:
        return [], week_days
    
    # Leer reporte
    try:
        if best_path.endswith(".xlsx"):
            df = pd.read_excel(best_path)
        else:
            df = pd.read_csv(best_path, encoding="utf-8-sig")
    except Exception as e:
        print(f"[Weekly v2] Error leyendo {best_path}: {e}")
        return [], week_days
    
    if df.empty:
        return [], week_days
    
    # Normalizar columnas
    df.columns = [str(c).lower().strip() for c in df.columns]
    
    # Parsear fechas
    if "fecha" not in df.columns:
        return [], week_days
    
    df["fecha_parsed"] = pd.to_datetime(df["fecha"], errors="coerce").dt.date
    
    # Filtrar solo días de esta semana
    week_dates = set(d for d in days)
    df_week = df[df["fecha_parsed"].isin(week_dates)].copy()
    
    if df_week.empty:
        return [], week_days
    
    # Columnas de requeridos y faltantes
    req_col = None
    falt_col = None
    for c in ["requeridos_personas", "requeridos_min_personas", "required"]:
        if c in df_week.columns:
            req_col = c
            break
    for c in ["faltantes_personas", "faltantes_vs_min_personas", "faltantes"]:
        if c in df_week.columns:
            falt_col = c
            break
    
    if not req_col:
        return [], week_days
    
    # Si no hay columna de faltantes, intentar calcular desde cubiertos
    if not falt_col:
        for c in ["cubiertos_personas", "covered", "assigned"]:
            if c in df_week.columns:
                df_week["_faltantes"] = (df_week[req_col].fillna(0) - df_week[c].fillna(0)).clip(lower=0)
                falt_col = "_faltantes"
                break
    
    if not falt_col:
        return [], week_days
    
    # Normalizar org_unit_id
    df_week["ou_norm"] = df_week["org_unit_id"].astype(str).str.strip().str.upper()
    
    # Marcar slots con demanda > 0
    df_week["tiene_demanda"] = df_week[req_col].fillna(0).astype(float) > 0
    # Marcar slots cubiertos (sin faltantes)
    df_week["cubierto_ok"] = (df_week[falt_col].fillna(0).astype(float) <= 0) & df_week["tiene_demanda"]
    
    # Obtener lista de OUs desde BD para nombres bonitos
    ous_db = db.query(OrgUnit).join(Branch).join(Company).all()
    ou_info = {}
    for ou in ous_db:
        key = str(ou.name).strip().upper()
        ou_info[key] = {
            "name": ou.name,
            "branch_name": ou.branch.name if ou.branch else "",
            "company_id": ou.branch.company_id if ou.branch else None,
            "company_name": ou.branch.company.name if ou.branch and ou.branch.company else "",
        }
    
    # Agregar por OU y fecha: contar slots totales y slots cubiertos
    agg = df_week.groupby(["ou_norm", "fecha_parsed"]).agg({
        "tiene_demanda": "sum",  # Total slots con demanda
        "cubierto_ok": "sum",    # Slots cubiertos OK
    }).reset_index()
    agg.columns = ["ou_norm", "fecha", "total_slots", "slots_ok"]
    
    # Calcular cobertura por OU y día
    ou_list = sorted(agg["ou_norm"].unique())
    
    # Filtrar por empresa si está seleccionada
    if active_company:
        ou_list = [
            ou for ou in ou_list
            if ou in ou_info and ou_info[ou].get("company_id") == active_company.id
        ]
    
    # Calcular cobertura promedio por OU para ordenar
    ou_avg_cov = {}
    for ou in ou_list:
        ou_data = agg[agg["ou_norm"] == ou]
        total_slots = ou_data["total_slots"].sum()
        total_ok = ou_data["slots_ok"].sum()
        ou_avg_cov[ou] = (total_ok / total_slots * 100) if total_slots > 0 else 100
    
    # Ordenar por peor cobertura primero
    ou_list_sorted = sorted(ou_list, key=lambda x: ou_avg_cov.get(x, 100))[:10]
    
    weekly_data = []
    for ou in ou_list_sorted:
        info = ou_info.get(ou, {"name": ou, "branch_name": "", "company_name": ""})
        ou_data = agg[agg["ou_norm"] == ou]
        
        days_data = []
        total_slots_week = 0
        total_ok_week = 0
        
        for d in days:
            day_row = ou_data[ou_data["fecha"] == d]
            if day_row.empty:
                days_data.append({
                    "coverage_pct": None,
                    "status": "no_data",
                })
            else:
                total_slots = int(day_row["total_slots"].iloc[0])
                slots_ok = int(day_row["slots_ok"].iloc[0])
                
                total_slots_week += total_slots
                total_ok_week += slots_ok
                
                if total_slots == 0:
                    pct = None
                    status = "no_demand"
                else:
                    pct = round(slots_ok / total_slots * 100)
                    if pct >= 90:
                        status = "ok"
                    elif pct >= 70:
                        status = "warning"
                    else:
                        status = "critical"
                
                days_data.append({
                    "coverage_pct": pct,
                    "status": status,
                })
        
        # Cobertura promedio de la semana
        avg_pct = round(total_ok_week / total_slots_week * 100) if total_slots_week > 0 else None
        
        weekly_data.append({
            "ou_name": info["name"],
            "branch_name": info["branch_name"],
            "company_name": info["company_name"],
            "days": days_data,
            "avg_coverage": avg_pct,
        })
    
    return weekly_data, week_days


# ─── Vista semanal legacy (mantener por compatibilidad) ─────────────────────

def _build_weekly(db: Session, active_ou: Optional[OrgUnit]):
    today      = date.today()
    week_start = today - timedelta(days=today.weekday())
    days       = [week_start + timedelta(days=i) for i in range(7)]
    DAY_LABELS = ["Lu", "Ma", "Mi", "Ju", "Vi", "Sa", "Do"]
    week_days  = [{"label": DAY_LABELS[i], "is_today": d == today} for i, d in enumerate(days)]

    if not active_ou:
        return [], week_days

    employees = (
        db.query(Employee)
        .filter(Employee.org_unit_id == active_ou.id, Employee.active == True)
        .order_by(Employee.nombre)
        .limit(6).all()
    )
    if not employees:
        return [], week_days

    # Ausentismos de la semana
    emp_keys   = [e.employee_key for e in employees]
    week_end   = days[-1].isoformat()
    week_start_s = days[0].isoformat()
    ausencias  = (
        db.query(AusentismoEmpleado)
        .filter(
            AusentismoEmpleado.employee_id.in_(emp_keys),
            AusentismoEmpleado.fecha_inicio <= week_end,
            AusentismoEmpleado.fecha_fin    >= week_start_s,
        ).all()
    )
    # Indexar por (employee_key, date_str)
    aus_set = set()
    for a in ausencias:
        try:
            fi = date.fromisoformat(a.fecha_inicio)
            ff = date.fromisoformat(a.fecha_fin)
            for d in days:
                if fi <= d <= ff:
                    aus_set.add((a.employee_id, d.isoformat()))
        except Exception:
            pass

    schedule = []
    for emp in employees:
        shifts = []
        for d in days:
            is_today = d == today
            if (emp.employee_key, d.isoformat()) in aus_set:
                shifts.append({"code": "a", "label": "Ausentismo", "is_today": is_today})
            else:
                # Sin plan_mensual.csv disponible en BD, mostramos celda vacía
                shifts.append({"code": "empty", "label": "—", "is_today": is_today})
        schedule.append({"name": emp.nombre, "shifts": shifts})

    return schedule, week_days


# ─── Runs recientes ───────────────────────────────────────────────────────────

def _build_runs(db: Session) -> list[dict]:
    runs = db.query(Run).order_by(Run.created_at.desc()).limit(4).all()
    result = []
    for r in runs:
        if r.status == "success":
            status, label = "ok",    f"Run completado"
            desc = r.original_filename or "—"
        elif r.status in ("queued", "running"):
            status, label = "queue", f"{'En cola' if r.status == 'queued' else 'Ejecutando'}"
            desc = r.original_filename or "—"
        else:
            status, label = "error", "Run fallido"
            desc = r.error_message or "Revisa el log."
        result.append({
            "status":      status,
            "label":       label,
            "description": desc,
            "time_ago":    _time_ago(r.created_at),
        })
    return result


# ─── Colaboradores de hoy ─────────────────────────────────────────────────────

def _build_todays(db: Session, active_ou: Optional[OrgUnit]) -> list[dict]:
    """
    Sin plan_mensual en BD, mostramos los primeros colaboradores activos
    de la OU seleccionada como referencia.
    """
    q = db.query(Employee).filter(Employee.active == True)
    if active_ou:
        q = q.filter(Employee.org_unit_id == active_ou.id)
    employees = q.order_by(Employee.nombre).limit(6).all()

    result = []
    for emp in employees:
        parts    = emp.nombre.split()
        initials = (parts[0][0] + parts[-1][0]).upper() if len(parts) >= 2 else emp.nombre[:2].upper()
        ou_name  = emp.org_unit.name if emp.org_unit else "—"
        result.append({
            "name":       emp.nombre,
            "initials":   initials,
            "ou":         ou_name,
            "time_range": emp.jornada_id or "—",
            "type":       "m",
            "type_label": emp.cargo_id or "—",
        })
    return result



# ─── Curva demanda vs dotación ────────────────────────────────────────────────

def _build_demand_curve(db: Session) -> dict:
    """
    Construye datos para gráfico de demanda vs asignados por día de semana.
    Muestra el pico promedio de personas requeridas y asignadas por tipo de día.
    Lee desde reporte_brechas.csv del run con más cobertura real.
    """
    import csv as csv_mod
    import glob
    from collections import defaultdict

    DAYS = ["LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"]
    DAY_LABELS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

    # Buscar el run con más asignaciones reales
    best_path = None
    best_total = 0

    for path in glob.glob("/app/storage/runs/*/out/reporte_brechas.csv"):
        try:
            with open(path, encoding="utf-8-sig") as f:
                rows = list(csv_mod.DictReader(f))
            total = sum(int(r.get("cubiertos_personas", 0)) for r in rows if r.get("cubiertos_personas", "0").isdigit())
            if total > best_total:
                best_total = total
                best_path = path
        except Exception:
            continue

    if not best_path:
        return {}

    # Por cada (fecha, dia_semana), tomar el MAX de requeridos y cubiertos del día
    fechas_by_day = defaultdict(set)
    req_by_fecha = defaultdict(int)
    cub_by_fecha = defaultdict(int)

    with open(best_path, encoding="utf-8-sig") as f:
        for r in csv_mod.DictReader(f):
            d = (r.get("dia_semana") or "").upper().strip()[:3]
            fecha = r.get("fecha", "")
            if d not in DAYS or not fecha:
                continue
            fechas_by_day[d].add(fecha)
            req_by_fecha[(fecha, d)] = max(
                req_by_fecha[(fecha, d)],
                int(r.get("requeridos_personas", 0) or 0)
            )
            cub_by_fecha[(fecha, d)] = max(
                cub_by_fecha[(fecha, d)],
                int(r.get("cubiertos_personas", 0) or 0)
            )

    demand_list = []
    asignados_list = []
    for d in DAYS:
        fechas = fechas_by_day[d]
        if fechas:
            avg_req = round(sum(req_by_fecha[(f, d)] for f in fechas) / len(fechas), 1)
            avg_cub = round(sum(cub_by_fecha[(f, d)] for f in fechas) / len(fechas), 1)
        else:
            avg_req = 0
            avg_cub = 0
        demand_list.append(avg_req)
        asignados_list.append(avg_cub)

    return {
        "labels":    DAY_LABELS,
        "demand":    demand_list,
        "asignados": asignados_list,
    }


# ─── Cobertura v3: Mínimo vs Ideal ────────────────────────────────────────────

def _build_coverage_summary_v3(db: Session) -> dict:
    """
    Construye resumen de cobertura Mínimo vs Ideal para el Dashboard.
    Lee desde reporte_brechas del último run exitoso.
    
    Columnas esperadas del reporte:
    - requeridos_personas (o requeridos_min_personas)
    - cubiertos_personas
    - faltantes_personas (o faltantes_vs_min_personas)
    - diagnostic (valores: OK, FALTANTE, OVER, etc.)
    
    Retorna dict con: total_slots, sobre_ideal, bajo_ideal, bajo_minimo,
    porcentajes, run_id, run_fecha, tiene_datos
    """
    best_path = None
    best_mtime = 0
    
    # Buscar el reporte más reciente (xlsx o csv)
    for pattern in ["/app/storage/runs/*/out/reporte_brechas.xlsx", 
                    "/app/storage/runs/*/out/reporte_brechas.csv"]:
        for path in glob.glob(pattern):
            try:
                mtime = Path(path).stat().st_mtime
                if mtime > best_mtime:
                    best_mtime = mtime
                    best_path = path
            except Exception:
                continue
    
    if not best_path:
        return {"tiene_datos": False}
    
    # Leer archivo
    rows = []
    try:
        if best_path.endswith(".xlsx"):
            wb = load_workbook(best_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            wb.close()
        else:
            with open(best_path, encoding="utf-8-sig") as f:
                rows = list(csv.DictReader(f))
    except Exception as e:
        print(f"[Dashboard v3] Error leyendo {best_path}: {e}")
        return {"tiene_datos": False}
    
    if not rows:
        return {"tiene_datos": False}
    
    # Contar por diagnóstico
    sobre_ideal = 0  # OK o OVER (cubiertos >= requeridos)
    bajo_ideal = 0   # Ligeramente bajo (faltantes <= 1)
    bajo_minimo = 0  # FALTANTE (faltantes > 0)
    total_slots = 0
    
    for r in rows:
        r_lower = {str(k).lower().strip(): v for k, v in r.items()}
        
        # Obtener requeridos (probar ambos nombres de columna)
        req = int(float(r_lower.get("requeridos_personas", 0) or 
                       r_lower.get("requeridos_min_personas", 0) or 0))
        if req == 0:
            continue
            
        total_slots += 1
        
        # Obtener diagnóstico - probar ambos nombres
        diag = str(r_lower.get("diagnostic", "") or 
                   r_lower.get("diagnostic_ideal", "")).upper().strip()
        
        # Obtener faltantes para clasificación más granular
        faltantes = int(float(r_lower.get("faltantes_personas", 0) or 
                              r_lower.get("faltantes_vs_min_personas", 0) or 0))
        cubiertos = int(float(r_lower.get("cubiertos_personas", 0) or 0))
        
        # Clasificar según diagnostic y faltantes
        if diag in ("OK", "SOBRE_IDEAL", "OVER"):
            sobre_ideal += 1
        elif diag == "BAJO_IDEAL":
            bajo_ideal += 1
        elif diag in ("FALTANTE", "BAJO_MINIMO"):
            # Subclasificar: si faltan pocos es "bajo ideal", si faltan muchos es "bajo mínimo"
            if faltantes == 1:
                bajo_ideal += 1
            else:
                bajo_minimo += 1
        else:
            # Fallback basado en números si no hay diagnostic claro
            if faltantes <= 0:
                sobre_ideal += 1
            elif faltantes == 1:
                bajo_ideal += 1
            else:
                bajo_minimo += 1
    
    if total_slots == 0:
        return {"tiene_datos": False}
    
    pct_sobre = round(sobre_ideal / total_slots * 100, 1)
    pct_bajo_i = round(bajo_ideal / total_slots * 100, 1)
    pct_bajo_m = round(bajo_minimo / total_slots * 100, 1)
    
    run_id = Path(best_path).parent.parent.name
    run_fecha = datetime.fromtimestamp(best_mtime).strftime("%-d %b %Y")
    
    return {
        "tiene_datos": True,
        "total_slots": total_slots,
        "sobre_ideal": sobre_ideal,
        "bajo_ideal": bajo_ideal,
        "bajo_minimo": bajo_minimo,
        "pct_sobre_ideal": pct_sobre,
        "pct_bajo_ideal": pct_bajo_i,
        "pct_bajo_minimo": pct_bajo_m,
        "run_id": run_id,
        "run_fecha": run_fecha,
    }


# -------------------------
# HOME (Dashboard)
# -------------------------
@router.get("/", response_class=HTMLResponse)
@router.get("/home", response_class=HTMLResponse)
def ui_home(request: Request, db: Session = Depends(get_db),
            company_id: str = "", branch_id: str = ""):
    ctx = _load_ctx(request)
    # Allow override via query param for dashboard filter
    if company_id:
        ctx = type(ctx)(user_id=ctx.user_id, company_id=company_id,
                        branch_id=branch_id or ctx.branch_id,
                        org_unit_id=ctx.org_unit_id)
    current_user = get_current_user(db, ctx.user_id)

    # Si no hay usuario, redirige a login
    if not current_user:
        demo = db.query(User).filter(User.is_active == True).order_by(User.created_at.asc()).first()
        if demo:
            resp = RedirectResponse(url="/ui/home", status_code=303)
            resp.set_cookie("ctx_user_id", str(demo.id))
            return resp
        return _redirect("/ui/system/users")

    active_company = db.get(Company, ctx.company_id) if ctx.company_id else None
    active_branch  = db.get(Branch,  ctx.branch_id)  if ctx.branch_id  else None
    active_ou      = db.get(OrgUnit, ctx.org_unit_id) if ctx.org_unit_id else None

    alerts = _build_alerts(db)
    weekly_data, week_days = _build_weekly_v2(db, active_company)
    
    return TEMPLATES.TemplateResponse("home.html", {
        "request":        request,
        "active_nav":     "home",
        "current_user":   current_user,
        "active_company": active_company,
        "active_branch":  active_branch,
        "active_ou":      active_ou,
        "alerts_count":   len(alerts),
        "stats":          _build_stats(db, active_ou),
        "ou_coverage":    _build_ou_coverage(db),
        "alerts":         alerts,
        "weekly_data":    weekly_data,
        "week_days":      week_days,
        "recent_runs":    _build_runs(db),
        "todays_shifts":  _build_todays(db, active_ou),
        "now":            datetime.now(),
        "fecha_hoy_es":   _fecha_es(datetime.now()),
        "demand_curve":   _build_demand_curve(db),
        "coverage_v3":    _build_coverage_summary_v3(db),
    })


# =====================================================
# Sidebar legacy aliases (para no romper tu menu actual)
# =====================================================
@router.get("/app/colaboradores")
def alias_app_colaboradores():
    return _redirect("/ui/colaboradores")


@router.get("/app/demanda")
def alias_app_demanda():
    return _redirect("/ui/demanda")


@router.get("/app/pool")
def alias_app_pool():
    return _redirect("/ui/pool")


@router.get("/app/restricciones")
def alias_app_restricciones():
    return _redirect("/ui/restricciones")


@router.get("/app/ausentismos")
def alias_app_ausentismos():
    return _redirect("/ui/ausentismos")


@router.get("/app/turnos")
def alias_app_turnos():
    return _redirect("/ui/runs")


# -------------------------
# Storage helpers
# -------------------------
def _company_dir(company_id: uuid.UUID) -> Path:
    p = Path("/app/storage/companies") / str(company_id)
    p.mkdir(parents=True, exist_ok=True)
    return p


def _company_template_path(company_id: uuid.UUID) -> Path:
    return _company_dir(company_id) / "case_template.xlsx"

def _template_backup(path: Path) -> None:
    if not path.exists():
        return
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    bak = path.parent / f"{path.stem}.bak_{ts}{path.suffix}"
    try:
        shutil.copyfile(path, bak)
    except Exception:
        # Si falla el backup, no detenemos (MVP), pero lo ideal es loggear.
        pass


def _template_read_rows(template_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Lee una hoja del template como lista de dict usando fila 1 como headers."""
    if not template_path.exists():
        return []
    wb = load_workbook(filename=str(template_path), data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    if ws.max_row < 2:
        return []
    headers: list[str] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        headers.append(str(v).strip() if v is not None else "")
    out: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row: dict[str, Any] = {}
        empty = True
        for i, h in enumerate(headers, start=1):
            if not h:
                continue
            v = ws.cell(row=r, column=i).value
            if v is not None and str(v).strip() != "":
                empty = False
            row[h] = v
        if not empty:
            out.append(row)
    return out


def _norm_str(v: Any) -> str:
    return (str(v).strip() if v is not None else "")


def _template_append_row(
    template_path: Path,
    sheet_name: str,
    key_col: str,
    payload: dict[str, Any],
) -> None:
    """Agrega una fila al template (hoja sheet_name) si el key_col no existe."""
    if not template_path.exists():
        raise ValueError("Template no existe")
    wb = load_workbook(filename=str(template_path))
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"El template no tiene la hoja {sheet_name}")
    ws = wb[sheet_name]

    # headers
    headers: list[str] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        headers.append(str(v).strip() if v is not None else "")
    idx = {h: i for i, h in enumerate(headers, start=1) if h}

    if key_col not in idx:
        raise ValueError(f"Hoja {sheet_name} no tiene columna {key_col}")

    key_val = _norm_str(payload.get(key_col))
    if not key_val:
        raise ValueError(f"{key_col} es requerido")

    # detectar duplicado (case-insensitive)
    key_i = idx[key_col]
    existing = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=key_i).value
        s = _norm_str(v).lower()
        if s:
            existing.add(s)
    if key_val.lower() in existing:
        raise ValueError(f"Ya existe {key_col}={key_val} en {sheet_name}")

    # construir row en el orden del header
    row_values: list[Any] = []
    for h in headers:
        if not h:
            row_values.append(None)
            continue
        row_values.append(payload.get(h))
    ws.append(row_values)

    _template_backup(template_path)
    wb.save(str(template_path))


def _company_vocab_template_name() -> str:
    tpl_dir = Path(__file__).parent / "templates"
    if (tpl_dir / "company_vocab_new.html").exists():
        return "company_vocab_new.html"
    return "company_vocab.html"

def _collect_artifacts(run: Run) -> List[Dict[str, Any]]:
    out_dir = Path(run.out_dir)
    log_path = Path(run.log_path)
    case_path = Path(run.case_path)

    artifacts: List[Dict[str, Any]] = []

    if case_path.exists() and case_path.is_file():
        artifacts.append(
            {
                "name": "input/case.xlsx",
                "size_bytes": case_path.stat().st_size,
                "modified_at": file_mtime_utc(case_path),
                "download_url": f"/ui/runs/{run.id}/input-case",
            }
        )

    if out_dir.exists():
        for p in sorted(out_dir.glob("**/*")):
            if p.is_file():
                rel = p.relative_to(out_dir).as_posix()
                artifacts.append(
                    {
                        "name": rel,
                        "size_bytes": p.stat().st_size,
                        "modified_at": file_mtime_utc(p),
                        "download_url": f"/runs/{run.id}/artifacts/{rel}",
                    }
                )

    if log_path.exists() and log_path.is_file():
        artifacts.append(
            {
                "name": "logs/solver.log",
                "size_bytes": log_path.stat().st_size,
                "modified_at": file_mtime_utc(log_path),
                "download_url": f"/runs/{run.id}/artifacts/logs/solver.log",
                    }
        )

    return artifacts


def _read_shift_ids_from_template(template_path: Path) -> List[str]:
    """
    Lee CatalogoTurnos.shift_id desde /app/storage/companies/<company>/case_template.xlsx
    Devuelve lista ordenada, sin duplicados.
    """
    if not template_path.exists():
        return []
    try:
        wb = load_workbook(filename=str(template_path), data_only=True, read_only=True)
        if "CatalogoTurnos" not in wb.sheetnames:
            return []
        ws = wb["CatalogoTurnos"]

        # buscar columna "shift_id" en la fila 1
        headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            headers.append(str(v).strip() if v is not None else "")
        if "shift_id" not in headers:
            return []
        idx = headers.index("shift_id") + 1

        out = []
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=idx).value
            if v is None:
                continue
            s = str(v).strip()
            if s:
                out.append(s)
        # únicos manteniendo orden
        seen = set()
        uniq = []
        for s in out:
            if s not in seen:
                seen.add(s)
                uniq.append(s)
        return uniq
    except Exception:
        return []


def _normalize_dow(d: str) -> str:
    d = (d or "").strip().upper()
    mapping = {
        "LUNES": "LUN",
        "MARTES": "MAR",
        "MIERCOLES": "MIE",
        "MIÉRCOLES": "MIE",
        "JUEVES": "JUE",
        "VIERNES": "VIE",
        "SABADO": "SAB",
        "SÁBADO": "SAB",
        "DOMINGO": "DOM",
    }
    if d in mapping:
        return mapping[d]
    return d

def _count_sundays_in_month(year: int, month: int) -> int:
    cal = calendar.monthcalendar(year, month)
    # En Python: lunes=0 ... domingo=6
    return sum(1 for week in cal if week[6] != 0)

def _monday_of_week_containing(d: date) -> date:
    return d.replace(day=d.day) - (d.weekday() * timedelta(days=1))


# =========================
# SISTEMA -> Usuarios (RBAC)
# =========================
def _chk(v: Optional[str]) -> bool:
    return bool(v)


@router.get("/system/users", response_class=HTMLResponse)
def ui_users(request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    users = db.query(User).order_by(User.created_at.desc()).all()
    return TEMPLATES.TemplateResponse(
        "users.html",
        _enrich(db, request, {"request": request, "users": users, "current_user": current_user}),
    )

@router.get("/turnos", response_class=HTMLResponse)
def ui_turnos(request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    companies = filter_companies(db, current_user)

    # Selectores en cascada
    q_company_id = _parse_uuid(request.query_params.get("company_id"))
    q_branch_id  = _parse_uuid(request.query_params.get("branch_id"))

    branches: list[Branch] = []
    org_units: list[OrgUnit] = []
    if q_company_id:
        branches = db.query(Branch).filter(Branch.company_id == q_company_id).order_by(Branch.code.asc()).all()
    if q_branch_id:
        org_units = db.query(OrgUnit).filter(OrgUnit.branch_id == q_branch_id).order_by(OrgUnit.name.asc()).all()

    now = datetime.now(timezone.utc)
    default_month = f"{now.year:04d}-{now.month:02d}"

    return TEMPLATES.TemplateResponse(
        "turnos.html",
        _enrich(
            db,
            request,
            {
                "request": request,
                "companies": companies,
                "branches": branches,
                "org_units": org_units,
                "selected_company_id": str(q_company_id) if q_company_id else "",
                "selected_branch_id":  str(q_branch_id)  if q_branch_id  else "",
                "default_month": default_month,
                "err": request.query_params.get("err"),
            },
        ),
    )


@router.post("/turnos/request")
def ui_turnos_request(
    request: Request,
    company_id: str = Form(...),
    month: str = Form(...),  # YYYY-MM
    branch_id: str = Form(""),
    org_unit_id: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    cid = _parse_uuid(company_id)
    if not cid:
        return _redirect("/ui/turnos?err=Empresa+inválida")

    bid  = _parse_uuid(branch_id)
    ouid = _parse_uuid(org_unit_id)

    # Validar mes YYYY-MM
    try:
        y, m = month.strip().split("-")
        year = int(y)
        mon = int(m)
        if mon < 1 or mon > 12:
            raise ValueError()
    except Exception:
        return _redirect("/ui/turnos?err=Mes+inválido+(usa+YYYY-MM)")

    # Semanas = número de domingos del mes (tu regla)
    semanas = _count_sundays_in_month(year, mon)

    # Construir case desde DB y lanzar Run
    run_id = uuid.uuid4()
    paths = get_run_paths(run_id)
    ensure_dirs(paths)

    # 1) Construir case.xlsx en paths.case_path
    template_path = _company_template_path(cid)
    if not template_path.exists():
        return _redirect("/ui/turnos?err=Primero+sube+el+template+base+del+solver+en+Empresas")
    try:
        build_case_from_db(
            db=db,
            company_id=cid,
            month=f"{year:04d}-{mon:02d}",
            template_path=template_path,
            out_case_path=paths.case_path,
            branch_id=bid,
            org_unit_id=ouid,
        )
    except Exception as e:
        return _redirect(f"/ui/turnos?err=No+se+pudo+generar+case:+{e}")

    # 2) Persistir Run en DB (igual que /runs POST)
    run = Run(
        id=run_id,
        status="queued",
        original_filename=f"case_{year:04d}-{mon:02d}.xlsx",
        case_path=str(paths.case_path),
        out_dir=str(paths.out_dir),
        log_path=str(paths.log_path),
    )
    db.add(run)
    db.commit()

    # 3) Encolar worker
    celery_app.send_task('execute_run', args=[str(run_id)])

    # 4) Ir al detalle del run
    return _redirect(f"/ui/runs/{run_id}")

@router.post("/system/users")
def ui_users_create(
    email: str = Form(...),
    full_name: str = Form(...),
    role: str = Form(...),
    can_manage_catalogs: Optional[str] = Form(None),
    can_manage_companies: Optional[str] = Form(None),
    can_edit_employees: Optional[str] = Form(None),
    can_edit_demand: Optional[str] = Form(None),
    can_edit_pool: Optional[str] = Form(None),
    can_edit_restrictions: Optional[str] = Form(None),
    can_edit_absences: Optional[str] = Form(None),
    can_request_turnos: Optional[str] = Form(None),
    can_edit_turnos: Optional[str] = Form(None),
    can_view_all_runs: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    email = email.strip().lower()
    full_name = full_name.strip()

    exists = db.query(User).filter(User.email == email).first()
    if exists:
        return _redirect("/ui/system/users")

    u = User(
        email=email,
        full_name=full_name,
        role=role.strip(),
        can_manage_catalogs=_chk(can_manage_catalogs),
        can_manage_companies=_chk(can_manage_companies),
        can_edit_employees=bool(can_edit_employees),
        can_edit_demand=bool(can_edit_demand),
        can_edit_pool=bool(can_edit_pool),
        can_edit_restrictions=bool(can_edit_restrictions),
        can_edit_absences=bool(can_edit_absences),
        can_request_turnos=bool(can_request_turnos),
        can_edit_turnos=_chk(can_edit_turnos),
        can_view_all_runs=_chk(can_view_all_runs),
    )
    db.add(u)
    db.commit()
    return _redirect("/ui/system/users")


@router.get("/system/users/{user_id}", response_class=HTMLResponse)
def ui_user_detail(user_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    u = db.get(User, user_id)
    if not u:
        raise HTTPException(status_code=404, detail="User not found")

    scopes = db.query(UserScope).filter(UserScope.user_id == user_id).order_by(UserScope.created_at.desc()).all()
    companies = db.query(Company).order_by(Company.name.asc()).all()
    branches = db.query(Branch).order_by(Branch.code.asc()).all()
    org_units = db.query(OrgUnit).order_by(OrgUnit.org_unit_key.asc()).all()

    return TEMPLATES.TemplateResponse(
        "user_detail.html",
        _enrich(
            db,
            request,
            {
                "request": request,
                "user": u,
                "scopes": scopes,
                "companies": companies,
                "branches": branches,
                "org_units": org_units,
                "err": request.query_params.get("err"),
            },
        ),
    )


@router.post("/system/users/{user_id}/impersonate")
def ui_impersonate(user_id: uuid.UUID):
    resp = RedirectResponse(url="/ui/companies", status_code=303)
    resp.set_cookie("ctx_user_id", str(user_id))
    return resp


@router.post("/system/users/clear-impersonation")
def ui_clear_impersonation():
    resp = RedirectResponse(url="/ui/system/users", status_code=303)
    resp.delete_cookie("ctx_user_id")
    return resp


# =========================
# EMPRESAS
# =========================
@router.get("/companies", response_class=HTMLResponse)
def ui_companies(request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")
    companies = filter_companies(db, current_user)
    return TEMPLATES.TemplateResponse("companies.html", _enrich(db, request, {"request": request, "companies": companies}))


@router.get("/companies/{company_id}", response_class=HTMLResponse)
def ui_company_detail(company_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")
    if not can_see_company(db, current_user, company_id):
        return _redirect("/ui/companies?err=Sin+acceso+a+empresa")

    c = db.get(Company, company_id)
    if not c:
        raise HTTPException(status_code=404, detail="Company not found")

    branches = filter_branches(db, current_user, company_id)

    template_path = _company_template_path(company_id)
    template_exists = template_path.exists()
    seed_company_vocab(db, company_id, template_path if template_exists else None)

    all_org_units = filter_org_units(db, current_user, None, company_id)

    # Wizard counts (informativo)
    branch_count = len(branches)
    ou_count = len(all_org_units)
    emp_count = (
        db.query(Employee).filter(Employee.org_unit_id.in_([ou.id for ou in all_org_units])).count()
        if all_org_units
        else 0
    )
    demand_count = (
        db.query(DemandUnit).filter(DemandUnit.org_unit_id.in_([ou.id for ou in all_org_units]), DemandUnit.active == True).count()  # noqa: E712
        if all_org_units
        else 0
    )
    pool_count = (
        db.query(PoolTurno).filter(PoolTurno.org_unit_id.in_([ou.id for ou in all_org_units]), PoolTurno.habilitado == True).count()  # noqa: E712
        if all_org_units
        else 0
    )

    steps = {
        "template": template_exists,
        "branches": branch_count > 0,
        "org_units": ou_count > 0,
        "employees": emp_count > 0,
        "demand": demand_count > 0,
        "pool": pool_count > 0,
    }
    ready_to_run = all(steps.values())

    return TEMPLATES.TemplateResponse(
        "company_detail.html",
        _enrich(
            db,
            request,
            {
                "request": request,
                "company": c,
                "branches": branches,
                "template_exists": template_exists,
                "steps": steps,
                "ready_to_run": ready_to_run,
                "all_org_units": all_org_units,
                "err": request.query_params.get("err"),
            },
        ),
    )


# =========================
# SUCURSALES
# =========================

@router.post("/companies/{company_id}/branches")
def ui_branch_create(
    company_id: uuid.UUID,
    request: Request,
    code: str = Form(...),
    name: str = Form(...),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")
    if not can_see_company(db, current_user, company_id):
        raise HTTPException(403, "Forbidden")

    code = code.strip()
    name = name.strip()
    if not code or not name:
        return _redirect(f"/ui/companies/{company_id}?err=Código+y+nombre+son+requeridos")

    exists = db.query(Branch).filter(Branch.company_id == company_id, Branch.code == code).first()
    if exists:
        return _redirect(f"/ui/companies/{company_id}?err=Ya+existe+una+sucursal+con+código+{code}")

    db.add(Branch(company_id=company_id, code=code, name=name))
    db.commit()
    return _redirect(f"/ui/companies/{company_id}?ok=Sucursal+creada")


@router.get("/branches/{branch_id}", response_class=HTMLResponse)
def ui_branch_detail(branch_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    branch = db.get(Branch, branch_id)
    if not branch:
        raise HTTPException(404, "Sucursal no encontrada")

    if not can_see_company(db, current_user, branch.company_id):
        raise HTTPException(403, "Forbidden")

    org_units = (
        db.query(OrgUnit)
        .filter(OrgUnit.branch_id == branch_id)
        .order_by(OrgUnit.org_unit_key.asc())
        .all()
    )

    return TEMPLATES.TemplateResponse(
        "branch_detail.html",
        _enrich(db, request, {
            "request": request,
            "branch": branch,
            "company_id": str(branch.company_id),
            "org_units": org_units,
            "ok": request.query_params.get("ok"),
            "err": request.query_params.get("err"),
        }),
    )


@router.post("/branches/{branch_id}/settings")
def ui_branch_settings(
    branch_id: uuid.UUID,
    request: Request,
    company_id: str = Form(""),
    opera_en_feriados: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    branch = db.get(Branch, branch_id)
    if not branch:
        raise HTTPException(404, "Sucursal no encontrada")

    if not can_see_company(db, current_user, branch.company_id):
        raise HTTPException(403, "Forbidden")

    branch.opera_en_feriados = bool(opera_en_feriados)
    db.commit()

    return _redirect(f"/ui/branches/{branch_id}?ok=Configuración+guardada")


@router.post("/branches/{branch_id}/org-units")
def ui_org_unit_create(
    branch_id: uuid.UUID,
    request: Request,
    org_unit_key: str = Form(...),
    name: str = Form(...),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    branch = db.get(Branch, branch_id)
    if not branch:
        raise HTTPException(404, "Sucursal no encontrada")

    if not can_see_company(db, current_user, branch.company_id):
        raise HTTPException(403, "Forbidden")

    org_unit_key = org_unit_key.strip()
    name = name.strip()
    if not org_unit_key or not name:
        return _redirect(f"/ui/branches/{branch_id}?err=org_unit_key+y+nombre+son+requeridos")

    exists = db.query(OrgUnit).filter(
        OrgUnit.branch_id == branch_id,
        OrgUnit.org_unit_key == org_unit_key
    ).first()
    if exists:
        return _redirect(f"/ui/branches/{branch_id}?err=Ya+existe+una+OU+con+key+{org_unit_key}")

    db.add(OrgUnit(branch_id=branch_id, org_unit_key=org_unit_key, name=name))
    db.commit()
    return _redirect(f"/ui/branches/{branch_id}?ok=OU+creada")


# =========================
# ORG UNITS (FICHA COMPLETA)
# =========================
@router.get("/org-units/{org_unit_id}", response_class=HTMLResponse)
def ui_org_unit_detail(org_unit_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    org_unit = db.get(OrgUnit, org_unit_id)
    if not org_unit:
        raise HTTPException(404, "OU no encontrada")

    branch = db.get(Branch, org_unit.branch_id)
    if not branch:
        raise HTTPException(404, "Sucursal no encontrada")

    if not can_see_company(db, current_user, branch.company_id):
        raise HTTPException(403, "Forbidden")

    company_id = branch.company_id
    template_path = _company_template_path(company_id)
    seed_company_vocab(db, company_id, template_path if template_path.exists() else None)
    vocab = list_vocab(db, company_id)

    employees = (
        db.query(Employee)
        .filter(Employee.org_unit_id == org_unit_id)
        .order_by(Employee.nombre.asc())
        .all()
    )

    demand = (
        db.query(DemandUnit)
        .filter(DemandUnit.org_unit_id == org_unit_id)
        .order_by(DemandUnit.dia_semana.asc(), DemandUnit.inicio.asc())
        .all()
    )

    pool_count = (
        db.query(PoolTurno)
        .filter(PoolTurno.org_unit_id == org_unit_id, PoolTurno.habilitado == True)  # noqa: E712
        .count()
    )

    cargos = sorted({e.cargo_id for e in employees if e.cargo_id})
    emp_keys = [e.employee_key for e in employees]

    restricciones = (
        db.query(RestriccionEmpleado)
        .filter(RestriccionEmpleado.company_id == company_id)
        .order_by(RestriccionEmpleado.created_at.desc())
        .limit(100)
        .all()
    )

    ausentismos = (
        db.query(AusentismoEmpleado)
        .filter(
            AusentismoEmpleado.company_id == company_id,
            AusentismoEmpleado.employee_id.in_(emp_keys),
        )
        .order_by(AusentismoEmpleado.created_at.desc())
        .limit(100)
        .all()
    ) if emp_keys else []

    # Construir jornada_cap desde el template Excel
    jornadas_rows = _template_read_rows(template_path, "Jornadas") if template_path.exists() else []
    jornada_cap: dict[str, int] = {}
    for row in jornadas_rows:
        jid = str(row.get("jornada_id") or row.get("id") or "").strip()
        cap = row.get("cap_min_semana") or row.get("contrato_max_min_semana") or 0
        if jid:
            try:
                jornada_cap[jid] = int(cap)
            except (ValueError, TypeError):
                jornada_cap[jid] = 0

    return TEMPLATES.TemplateResponse(
        "orgunit_detail.html",
        _enrich(db, request, {
            "request": request,
            "org_unit": org_unit,
            "branch": branch,
            "company_id": str(company_id),
            "employees": employees,
            "emp_count": len(employees),
            "demand": demand,
            "pool_count": pool_count,
            "cargos": cargos,
            "jornadas": vocab.get(CAT_JORNADA, []),
            "jornada_cap": jornada_cap,
            "restr_tipos": sorted(set([i.value for i in vocab.get(CAT_RESTR_TIPO, [])] + RESTR_TIPOS_FALLBACK)),
            "aus_codes": sorted(set([i.value for i in vocab.get(CAT_AUSENTISMO, [])] + AUS_CODES_FALLBACK)),
            "restricciones": restricciones,
            "ausentismos": ausentismos,
            "ok": request.query_params.get("ok"),
            "err": request.query_params.get("err"),
        }),
    )


@router.post("/org-units/{org_unit_id}/employees")
def ui_org_unit_add_employee(
    org_unit_id: uuid.UUID,
    request: Request,
    employee_key: str = Form(...),
    nombre: str = Form(...),
    cargo_id: str = Form(...),
    jornada_id: str = Form(...),
    contrato_max_min_semana: int = Form(2640),
    expertise: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    org_unit = db.get(OrgUnit, org_unit_id)
    if not org_unit:
        raise HTTPException(404, "OU no encontrada")
    branch = db.get(Branch, org_unit.branch_id)
    if not branch or not can_see_company(db, current_user, branch.company_id):
        raise HTTPException(403, "Forbidden")

    employee_key = employee_key.strip()
    nombre = nombre.strip()

    exists = db.query(Employee).filter(
        Employee.org_unit_id == org_unit_id,
        Employee.employee_key == employee_key,
    ).first()
    if exists:
        return _redirect(f"/ui/org-units/{org_unit_id}?err=Ya+existe+un+colaborador+con+key+{employee_key}")

    emp = Employee(
        org_unit_id=org_unit_id,
        employee_key=employee_key,
        nombre=nombre,
        cargo_id=cargo_id.strip(),
        jornada_id=jornada_id.strip(),
    )
    if hasattr(emp, "contrato_max_min_semana"):
        emp.contrato_max_min_semana = contrato_max_min_semana
    if hasattr(emp, "expertise") and expertise:
        emp.expertise = expertise.strip()

    db.add(emp)
    db.commit()
    return _redirect(f"/ui/org-units/{org_unit_id}?ok=Colaborador+creado")



@router.post("/org-units/{org_unit_id}/employees/{employee_id}/edit")
def ui_org_unit_edit_employee(
    org_unit_id: uuid.UUID,
    employee_id: uuid.UUID,
    request: Request,
    nombre: str = Form(...),
    cargo_id: str = Form(...),
    jornada_id: str = Form(...),
    contrato_max_min_semana: int = Form(2640),
    expertise: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    emp = db.get(Employee, employee_id)
    if not emp:
        raise HTTPException(404, "Colaborador no encontrado")

    org_unit = db.get(OrgUnit, org_unit_id)
    if not org_unit:
        raise HTTPException(404, "OU no encontrada")
    branch = db.get(Branch, org_unit.branch_id)
    if not branch or not can_see_company(db, current_user, branch.company_id):
        raise HTTPException(403, "Forbidden")

    emp.nombre = nombre.strip()
    emp.cargo_id = cargo_id.strip()
    emp.jornada_id = jornada_id.strip()
    if hasattr(emp, "contrato_max_min_semana"):
        emp.contrato_max_min_semana = contrato_max_min_semana
    if hasattr(emp, "expertise"):
        emp.expertise = expertise.strip()

    db.commit()
    return _redirect(f"/ui/org-units/{org_unit_id}?ok=Colaborador+actualizado#empleados")

@router.post("/org-units/{org_unit_id}/demand")
def ui_org_unit_add_demand(
    org_unit_id: uuid.UUID,
    request: Request,
    dia_semana: str = Form(...),
    inicio: str = Form(...),
    fin: str = Form(...),
    requeridos: int = Form(...),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    org_unit = db.get(OrgUnit, org_unit_id)
    if not org_unit:
        raise HTTPException(404, "OU no encontrada")
    branch = db.get(Branch, org_unit.branch_id)
    if not branch or not can_see_company(db, current_user, branch.company_id):
        raise HTTPException(403, "Forbidden")

    # Upsert: actualiza si ya existe el tramo (dia+inicio+fin)
    existing = db.query(DemandUnit).filter(
        DemandUnit.org_unit_id == org_unit_id,
        DemandUnit.dia_semana == dia_semana.strip().upper(),
        DemandUnit.inicio == inicio.strip(),
        DemandUnit.fin == fin.strip(),
    ).first()

    if existing:
        existing.requeridos = requeridos
        existing.active = True
    else:
        db.add(DemandUnit(
            org_unit_id=org_unit_id,
            dia_semana=dia_semana.strip().upper(),
            inicio=inicio.strip(),
            fin=fin.strip(),
            requeridos=requeridos,
            active=True,
        ))
    db.commit()
    return _redirect(f"/ui/org-units/{org_unit_id}?ok=Demanda+guardada#demanda")


@router.post("/org-units/{org_unit_id}/pool-expand")
def ui_org_unit_pool_expand(
    org_unit_id: uuid.UUID,
    request: Request,
    cargo_id: str = Form(...),
    dias_semana: List[str] = Form(...),
    shift_ids_text: str = Form(...),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    org_unit = db.get(OrgUnit, org_unit_id)
    if not org_unit:
        raise HTTPException(404, "OU no encontrada")
    branch = db.get(Branch, org_unit.branch_id)
    if not branch or not can_see_company(db, current_user, branch.company_id):
        raise HTTPException(403, "Forbidden")

    cargo_id = cargo_id.strip()
    dias = [_normalize_dow(d) for d in dias_semana if d.strip()]

    added = 0
    for line in shift_ids_text.splitlines():
        shift_id = line.strip()
        if not shift_id:
            continue
        for dia in dias:
            exists = db.query(PoolTurno).filter(
                PoolTurno.org_unit_id == org_unit_id,
                PoolTurno.cargo_id == cargo_id,
                PoolTurno.dia_semana == dia,
                PoolTurno.shift_id == shift_id,
            ).first()
            if not exists:
                db.add(PoolTurno(
                    org_unit_id=org_unit_id,
                    cargo_id=cargo_id,
                    cargo=cargo_id,
                    dia_semana=dia,
                    shift_id=shift_id,
                    habilitado=1,
                ))
                added += 1

    db.commit()
    return _redirect(f"/ui/org-units/{org_unit_id}?ok={added}+filas+de+pool+agregadas#pool")


@router.post("/org-units/{org_unit_id}/restricciones")
def ui_org_unit_add_restriccion(
    org_unit_id: uuid.UUID,
    request: Request,
    employee_id: str = Form("GLOBAL"),
    tipo: str = Form(...),
    valor1: str = Form(""),
    valor2: str = Form(""),
    dia_semana: str = Form(""),
    fecha: str = Form(""),
    hard: int = Form(0),
    penalizacion: int = Form(0),
    detalle: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    org_unit = db.get(OrgUnit, org_unit_id)
    if not org_unit:
        raise HTTPException(404, "OU no encontrada")
    branch = db.get(Branch, org_unit.branch_id)
    if not branch or not can_see_company(db, current_user, branch.company_id):
        raise HTTPException(403, "Forbidden")

    emp_id = None if employee_id == "GLOBAL" else employee_id.strip()

    db.add(RestriccionEmpleado(
        company_id=branch.company_id,
        employee_id=emp_id,
        tipo=tipo.strip(),
        valor1=valor1.strip() if valor1 else None,
        valor2=valor2.strip() if valor2 else None,
        dia_semana=dia_semana.strip().upper() if dia_semana else None,
        fecha=fecha.strip() if fecha else None,
        hard=int(hard),
        penalizacion=int(penalizacion),
        detalle=detalle.strip() if detalle else None,
    ))
    db.commit()
    return _redirect(f"/ui/org-units/{org_unit_id}?ok=Restricción+guardada#restricciones")


@router.post("/org-units/{org_unit_id}/ausentismos")
def ui_org_unit_add_ausentismo(
    org_unit_id: uuid.UUID,
    request: Request,
    employee_id: str = Form(...),
    fecha_inicio: str = Form(...),
    fecha_fin: str = Form(""),
    ausentismo: str = Form(...),
    detalle: str = Form(""),
    hard: int = Form(1),
    penalizacion: int = Form(0),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    org_unit = db.get(OrgUnit, org_unit_id)
    if not org_unit:
        raise HTTPException(404, "OU no encontrada")
    branch = db.get(Branch, org_unit.branch_id)
    if not branch or not can_see_company(db, current_user, branch.company_id):
        raise HTTPException(403, "Forbidden")

    db.add(AusentismoEmpleado(
        company_id=branch.company_id,
        employee_id=employee_id.strip(),
        fecha_inicio=fecha_inicio.strip(),
        fecha_fin=fecha_fin.strip() if fecha_fin else None,
        ausentismo=ausentismo.strip(),
        detalle=detalle.strip() if detalle else None,
        hard=int(hard),
        penalizacion=int(penalizacion),
    ))
    db.commit()
    return _redirect(f"/ui/org-units/{org_unit_id}?ok=Ausentismo+guardado#ausentismos")

@router.get("/colaboradores", response_class=HTMLResponse)
def ui_colaboradores(request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    q_company_id = _parse_uuid(request.query_params.get("company_id"))
    q_branch_id = _parse_uuid(request.query_params.get("branch_id"))
    q_org_unit_id = _parse_uuid(request.query_params.get("org_unit_id"))
    q_search = (request.query_params.get("q") or "").strip()

    companies = filter_companies(db, current_user)
    branches: List[Branch] = []
    org_units: List[OrgUnit] = []

    if q_company_id:
        branches = filter_branches(db, current_user, q_company_id)
    if q_branch_id and q_company_id:
        org_units = filter_org_units(db, current_user, q_branch_id, q_company_id)
    elif q_company_id:
        org_units = filter_org_units(db, current_user, None, q_company_id)

    emp_q = (
        db.query(Employee)
        .join(OrgUnit, Employee.org_unit_id == OrgUnit.id)
        .join(Branch, OrgUnit.branch_id == Branch.id)
        .join(Company, Branch.company_id == Company.id)
    )

    allowed_company_ids = [c.id for c in companies]
    if allowed_company_ids:
        emp_q = emp_q.filter(Company.id.in_(allowed_company_ids))
    else:
        emp_q = emp_q.filter(False)

    if q_company_id:
        emp_q = emp_q.filter(Company.id == q_company_id)
    if q_branch_id:
        emp_q = emp_q.filter(Branch.id == q_branch_id)
    if q_org_unit_id:
        emp_q = emp_q.filter(OrgUnit.id == q_org_unit_id)

    if q_search:
        like = f"%{q_search}%"
        emp_q = emp_q.filter((Employee.nombre.ilike(like)) | (Employee.employee_key.ilike(like)))

    employees = emp_q.order_by(Employee.created_at.desc()).limit(200).all()

    ou_map = {ou.id: ou for ou in db.query(OrgUnit).all()}
    br_map = {b.id: b for b in db.query(Branch).all()}
    comp_map = {c.id: c for c in db.query(Company).all()}

    return TEMPLATES.TemplateResponse(
        "collaborators.html",
        _enrich(
            db,
            request,
            {
                "request": request,
                "companies": companies,
                "branches": branches,
                "org_units": org_units,
                "selected_company_id": str(q_company_id) if q_company_id else "",
                "selected_branch_id": str(q_branch_id) if q_branch_id else "",
                "selected_org_unit_id": str(q_org_unit_id) if q_org_unit_id else "",
                "q": q_search,
                "employees": employees,
                "ou_map": ou_map,
                "br_map": br_map,
                "comp_map": comp_map,
            },
        ),
    )


@router.post("/colaboradores/add")
def ui_colaboradores_add(
    request: Request,
    company_id: str = Form(...),
    branch_id: str = Form(...),
    org_unit_id: str = Form(...),
    employee_key: str = Form(...),
    nombre: str = Form(...),
    cargo_id: str = Form(...),
    jornada_id: str = Form(...),
    contrato_max_min_semana: int = Form(2640),
    email: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    ouid = _parse_uuid(org_unit_id)
    if not ouid:
        return _redirect(
            f"/ui/colaboradores?company_id={company_id}&branch_id={branch_id}"
            f"&org_unit_id={org_unit_id}&err=OU+inválida"
        )

    org_unit = db.get(OrgUnit, ouid)
    if not org_unit:
        return _redirect(
            f"/ui/colaboradores?company_id={company_id}&branch_id={branch_id}&err=OU+no+encontrada"
        )

    branch = db.get(Branch, org_unit.branch_id)
    if not branch or not can_see_company(db, current_user, branch.company_id):
        raise HTTPException(403, "Forbidden")

    employee_key = employee_key.strip()
    nombre = nombre.strip()

    if not employee_key or not nombre:
        return _redirect(
            f"/ui/colaboradores?company_id={company_id}&branch_id={branch_id}"
            f"&org_unit_id={org_unit_id}&err=RUT+y+nombre+son+requeridos"
        )

    exists = db.query(Employee).filter(
        Employee.org_unit_id == ouid,
        Employee.employee_key == employee_key,
    ).first()
    if exists:
        return _redirect(
            f"/ui/colaboradores?company_id={company_id}&branch_id={branch_id}"
            f"&org_unit_id={org_unit_id}&err=Ya+existe+un+colaborador+con+key+{employee_key}"
        )

    emp = Employee(
        org_unit_id=ouid,
        employee_key=employee_key,
        nombre=nombre,
        cargo_id=cargo_id.strip(),
        jornada_id=jornada_id.strip(),
    )
    if hasattr(emp, "contrato_max_min_semana"):
        emp.contrato_max_min_semana = contrato_max_min_semana
    if hasattr(emp, "email") and email:
        emp.email = email.strip()

    db.add(emp)
    db.commit()
    return _redirect(
        f"/ui/colaboradores?company_id={company_id}&branch_id={branch_id}"
        f"&org_unit_id={org_unit_id}&ok=Colaborador+{employee_key}+creado+exitosamente"
    )


@router.get("/colaboradores/{employee_id}", response_class=HTMLResponse)
def ui_colaborador_detail(employee_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    e = db.get(Employee, employee_id)
    if not e:
        raise HTTPException(status_code=404, detail="Employee not found")

    ou = db.get(OrgUnit, e.org_unit_id)
    br = db.get(Branch, ou.branch_id) if ou else None
    comp = db.get(Company, br.company_id) if br else None
    if not comp:
        raise HTTPException(status_code=400, detail="Employee without company context")

    if not can_see_company(db, current_user, comp.id):
        return _redirect("/ui/companies?err=Sin+acceso")

    template_path = _company_template_path(comp.id)
    seed_company_vocab(db, comp.id, template_path if template_path.exists() else None)
    vocab = list_vocab(db, comp.id)
    aus_codes = sorted(set([i.value for i in vocab.get(CAT_AUSENTISMO, [])] + AUS_CODES_FALLBACK))
    restr_tipos = sorted(set([i.value for i in vocab.get(CAT_RESTR_TIPO, [])] + RESTR_TIPOS_FALLBACK))
    jornadas = vocab.get(CAT_JORNADA, [])

    emp_key = e.employee_key

    ausentismos = (
        db.query(AusentismoEmpleado)
        .filter(AusentismoEmpleado.company_id == comp.id, AusentismoEmpleado.employee_id == emp_key)
        .order_by(AusentismoEmpleado.created_at.desc())
        .limit(200)
        .all()
    )

    restricciones = (
        db.query(RestriccionEmpleado)
        .filter(RestriccionEmpleado.company_id == comp.id, RestriccionEmpleado.employee_id == emp_key)
        .order_by(RestriccionEmpleado.created_at.desc())
        .limit(200)
        .all()
    )

    return TEMPLATES.TemplateResponse(
        "employee_detail.html",
        _enrich(
            db,
            request,
            {
                "request": request,
                "employee": e,
                "org_unit": ou,
                "branch": br,
                "company": comp,
                "aus_codes": aus_codes,
                "restr_tipos": restr_tipos,
                "jornadas": jornadas,
                "ausentismos": ausentismos,
                "restricciones": restricciones,
                "err": request.query_params.get("err"),
            },
        ),
    )


@router.post("/colaboradores/{employee_id}/ausentismos")
def ui_add_ausentismo(
    employee_id: uuid.UUID,
    request: Request,
    fecha_inicio: str = Form(...),
    fecha_fin: str = Form(...),
    ausentismo: str = Form(...),
    detalle: str = Form(""),
    hard: int = Form(1),
    penalizacion: int = Form(100000000),
    db: Session = Depends(get_db),
):
    e = db.get(Employee, employee_id)
    if not e:
        raise HTTPException(status_code=404, detail="Employee not found")

    ou = db.get(OrgUnit, e.org_unit_id)
    br = db.get(Branch, ou.branch_id) if ou else None
    comp = db.get(Company, br.company_id) if br else None
    if not comp:
        raise HTTPException(status_code=400, detail="Employee without company context")

    row = AusentismoEmpleado(
        company_id=comp.id,
        employee_id=e.employee_key,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        ausentismo=ausentismo.strip(),
        detalle=detalle.strip() if detalle else None,
        hard=int(hard),
        penalizacion=int(penalizacion),
    )
    db.add(row)
    db.commit()
    return _redirect(f"/ui/colaboradores/{employee_id}")


@router.post("/colaboradores/{employee_id}/ausentismos/{aus_id}/delete")
def ui_delete_ausentismo(employee_id: uuid.UUID, aus_id: uuid.UUID, db: Session = Depends(get_db)):
    row = db.get(AusentismoEmpleado, aus_id)
    if not row:
        raise HTTPException(status_code=404, detail="Ausentismo not found")
    db.delete(row)
    db.commit()
    return _redirect(f"/ui/colaboradores/{employee_id}")


@router.post("/colaboradores/{employee_id}/restricciones")
def ui_add_restriccion(
    employee_id: uuid.UUID,
    request: Request,
    tipo: str = Form(...),
    valor1: str = Form(""),
    valor2: str = Form(""),
    dia_semana: str = Form(""),
    fecha: str = Form(""),
    hard: int = Form(1),
    penalizacion: int = Form(100000000),
    detalle: str = Form(""),
    db: Session = Depends(get_db),
):
    e = db.get(Employee, employee_id)
    if not e:
        raise HTTPException(status_code=404, detail="Employee not found")

    ou = db.get(OrgUnit, e.org_unit_id)
    br = db.get(Branch, ou.branch_id) if ou else None
    comp = db.get(Company, br.company_id) if br else None
    if not comp:
        raise HTTPException(status_code=400, detail="Employee without company context")

    row = RestriccionEmpleado(
        company_id=comp.id,
        employee_id=e.employee_key,
        tipo=tipo.strip(),
        valor1=valor1.strip() if valor1 else None,
        valor2=valor2.strip() if valor2 else None,
        dia_semana=dia_semana.strip().upper() if dia_semana else None,
        fecha=fecha.strip() if fecha else None,
        hard=int(hard),
        penalizacion=int(penalizacion),
        detalle=detalle.strip() if detalle else None,
    )
    db.add(row)
    db.commit()
    return _redirect(f"/ui/colaboradores/{employee_id}")


@router.post("/colaboradores/{employee_id}/restricciones/{rid}/delete")
def ui_delete_restriccion(employee_id: uuid.UUID, rid: uuid.UUID, db: Session = Depends(get_db)):
    row = db.get(RestriccionEmpleado, rid)
    if not row:
        raise HTTPException(status_code=404, detail="Restriccion not found")
    db.delete(row)
    db.commit()
    return _redirect(f"/ui/colaboradores/{employee_id}")


@router.post("/colaboradores/{employee_id}/email")
def ui_colaborador_save_email(
    employee_id: uuid.UUID,
    email: str = Form(""),
    db: Session = Depends(get_db),
):
    emp = db.get(Employee, employee_id)
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    emp.email = email.strip() if email else None
    db.commit()
    return _redirect(f"/ui/colaboradores/{employee_id}?ok=Email+guardado")


# =========================
# AUSENTISMOS (BANDEJA)
# =========================
@router.get("/ausentismos", response_class=HTMLResponse)
def ui_ausentismos(request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    q_company_id = _parse_uuid(request.query_params.get("company_id"))
    q_branch_id = _parse_uuid(request.query_params.get("branch_id"))
    q_org_unit_id = _parse_uuid(request.query_params.get("org_unit_id"))
    q_tipo = (request.query_params.get("tipo") or "").strip()
    q_from = (request.query_params.get("from") or "").strip()  # YYYY-MM-DD
    q_to = (request.query_params.get("to") or "").strip()      # YYYY-MM-DD

    companies = filter_companies(db, current_user)
    branches: List[Branch] = []
    org_units: List[OrgUnit] = []

    aus_codes: List[str] = []
    employees: List[Employee] = []
    employee_by_key: Dict[str, Employee] = {}
    rows: List[AusentismoEmpleado] = []

    if q_company_id:
        branches = filter_branches(db, current_user, q_company_id)
        org_units = (
            filter_org_units(db, current_user, q_branch_id, q_company_id)
            if q_branch_id
            else filter_org_units(db, current_user, None, q_company_id)
        )

        # vocab por empresa (tipos ausentismo)
        template_path = _company_template_path(q_company_id)
        seed_company_vocab(db, q_company_id, template_path if template_path.exists() else None)
        vocab = list_vocab(db, q_company_id)
        aus_codes = sorted(set([i.value for i in vocab.get(CAT_AUSENTISMO, [])] + AUS_CODES_FALLBACK))

        # empleados visibles según filtros (para dropdown y map)
        emp_q = (
            db.query(Employee)
            .join(OrgUnit, Employee.org_unit_id == OrgUnit.id)
            .join(Branch, OrgUnit.branch_id == Branch.id)
            .filter(Branch.company_id == q_company_id)
        )
        if q_branch_id:
            emp_q = emp_q.filter(Branch.id == q_branch_id)
        if q_org_unit_id:
            emp_q = emp_q.filter(OrgUnit.id == q_org_unit_id)

        employees = emp_q.order_by(Employee.nombre.asc()).all()
        employee_by_key = {e.employee_key: e for e in employees}

        # query ausentismos (por company + empleados visibles)
        emp_keys = [e.employee_key for e in employees]
        a_q = db.query(AusentismoEmpleado).filter(AusentismoEmpleado.company_id == q_company_id)

        if emp_keys:
            a_q = a_q.filter(AusentismoEmpleado.employee_id.in_(emp_keys))
        else:
            a_q = a_q.filter(False)

        if q_tipo:
            a_q = a_q.filter(AusentismoEmpleado.ausentismo == q_tipo)

        # filtro por rango (si fechas son YYYY-MM-DD, comparar string sirve)
        if q_from:
            a_q = a_q.filter(AusentismoEmpleado.fecha_inicio >= q_from)
        if q_to:
            a_q = a_q.filter(AusentismoEmpleado.fecha_fin <= q_to)

        rows = a_q.order_by(AusentismoEmpleado.created_at.desc()).limit(300).all()

    return TEMPLATES.TemplateResponse(
        "ausentismos.html",
        _enrich(
            db,
            request,
            {
                "request": request,
                "companies": companies,
                "branches": branches,
                "org_units": org_units,
                "selected_company_id": str(q_company_id) if q_company_id else "",
                "selected_branch_id": str(q_branch_id) if q_branch_id else "",
                "selected_org_unit_id": str(q_org_unit_id) if q_org_unit_id else "",
                "tipo": q_tipo,
                "from": q_from,
                "to": q_to,
                "aus_codes": aus_codes,
                "employees": employees,
                "employee_by_key": employee_by_key,
                "rows": rows,
                "err": request.query_params.get("err"),
            },
        ),
    )


@router.post("/ausentismos/add")
def ui_ausentismos_add(
    request: Request,
    company_id: str = Form(...),
    employee_uuid: str = Form(...),
    ausentismo: str = Form(...),
    fecha_inicio: str = Form(...),
    fecha_fin: str = Form(...),
    detalle: str = Form(""),
    hard: int = Form(1),
    penalizacion: int = Form(100000000),
    db: Session = Depends(get_db),
):
    cid = _parse_uuid(company_id)
    eid = _parse_uuid(employee_uuid)
    if not cid or not eid:
        return _redirect("/ui/ausentismos?err=Empresa+o+Empleado+inválido")

    emp = db.get(Employee, eid)
    if not emp:
        return _redirect(f"/ui/ausentismos?company_id={cid}&err=Empleado+no+existe")

    row = AusentismoEmpleado(
        company_id=cid,
        employee_id=emp.employee_key,  # <- IMPORTANT: el solver lee esto como employee_id
        fecha_inicio=fecha_inicio.strip(),
        fecha_fin=fecha_fin.strip(),
        ausentismo=ausentismo.strip(),
        detalle=detalle.strip() if detalle else None,
        hard=int(hard),
        penalizacion=int(penalizacion),
    )
    db.add(row)
    db.commit()

    # UX: al guardar, te manda a la ficha del colaborador
    return _redirect(f"/ui/colaboradores/{emp.id}")


@router.post("/ausentismos/{aus_id}/delete")
def ui_ausentismos_delete(aus_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    row = db.get(AusentismoEmpleado, aus_id)
    if not row:
        raise HTTPException(status_code=404, detail="Ausentismo no encontrado")

    emp = db.query(Employee).filter(Employee.employee_key == row.employee_id).first()

    db.delete(row)
    db.commit()

    # UX: al eliminar, te manda a ficha si se puede
    if emp:
        return _redirect(f"/ui/colaboradores/{emp.id}")
    return _redirect("/ui/ausentismos")


# =========================
# DEMANDA
# =========================
@router.get("/demanda", response_class=HTMLResponse)
def ui_demanda(request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    q_company_id = _parse_uuid(request.query_params.get("company_id"))
    q_branch_id = _parse_uuid(request.query_params.get("branch_id"))
    q_org_unit_id = _parse_uuid(request.query_params.get("org_unit_id"))

    companies = filter_companies(db, current_user)
    branches: List[Branch] = []
    org_units: List[OrgUnit] = []

    if q_company_id:
        branches = filter_branches(db, current_user, q_company_id)
        org_units = (
            filter_org_units(db, current_user, q_branch_id, q_company_id)
            if q_branch_id
            else filter_org_units(db, current_user, None, q_company_id)
        )

    demand_rows: List[DemandUnit] = []
    ou_summary: List[Dict[str, Any]] = []

    if q_org_unit_id:
        demand_rows = (
            db.query(DemandUnit)
            .filter(DemandUnit.org_unit_id == q_org_unit_id)
            .order_by(DemandUnit.dia_semana.asc(), DemandUnit.inicio.asc())
            .all()
        )
    else:
        if q_company_id:
            ous = org_units
            for ou in ous:
                rows = (
                    db.query(DemandUnit)
                    .filter(DemandUnit.org_unit_id == ou.id)
                    .order_by(DemandUnit.dia_semana.asc(), DemandUnit.inicio.asc())
                    .all()
                )
                cnt = sum(1 for r in rows if r.active)
                # Contar tramos con holgura (ideal > mínimo)
                con_holgura = sum(
                    1 for r in rows 
                    if r.active and r.requeridos_ideal is not None and r.requeridos_ideal > r.requeridos
                )
                ou_summary.append({
                    "ou": ou, 
                    "rows_active": cnt, 
                    "rows": rows,
                    "con_holgura": con_holgura
                })

    dias = ["LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"]

    return TEMPLATES.TemplateResponse(
        "demand.html",
        _enrich(
            db,
            request,
            {
                "request": request,
                "companies": companies,
                "branches": branches,
                "org_units": org_units,
                "selected_company_id": str(q_company_id) if q_company_id else "",
                "selected_branch_id": str(q_branch_id) if q_branch_id else "",
                "selected_org_unit_id": str(q_org_unit_id) if q_org_unit_id else "",
                "dias": dias,
                "demand_rows": demand_rows,
                "ou_summary": ou_summary,
                "ok": request.query_params.get("ok"),
                "err": request.query_params.get("err"),
            },
        ),
    )


@router.post("/demanda/add")
def ui_demanda_add(
    request: Request,
    org_unit_id: str = Form(...),
    dia_semana: str = Form(...),
    inicio: str = Form(...),
    fin: str = Form(...),
    requeridos: int = Form(...),
    requeridos_ideal: Optional[str] = Form(None),  # NUEVO: Demanda ideal (puede venir vacío)
    active: int = Form(1),
    db: Session = Depends(get_db),
):
    ouid = _parse_uuid(org_unit_id)
    if not ouid:
        return _redirect("/ui/demanda?err=OU+inválida")

    # Obtener company_id y branch_id desde la OU para mantener los filtros
    ou = db.get(OrgUnit, ouid)
    company_id = ""
    branch_id = ""
    if ou and ou.branch:
        branch_id = str(ou.branch_id)
        if ou.branch.company:
            company_id = str(ou.branch.company_id)

    # Normalizar valores
    dia = dia_semana.strip().upper()
    ini = inicio.strip()
    fi = fin.strip()

    # Validación: ideal debe ser >= mínimo si está definido
    ideal_val = None
    if requeridos_ideal is not None and requeridos_ideal.strip() != "":
        ideal_val = int(requeridos_ideal)
        if ideal_val < int(requeridos):
            return _redirect(f"/ui/demanda?company_id={company_id}&branch_id={branch_id}&org_unit_id={ouid}&err=Ideal+debe+ser+mayor+o+igual+al+mínimo")

    # Buscar si ya existe esta combinación (UPSERT)
    existing = (
        db.query(DemandUnit)
        .filter(
            DemandUnit.org_unit_id == ouid,
            DemandUnit.dia_semana == dia,
            DemandUnit.inicio == ini,
            DemandUnit.fin == fi,
        )
        .first()
    )

    # Construir URL de redirección con todos los filtros
    redirect_base = f"/ui/demanda?company_id={company_id}&branch_id={branch_id}&org_unit_id={ouid}"

    if existing:
        # Actualizar existente
        existing.requeridos = int(requeridos)
        existing.requeridos_ideal = ideal_val
        existing.active = bool(int(active))
        db.commit()
        return _redirect(f"{redirect_base}&ok=Demanda+actualizada")
    else:
        # Crear nuevo
        row = DemandUnit(
            org_unit_id=ouid,
            dia_semana=dia,
            inicio=ini,
            fin=fi,
            requeridos=int(requeridos),
            requeridos_ideal=ideal_val,
            active=bool(int(active)),
        )
        db.add(row)
        db.commit()
        return _redirect(f"{redirect_base}&ok=Demanda+agregada")


@router.post("/demanda/{row_id}/delete")
def ui_demanda_delete(row_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    row = db.get(DemandUnit, row_id)
    if not row:
        raise HTTPException(status_code=404, detail="Demand row not found")
    ouid = row.org_unit_id
    db.delete(row)
    db.commit()
    return _redirect(f"/ui/demanda?org_unit_id={ouid}")


# =========================
# POOL
# =========================
@router.get("/pool", response_class=HTMLResponse)
def ui_pool(request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    q_company_id = _parse_uuid(request.query_params.get("company_id"))
    q_branch_id = _parse_uuid(request.query_params.get("branch_id"))
    q_org_unit_id = _parse_uuid(request.query_params.get("org_unit_id"))
    q_cargo_id = (request.query_params.get("cargo_id") or "").strip()

    companies = filter_companies(db, current_user)
    branches: List[Branch] = []
    org_units: List[OrgUnit] = []

    shift_ids: List[str] = []
    if q_company_id:
        branches = filter_branches(db, current_user, q_company_id)
        org_units = (
            filter_org_units(db, current_user, q_branch_id, q_company_id)
            if q_branch_id
            else filter_org_units(db, current_user, None, q_company_id)
        )

        template_path = _company_template_path(q_company_id)
        shift_ids = _read_shift_ids_from_template(template_path)

    dias = ["LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"]

    pool_rows: List[PoolTurno] = []
    cargo_options: List[str] = []

    if q_org_unit_id:
        pool_q = db.query(PoolTurno).filter(PoolTurno.org_unit_id == q_org_unit_id)
        if q_cargo_id:
            pool_q = pool_q.filter(PoolTurno.cargo_id == q_cargo_id)
        pool_rows = pool_q.order_by(
            PoolTurno.cargo_id.asc(), PoolTurno.dia_semana.asc(), PoolTurno.shift_id.asc()
        ).all()

        cargo_options = [
            r[0]
            for r in db.query(PoolTurno.cargo_id)
            .filter(PoolTurno.org_unit_id == q_org_unit_id)
            .distinct()
            .order_by(PoolTurno.cargo_id.asc())
            .all()
            if r and r[0]
        ]

    return TEMPLATES.TemplateResponse(
        "pool.html",
        _enrich(
            db,
            request,
            {
                "request": request,
                "companies": companies,
                "branches": branches,
                "org_units": org_units,
                "selected_company_id": str(q_company_id) if q_company_id else "",
                "selected_branch_id": str(q_branch_id) if q_branch_id else "",
                "selected_org_unit_id": str(q_org_unit_id) if q_org_unit_id else "",
                "cargo_id": q_cargo_id,
                "cargo_options": cargo_options,
                "dias": dias,
                "pool_rows": pool_rows,
                "shift_ids": shift_ids,
                "err": request.query_params.get("err"),
            },
        ),
    )


@router.post("/pool/add")
def ui_pool_add(
    request: Request,
    org_unit_id: str = Form(...),
    cargo_id: str = Form(...),
    cargo: str = Form(""),
    dia_semana: str = Form(...),
    shift_id: str = Form(...),
    habilitado: int = Form(1),
    db: Session = Depends(get_db),
):
    ouid = _parse_uuid(org_unit_id)
    if not ouid:
        return _redirect("/ui/pool?err=OU+inválida")

    dia = _normalize_dow(dia_semana)
    row = PoolTurno(
        org_unit_id=ouid,
        cargo_id=cargo_id.strip(),
        cargo=(cargo.strip() if cargo else cargo_id.strip()),
        dia_semana=dia,
        shift_id=shift_id.strip(),
        habilitado=int(habilitado),
    )
    db.add(row)
    db.commit()
    return _redirect(f"/ui/pool?org_unit_id={ouid}&cargo_id={cargo_id.strip()}")


@router.post("/pool/bulk")
def ui_pool_bulk(
    request: Request,
    org_unit_id: str = Form(...),
    cargo_id: str = Form(...),
    cargo: str = Form(""),
    default_days: str = Form(""),
    lines: str = Form(...),
    habilitado: int = Form(1),
    db: Session = Depends(get_db),
):
    ouid = _parse_uuid(org_unit_id)
    if not ouid:
        return _redirect("/ui/pool?err=OU+inválida")

    cargo_id = cargo_id.strip()
    cargo = (cargo.strip() if cargo else cargo_id)

    days_default = [d for d in (_normalize_dow(x) for x in (default_days or "").split(",")) if d]
    if not days_default:
        days_default = ["LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"]

    for raw in (lines or "").splitlines():
        s = raw.strip()
        if not s:
            continue
        parts = re.split(r"[,\t;]+", s)
        parts = [p.strip() for p in parts if p.strip()]

        if len(parts) == 1:
            shift = parts[0]
            for d in days_default:
                db.add(
                    PoolTurno(
                        org_unit_id=ouid,
                        cargo_id=cargo_id,
                        cargo=cargo,
                        dia_semana=d,
                        shift_id=shift,
                        habilitado=int(habilitado),
                    )
                )
        else:
            d = _normalize_dow(parts[0])
            shift = parts[1]
            db.add(
                PoolTurno(
                    org_unit_id=ouid,
                    cargo_id=cargo_id,
                    cargo=cargo,
                    dia_semana=d,
                    shift_id=shift,
                    habilitado=int(habilitado),
                )
            )

    db.commit()
    return _redirect(f"/ui/pool?org_unit_id={ouid}&cargo_id={cargo_id}")


@router.post("/pool/{row_id}/delete")
def ui_pool_delete(row_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    row = db.get(PoolTurno, row_id)
    if not row:
        raise HTTPException(status_code=404, detail="Pool row not found")
    ouid = row.org_unit_id
    cargo_id = row.cargo_id
    db.delete(row)
    db.commit()
    return _redirect(f"/ui/pool?org_unit_id={ouid}&cargo_id={cargo_id}")


# =========================
# RESTRICCIONES GLOBALES
# =========================
@router.get("/restricciones", response_class=HTMLResponse)
def ui_restricciones(request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    q_company_id = _parse_uuid(request.query_params.get("company_id"))
    q_branch_id = _parse_uuid(request.query_params.get("branch_id"))
    q_org_unit_id = _parse_uuid(request.query_params.get("org_unit_id"))
    q_cargo_id = (request.query_params.get("cargo_id") or "").strip()

    companies = filter_companies(db, current_user)
    branches: List[Branch] = []
    org_units: List[OrgUnit] = []

    restr_tipos: List[str] = []

    if q_company_id:
        branches = filter_branches(db, current_user, q_company_id)
        org_units = (
            filter_org_units(db, current_user, q_branch_id, q_company_id)
            if q_branch_id
            else filter_org_units(db, current_user, None, q_company_id)
        )

        template_path = _company_template_path(q_company_id)
        seed_company_vocab(db, q_company_id, template_path if template_path.exists() else None)
        vocab = list_vocab(db, q_company_id)
        restr_tipos = sorted(set([i.value for i in vocab.get(CAT_RESTR_TIPO, [])] + RESTR_TIPOS_FALLBACK))

    dias = ["", "LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"]

    rows: List[RestriccionEmpleado] = []
    if q_company_id:
        q = db.query(RestriccionEmpleado).filter(
            RestriccionEmpleado.company_id == q_company_id,
            RestriccionEmpleado.employee_id == None,  # noqa: E711
        )

        if q_org_unit_id:
            ou = db.get(OrgUnit, q_org_unit_id)
            if ou:
                q = q.filter((RestriccionEmpleado.valor1 == ou.org_unit_key) | (RestriccionEmpleado.valor2 == ou.org_unit_key))
        if q_cargo_id:
            q = q.filter((RestriccionEmpleado.valor1 == q_cargo_id) | (RestriccionEmpleado.valor2 == q_cargo_id))

        rows = q.order_by(RestriccionEmpleado.created_at.desc()).limit(300).all()

    return TEMPLATES.TemplateResponse(
        "restricciones_globales.html",
        _enrich(
            db,
            request,
            {
                "request": request,
                "companies": companies,
                "branches": branches,
                "org_units": org_units,
                "selected_company_id": str(q_company_id) if q_company_id else "",
                "selected_branch_id": str(q_branch_id) if q_branch_id else "",
                "selected_org_unit_id": str(q_org_unit_id) if q_org_unit_id else "",
                "cargo_id": q_cargo_id,
                "dias": dias,
                "restr_tipos": restr_tipos,
                "rows": rows,
                "err": request.query_params.get("err"),
            },
        ),
    )


@router.post("/restricciones/add-global")
def ui_add_restriccion_global(
    request: Request,
    company_id: str = Form(...),
    tipo: str = Form(...),
    valor1: str = Form(""),
    valor2: str = Form(""),
    dia_semana: str = Form(""),
    fecha: str = Form(""),
    hard: int = Form(1),
    penalizacion: int = Form(100000000),
    detalle: str = Form(""),
    db: Session = Depends(get_db),
):
    cid = _parse_uuid(company_id)
    if not cid:
        return _redirect("/ui/restricciones?err=Empresa+inválida")

    row = RestriccionEmpleado(
        company_id=cid,
        employee_id=None,
        tipo=tipo.strip(),
        valor1=valor1.strip() if valor1 else None,
        valor2=valor2.strip() if valor2 else None,
        dia_semana=dia_semana.strip().upper() if dia_semana else None,
        fecha=fecha.strip() if fecha else None,
        hard=int(hard),
        penalizacion=int(penalizacion),
        detalle=detalle.strip() if detalle else None,
    )
    db.add(row)
    db.commit()
    return _redirect(f"/ui/restricciones?company_id={cid}")


@router.post("/restricciones/{rid}/delete")
def ui_delete_restriccion_global(rid: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    row = db.get(RestriccionEmpleado, rid)
    if not row:
        raise HTTPException(status_code=404, detail="Restricción no encontrada")
    cid = row.company_id
    db.delete(row)
    db.commit()
    return _redirect(f"/ui/restricciones?company_id={cid}")

# =========================
# PLAN VIEWER (grid semanal + overrides + validación)
# =========================

import csv
from datetime import datetime, timezone, date, timedelta
from pathlib import Path
import shutil
from urllib.parse import quote_plus
from typing import Optional

from openpyxl import load_workbook

from api.plan_models import PlanOverride


def _parse_date_ymd(s: str) -> date:
    return date.fromisoformat(str(s)[:10])


def _monday(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _daterange(start: date, days: int) -> list[date]:
    return [start + timedelta(days=i) for i in range(days)]


def _load_plan_rows_from_run(run: Run) -> list[dict]:
    """
    Lee plan_mensual.csv (preferido). No usamos pandas.
    """
    out_dir = Path(run.out_dir)
    csv_path = out_dir / "plan_mensual.csv"
    if not csv_path.exists():
        return []

    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [r for r in reader]


def _load_catalogo_turnos(template_path: Path) -> dict[str, dict]:
    """
    Carga CatalogoTurnos: shift_id, inicio, fin, cruza_medianoche, tipo
    """
    if not template_path.exists():
        return {}

    wb = load_workbook(filename=str(template_path), data_only=True, read_only=True)
    if "CatalogoTurnos" not in wb.sheetnames:
        return {}

    ws = wb["CatalogoTurnos"]

    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        headers.append(str(v).strip() if v is not None else "")
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    required = ["shift_id", "inicio", "fin", "cruza_medianoche", "tipo"]
    for k in required:
        if k not in idx:
            return {}

    out: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(row=r, column=idx["shift_id"]).value
        if sid is None:
            continue
        sid = str(sid).strip()
        if not sid:
            continue

        inicio = ws.cell(row=r, column=idx["inicio"]).value
        fin = ws.cell(row=r, column=idx["fin"]).value
        cruza = ws.cell(row=r, column=idx["cruza_medianoche"]).value
        tipo = ws.cell(row=r, column=idx["tipo"]).value

        out[sid] = {
            "inicio": str(inicio)[:5] if inicio else None,
            "fin": str(fin)[:5] if fin else None,
            "cruza": (str(cruza).strip().upper() if cruza else "NO"),
            "tipo": (str(tipo).strip().upper() if tipo else ""),
        }
    return out


def _load_jornadas_min_descanso(template_path: Path) -> dict[str, int]:
    """
    Carga Jornadas: jornada_id -> min_descanso_horas
    """
    if not template_path.exists():
        return {}

    wb = load_workbook(filename=str(template_path), data_only=True, read_only=True)
    if "Jornadas" not in wb.sheetnames:
        return {}

    ws = wb["Jornadas"]

    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        headers.append(str(v).strip() if v is not None else "")
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    if "jornada_id" not in idx or "min_descanso_horas" not in idx:
        return {}

    out: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        jid = ws.cell(row=r, column=idx["jornada_id"]).value
        if jid is None:
            continue
        jid = str(jid).strip()
        if not jid:
            continue
        v = ws.cell(row=r, column=idx["min_descanso_horas"]).value
        try:
            out[jid] = int(v)
        except Exception:
            continue
    return out


def _is_worked(shift_id: str, cat: dict[str, dict]) -> bool:
    """True si el turno implica trabajo efectivo."""
    if not shift_id:
        return False
    sid = shift_id.strip().upper()
    if sid in ("LIBRE", "SALIENTE", "LM", "VAC", "PA", ""):
        return False
    info = cat.get(shift_id)
    if info:
        tipo = (info.get("tipo") or "").strip().upper()
        if tipo:
            return tipo in ("TRABAJADO", "T", "WORK", "W")
        return bool(info.get("inicio") and info.get("fin"))
    return True


def _shift_end_dt(d: date, shift_id: str, cat: dict[str, dict]) -> Optional[datetime]:
    info = cat.get(shift_id)
    if not info or not info.get("fin"):
        return None
    fin_h, fin_m = map(int, info["fin"].split(":"))
    end = datetime(d.year, d.month, d.day, fin_h, fin_m, tzinfo=timezone.utc)
    if info.get("cruza") in ("SI", "SÍ", "YES", "TRUE"):
        end = end + timedelta(days=1)
    return end


def _shift_start_dt(d: date, shift_id: str, cat: dict[str, dict]) -> Optional[datetime]:
    info = cat.get(shift_id)
    if not info or not info.get("inicio"):
        return None
    ini_h, ini_m = map(int, info["inicio"].split(":"))
    return datetime(d.year, d.month, d.day, ini_h, ini_m, tzinfo=timezone.utc)


def _validate_edit(
    employee: Employee,
    d: date,
    new_shift_id: str,
    plan_map: dict[tuple[str, str], str],
    cat: dict[str, dict],
    min_descanso_by_jornada: dict[str, int],
) -> tuple[bool, str | None, str | None]:
    """
    Retorna (ok, err, warn)
    """
    # 1) Descanso mínimo vs día anterior
    min_desc = min_descanso_by_jornada.get(employee.jornada_id)
    if min_desc:
        prev_date = d - timedelta(days=1)
        prev_shift = plan_map.get((employee.employee_key, prev_date.isoformat()), "LIBRE")
        prev_end = _shift_end_dt(prev_date, prev_shift, cat)
        new_start = _shift_start_dt(d, new_shift_id, cat)
        if prev_end and new_start:
            rest_hours = (new_start - prev_end).total_seconds() / 3600.0
            if rest_hours < float(min_desc):
                return (
                    False,
                    f"Descanso insuficiente: {rest_hours:.1f}h (< {min_desc}h) entre {prev_shift} y {new_shift_id}",
                    None,
                )

    # 2) Máximo 6 días trabajados consecutivos (ventana +/- 10 días)
    window_start = d - timedelta(days=10)
    window_end = d + timedelta(days=10)

    cursor = window_start
    worked_flags: list[bool] = []
    while cursor <= window_end:
        sid = plan_map.get((employee.employee_key, cursor.isoformat()), "LIBRE")
        if cursor == d:
            sid = new_shift_id
        worked_flags.append(_is_worked(sid, cat))
        cursor += timedelta(days=1)

    streak = 0
    max_streak = 0
    for w in worked_flags:
        if w:
            streak += 1
            max_streak = max(max_streak, streak)
        else:
            streak = 0

    if max_streak > 6:
        return (False, f"Regla: más de 6 días trabajados consecutivos (racha {max_streak}).", None)

    # 3) Domingos libres — mínimo 2 para contratos > 30h semanales
    warn = None
    contrato_min_sem = getattr(employee, "contrato_max_min_semana", 0) or 0
    # 30 horas = 1800 minutos semanales
    if contrato_min_sem >= 1800:
        month_start = date(d.year, d.month, 1)
        next_month = month_start.replace(day=28) + timedelta(days=4)
        month_end = next_month - timedelta(days=next_month.day)

        sundays_free = 0
        cursor = month_start
        while cursor <= month_end:
            if cursor.weekday() == 6:
                if cursor == d:
                    sid = new_shift_id  # usar el turno nuevo propuesto
                else:
                    sid = plan_map.get((employee.employee_key, cursor.isoformat()), "LIBRE")
                if not _is_worked(sid, cat):
                    sundays_free += 1
            cursor += timedelta(days=1)

        if sundays_free < 2:
            return (
                False,
                f"Regla: este cambio dejaría solo {sundays_free} domingo{'s' if sundays_free != 1 else ''} libre en el mes (mínimo 2 para contratos sobre 30 horas).",
                None,
            )

    return (True, None, warn)


def _build_demand_curves_for_week(
    run,
    ws_date,
    org_unit_id=None,
    org_unit_key=None,  # FIX: Usar el código de la OU para filtrar el CSV
    branch_id=None,
    company_id=None,
) -> dict:
    """
    Lee reporte_brechas.csv del run activo y construye las curvas de demanda
    por tramo horario para el panel inferior de la vista semanal.

    Filtra por org_unit_id / branch_id / company_id según el filtro activo,
    evitando acumular toda la empresa cuando hay un filtro aplicado.

    El CSV tiene columnas (nombres exactos del solver):
      - fecha
      - tramo_inicio          : HH:MM (cada 30 min)
      - requerimientos_minimos_personas  (o requeridos_personas)
      - requerimiento_ideal_persona      (o requeridos_ideal_personas)
      - cubierto_personas                (o cubiertos_personas)
      - org_unit_id

    Devuelve:
    {
        "dias": [
            {
                "fecha":     "2026-03-30",
                "label":     "Lun 30/03",
                "franjas":   ["06:00", "06:30", ...],
                "minimo":    [2, 3, 3, ...],
                "ideal":     [3, 4, 4, ...],
                "cubiertos": [2, 3, 3, ...],
                "tiene_demanda": True,
            },
            ...  (7 días)
        ]
    }
    """
    import csv as _csv
    from pathlib import Path as _Path
    from datetime import timedelta as _td
    from collections import defaultdict

    DAY_LABELS = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

    week_dates = {(ws_date + _td(days=i)).isoformat() for i in range(7)}

    # FIX: Usar org_unit_key (código de la OU) para filtrar el CSV
    # El CSV tiene el código (ej: "UO_PRUEBA"), no el UUID
    allowed_ou_keys = None
    if org_unit_key:
        allowed_ou_keys = {str(org_unit_key).lower().strip()}

    brechas_path = _Path(run.out_dir) / "reporte_brechas.csv"
    if not brechas_path.exists():
        return {"dias": []}

    # Acumular: {fecha: {tramo: {min, ideal, cub}}}
    data: dict = defaultdict(lambda: defaultdict(lambda: {"min": 0, "ideal": 0, "cub": 0}))

    def _col(row, *names):
        for n in names:
            v = row.get(n) or row.get(n.lower())
            if v not in (None, "", "nan"):
                try:
                    return float(v)
                except Exception:
                    pass
        return 0.0

    try:
        with brechas_path.open(encoding="utf-8-sig", newline="") as f:
            for row in _csv.DictReader(f):
                fecha = str(row.get("fecha", "") or "")[:10]
                if fecha not in week_dates:
                    continue

                # FIX: Filtrar por org_unit_key (código), no por UUID
                if allowed_ou_keys is not None:
                    row_ou = str(row.get("org_unit_id", "") or "").lower().strip()
                    if row_ou not in allowed_ou_keys:
                        continue

                tramo = str(
                    row.get("tramo_inicio") or row.get("tramo") or ""
                ).strip()
                if not tramo:
                    continue

                req_min = _col(row,
                    "requerimientos_minimos_personas",
                    "requeridos_personas",
                    "requeridos_min_personas",
                    "required",
                )
                req_ideal = _col(row,
                    "requerimiento_ideal_persona",
                    "requeridos_ideal_personas",
                    "requeridos_ideal",
                )
                if req_ideal == 0:
                    req_ideal = req_min

                cubiertos = _col(row,
                    "cubierto_personas",
                    "cubiertos_personas",
                    "covered",
                    "assigned",
                )

                slot = data[fecha][tramo]
                slot["min"]   += int(req_min)
                slot["ideal"] += int(req_ideal)
                slot["cub"]   += int(cubiertos)

    except Exception:
        import traceback
        traceback.print_exc()
        return {"dias": []}

    dias = []
    for i in range(7):
        fecha = (ws_date + _td(days=i)).isoformat()
        label = DAY_LABELS[i] + " " + fecha[8:10] + "/" + fecha[5:7]
        tramos_dia = data.get(fecha, {})

        if not tramos_dia:
            dias.append({
                "fecha": fecha, "label": label,
                "franjas": [], "minimo": [], "ideal": [], "cubiertos": [],
                "tiene_demanda": False,
            })
            continue

        franjas_ord = sorted(tramos_dia.keys())
        dias.append({
            "fecha":     fecha,
            "label":     label,
            "franjas":   franjas_ord,
            "minimo":    [tramos_dia[t]["min"]   for t in franjas_ord],
            "ideal":     [tramos_dia[t]["ideal"] for t in franjas_ord],
            "cubiertos": [tramos_dia[t]["cub"]   for t in franjas_ord],
            "tiene_demanda": any(tramos_dia[t]["min"] > 0 for t in franjas_ord),
        })

    return {"dias": dias}


@router.get("/plan", response_class=HTMLResponse)
def ui_plan(request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    q_company_id  = _parse_uuid(request.query_params.get("company_id"))
    q_branch_id   = _parse_uuid(request.query_params.get("branch_id"))
    q_org_unit_id = _parse_uuid(request.query_params.get("org_unit_id"))
    q_run_id      = _parse_uuid(request.query_params.get("run_id"))
    q_week_start  = (request.query_params.get("week_start") or "").strip()
    q_search      = (request.query_params.get("q") or "").strip()

    companies = filter_companies(db, current_user)
    runs: list[Run] = []
    run: Run | None = None

    # Sucursales y OUs para los selectores del filtro
    plan_branches: list[Branch] = []
    plan_org_units: list[OrgUnit] = []
    if q_company_id:
        plan_branches = (
            db.query(Branch)
            .filter(Branch.company_id == q_company_id)
            .order_by(Branch.code.asc())
            .all()
        )
    if q_branch_id:
        plan_org_units = (
            db.query(OrgUnit)
            .filter(OrgUnit.branch_id == q_branch_id)
            .order_by(OrgUnit.name.asc())
            .all()
        )

    if q_company_id:
        runs = (
            db.query(Run)
            .filter(Run.status == "success")
            .order_by(Run.created_at.desc())
            .limit(30)
            .all()
        )
        if q_run_id:
            run = db.get(Run, q_run_id)
        elif runs:
            run = runs[0]
            params = f"company_id={q_company_id}&run_id={run.id}"
            if q_branch_id:
                params += f"&branch_id={q_branch_id}"
            if q_org_unit_id:
                params += f"&org_unit_id={q_org_unit_id}"
            if q_week_start:
                params += f"&week_start={q_week_start}"
            return _redirect(f"/ui/plan?{params}")
        else:
            run = None

    if not run:
        return TEMPLATES.TemplateResponse(
            "plan_grid.html",
            _enrich(
                db,
                request,
                {
                    "request": request,
                    "companies": companies,
                    "runs": runs,
                    "run": None,
                    "selected_company_id": str(q_company_id) if q_company_id else "",
                    "selected_branch_id": str(q_branch_id) if q_branch_id else "",
                    "selected_org_unit_id": str(q_org_unit_id) if q_org_unit_id else "",
                    "plan_branches": plan_branches,
                    "plan_org_units": plan_org_units,
                    "selected_run_id": str(q_run_id) if q_run_id else "",
                    "week_starts": [],
                    "week_start": "",
                    "prev_week": None,
                    "next_week": None,
                    "days": [],
                    "grid_rows": [],
                    "shift_options": [],
                    "q": q_search,
                    "err": request.query_params.get("err"),
                    "warn": request.query_params.get("warn"),
                    "ok": request.query_params.get("ok"),
                },
            ),
        )

    template_path = _company_template_path(q_company_id)
    cat = _load_catalogo_turnos(template_path)
    min_desc = _load_jornadas_min_descanso(template_path)

    # vocab ausentismos (puede venir como ORM objects)
    seed_company_vocab(db, q_company_id, template_path if template_path.exists() else None)
    vocab = list_vocab(db, q_company_id)
    aus_items = vocab.get(CAT_AUSENTISMO, [])

    aus_codes: list[str] = []
    for it in aus_items:
        if it is None:
            continue
        if isinstance(it, str):
            code = it.strip()
        else:
            code = getattr(it, "code", str(it)).strip()
        if code:
            aus_codes.append(code)

    shift_options = ["LIBRE"] + sorted(set(list(cat.keys()) + aus_codes))

    plan_rows = _load_plan_rows_from_run(run)
    base_map: dict[tuple[str, str], str] = {}
    colacion_map: dict[tuple[str, str], str] = {}  # (eid, fecha) -> colacion_inicio
    all_dates: list[date] = []

    for r in plan_rows:
        eid = (r.get("employee_id") or "").strip()
        f = (r.get("fecha") or "").strip()[:10]
        sid = (r.get("shift_id") or "LIBRE").strip()
        col_inicio = (r.get("colacion_inicio") or "").strip()  # Nueva columna
        if eid and f and sid:
            base_map[(eid, f)] = sid
            if col_inicio:
                colacion_map[(eid, f)] = col_inicio
            try:
                all_dates.append(_parse_date_ymd(f))
            except Exception:
                pass

    if not all_dates:
        return _redirect(f"/ui/plan?company_id={q_company_id}&err=El+run+no+tiene+plan_mensual.csv")

    min_d = min(all_dates)
    max_d = max(all_dates)

    week_starts: list[str] = []
    cursor = _monday(min_d)
    while cursor <= max_d:
        week_starts.append(cursor.isoformat())
        cursor += timedelta(days=7)

    # Auto-seleccionar la semana que contiene HOY si no hay selección
    today = date.today()
    today_monday = _monday(today).isoformat()
    
    if q_week_start and q_week_start in week_starts:
        week_start = q_week_start
    elif today_monday in week_starts:
        # Si hoy está dentro del rango del plan, usar esa semana
        week_start = today_monday
    else:
        # Si hoy no está en el plan, usar la primera semana
        week_start = week_starts[0]
    
    idx = week_starts.index(week_start)
    prev_week = week_starts[idx - 1] if idx > 0 else None
    next_week = week_starts[idx + 1] if idx + 1 < len(week_starts) else None

    ws_date = _parse_date_ymd(week_start)
    days = []
    for dd in _daterange(ws_date, 7):
        label = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"][dd.weekday()]
        days.append({"date": dd.isoformat(), "label": label})

    ov_rows = db.query(PlanOverride).filter(PlanOverride.run_id == run.id).all()
    ov_map = {(o.employee_id, o.fecha[:10]): o.shift_id for o in ov_rows}

    plan_map = dict(base_map)
    for k, v in ov_map.items():
        plan_map[k] = v

    # Estados de respuesta para mostrar en la grilla
    from sqlalchemy import text as _sql
    ov_ids = [str(o.id) for o in ov_rows if hasattr(o, "id") and o.id]
    override_states: dict[tuple[str, str], str] = {}
    if ov_ids:
        try:
            placeholders = ",".join(f"'{x}'" for x in ov_ids)
            resp_rows = db.execute(
                _sql(f"SELECT employee_id, fecha, estado FROM plan_override_responses WHERE override_id IN ({placeholders})")
            ).fetchall()
            for emp_id, fecha_r, estado in resp_rows:
                override_states[(emp_id, str(fecha_r)[:10])] = estado
        except Exception:
            db.rollback()  # tabla puede no existir aún — evitar transacción abortada

    emp_q = db.query(Employee)
    if q_org_unit_id:
        emp_q = emp_q.filter(Employee.org_unit_id == q_org_unit_id)
    elif q_branch_id:
        ou_ids = [ou.id for ou in db.query(OrgUnit).filter(OrgUnit.branch_id == q_branch_id).all()]
        if ou_ids:
            emp_q = emp_q.filter(Employee.org_unit_id.in_(ou_ids))
    if q_search:
        like = f"%{q_search}%"
        emp_q = emp_q.filter((Employee.nombre.ilike(like)) | (Employee.employee_key.ilike(like)))

    employees = emp_q.order_by(Employee.nombre.asc()).limit(200).all()

    grid_rows = []
    for e in employees:
        cells = []
        has_any = False
        for d0 in _daterange(ws_date, 7):
            key = (e.employee_key, d0.isoformat())
            sid = plan_map.get(key, "LIBRE")
            is_override = key in ov_map
            if key in base_map or is_override:
                has_any = True
            cells.append({"date": d0.isoformat(), "shift_id": sid, "is_override": is_override, "colacion_inicio": colacion_map.get(key, "")})
        if has_any:
            grid_rows.append({
    "employee_id": e.employee_key,
    "nombre": e.nombre,
    "cargo_id": e.cargo_id,
    "jornada_id": e.jornada_id,
    "contrato_min_semana": getattr(e, "contrato_max_min_semana", 0) or 0,  # ← ya lo tienes
    "cells": cells,
})

    return TEMPLATES.TemplateResponse(
        "plan_grid.html",
        _enrich(
            db,
            request,
            {
                "request": request,
                "companies": companies,
                "runs": runs,
                "run": run,
                "selected_company_id": str(q_company_id),
                "selected_branch_id": str(q_branch_id) if q_branch_id else "",
                "selected_org_unit_id": str(q_org_unit_id) if q_org_unit_id else "",
                "plan_branches": plan_branches,
                "plan_org_units": plan_org_units,
                "selected_run_id": str(run.id),
                "week_starts": week_starts,
                "week_start": week_start,
                "prev_week": prev_week,
                "next_week": next_week,
                "days": days,
                "grid_rows": grid_rows,
                "shift_options": shift_options,
                "override_states": {f"{k[0]}|{k[1]}": v for k, v in override_states.items()},
                "cat_minutos": {
                    sid: max(0, (
                        (int(v["fin"][:2]) * 60 + int(v["fin"][3:5])) -
                        (int(v["inicio"][:2]) * 60 + int(v["inicio"][3:5])) +
                        (1440 if v.get("cruza") == "SI" else 0) -
                        (int(sid.rsplit("_", 1)[-1]) if sid.startswith("S_") and sid.rsplit("_", 1)[-1].isdigit() else 0)
                    ))
                    if v.get("inicio") and v.get("fin") else 0
                    for sid, v in cat.items()
                },
                "q": q_search,
                "shift_labels": _shift_labels(shift_options),
                "aus_codes": aus_codes,
                "today_str": date.today().isoformat(),
                "q_view": (request.query_params.get("view") or "week").strip(),
                "q_date": (request.query_params.get("date") or date.today().isoformat()).strip(),
                "cat_turnos": {
                    sid: {"inicio": v.get("inicio",""), "fin": v.get("fin",""), "cruza": v.get("cruza","NO")}
                    for sid, v in cat.items() if v.get("inicio") and v.get("fin")
                },
                "demand_curves": _build_demand_curves_for_week(
                    run=run,
                    ws_date=ws_date,
                    org_unit_id=q_org_unit_id,
                    org_unit_key=(db.get(OrgUnit, q_org_unit_id).org_unit_key if q_org_unit_id and db.get(OrgUnit, q_org_unit_id) else None),
                    branch_id=q_branch_id,
                    company_id=q_company_id,
                ) if run else {"dias": []},
                "err": request.query_params.get("err"),
                "warn": request.query_params.get("warn"),
                "ok": request.query_params.get("ok"),
            },
        ),
    )


@router.post("/plan/edit")
async def ui_plan_edit(
    request: Request,
    db: Session = Depends(get_db),
):
    form = await request.form()
    company_id = str(form.get("company_id", "")).strip()
    run_id     = str(form.get("run_id", "")).strip()
    week_start = str(form.get("week_start", "")).strip()
    # Construir return_to limpio para evitar double-encoding
    return_to = f"/ui/plan?company_id={company_id}&run_id={run_id}"
    if week_start:
        return_to += f"&week_start={week_start}"

    cid = _parse_uuid(company_id)
    rid = _parse_uuid(run_id)
    if not cid or not rid:
        return _redirect(f"{return_to}&err=Company+o+Run+inválido")

    run = db.get(Run, rid)
    if not run:
        return _redirect(f"{return_to}&err=Run+no+existe")

    template_path = _company_template_path(cid)
    cat = _load_catalogo_turnos(template_path)
    min_desc = _load_jornadas_min_descanso(template_path)

    plan_rows = _load_plan_rows_from_run(run)
    base_map: dict[tuple[str, str], str] = {}
    for r in plan_rows:
        eid = (r.get("employee_id") or "").strip()
        f   = (r.get("fecha") or "").strip()[:10]
        sid = (r.get("shift_id") or "LIBRE").strip()
        if eid and f and sid:
            base_map[(eid, f)] = sid

    ov_rows = db.query(PlanOverride).filter(PlanOverride.run_id == run.id).all()
    ov_map  = {(o.employee_id, o.fecha[:10]): o.shift_id for o in ov_rows}
    plan_map = dict(base_map)
    for k, v in ov_map.items():
        plan_map[k] = v

    # Detectar campos cell_{employee_id}_{fecha} que cambiaron respecto al plan actual
    saved = 0
    warns = []
    _pending_notify = []
    for key, value in form.items():
        if not key.startswith("cell_"):
            continue
        # key = "cell_EMP001_2026-03-11"
        parts = key[5:].rsplit("_", 3)  # split desde la derecha para aislar la fecha YYYY-MM-DD
        if len(parts) < 2:
            continue
        # La fecha siempre son los últimos 10 chars del key tras "cell_"
        rest = key[5:]  # "EMP001_2026-03-11"
        fecha = rest[-10:]
        employee_id = rest[:-11]  # quitar "_YYYY-MM-DD"

        shift_id = str(value).strip()

        # Solo procesar si cambió respecto al plan actual
        current = plan_map.get((employee_id, fecha), "LIBRE")
        if shift_id == current:
            continue

        employee = db.query(Employee).filter(
            Employee.employee_key == employee_id,
        ).first()
        if not employee:
            continue

        d = _parse_date_ymd(fecha)
        if not d:
            continue

        ok, err, warn = _validate_edit(employee, d, shift_id, plan_map, cat, min_desc)
        if not ok:
            return _redirect(f"{return_to}&err={err.replace(' ', '+')}")
        if warn:
            warns.append(warn)

        existing = (
            db.query(PlanOverride)
            .filter(
                PlanOverride.run_id == run.id,
                PlanOverride.employee_id == employee_id,
                PlanOverride.fecha == d.isoformat(),
            )
            .first()
        )
        if existing:
            existing.shift_id = shift_id
            existing.is_valid = True
            existing.error_message = None
            db.add(existing)
        else:
            db.add(PlanOverride(
                company_id=cid,
                run_id=run.id,
                employee_id=employee_id,
                fecha=d.isoformat(),
                shift_id=shift_id,
                is_valid=True,
            ))
        plan_map[(employee_id, fecha)] = shift_id
        saved += 1

        # Guardar info para notificar después del commit
        _pending_notify.append({
            "employee_id": employee_id,
            "fecha": fecha,
            "shift_id_old": current,
            "shift_id_new": shift_id,
        })

    db.commit()

    # Disparar notificaciones DESPUÉS del commit (para que el PlanOverride exista en BD)
    for _pn in _pending_notify:
        try:
            _emp_obj = db.query(Employee).filter(Employee.employee_key == _pn["employee_id"]).first()
            _emp_email = (getattr(_emp_obj, "email", "") or "").strip() if _emp_obj else ""
            _emp_nombre = _emp_obj.nombre if _emp_obj else _pn["employee_id"]
            _hours = int(os.getenv("OVERRIDE_RESPONSE_HOURS", "24"))
            _ov_obj = (
                db.query(PlanOverride)
                .filter(
                    PlanOverride.run_id == rid,
                    PlanOverride.employee_id == _pn["employee_id"],
                    PlanOverride.fecha == _pn["fecha"],
                )
                .first()
            )
            if _ov_obj:
                celery_app.send_task("notify_override_change", kwargs={
                    "override_id":     str(_ov_obj.id),
                    "run_id":          str(rid),
                    "company_id":      str(cid),
                    "employee_id":     _pn["employee_id"],
                    "fecha":           _pn["fecha"],
                    "shift_id_old":    _pn["shift_id_old"],
                    "shift_id_new":    _pn["shift_id_new"],
                    "employee_email":  _emp_email,
                    "employee_nombre": _emp_nombre,
                    "supervisor_email": "",
                    "hours":           _hours,
                })
                print(f"[notify_override] Tarea disparada para {_pn['employee_id']} fecha {_pn['fecha']}")
        except Exception as _notify_err:
            import traceback as _tb
            print(f"[notify_override] ERROR: {_notify_err}\n{_tb.format_exc()}")

    if saved == 0:
        return _redirect(f"{return_to}&warn=Sin+cambios+detectados")
    if warns:
        w = warns[0].replace(" ", "+")
        return _redirect(f"{return_to}&warn={w}&ok={saved}+cambios+guardados")
    return _redirect(f"{return_to}&ok={saved}+cambios+guardados")

# =========================
# RUNS
# =========================
@router.get("/runs", response_class=HTMLResponse)
def ui_runs(request: Request, db: Session = Depends(get_db)):
    runs = db.query(Run).order_by(Run.created_at.desc()).limit(50).all()
    return TEMPLATES.TemplateResponse("runs.html", _enrich(db, request, {"request": request, "runs": runs}))


@router.get("/runs/{run_id}", response_class=HTMLResponse)
def ui_run_detail(run_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    run: Run | None = db.get(Run, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    arts = _collect_artifacts(run)

    qa = load_qa(str(run.id))
    return TEMPLATES.TemplateResponse(
        "run_detail.html",
        _enrich(
            db,
            request,
            {
                "request": request,
                "run": run,
                "artifacts": arts,
                "qa_status": qa_status(qa),
                "qa_message": qa_message(qa),
                "qa_summary": qa_summary(qa),
                "auto_refresh": run.status in ("queued", "running"),
            },
        ),
    )


@router.get("/runs/{run_id}/input-case")
def ui_download_input_case(run_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    run: Run | None = db.get(Run, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    p = Path(run.case_path)
    if not p.exists():
        raise HTTPException(status_code=404, detail="Input case not found")
    return FileResponse(
        str(p),
        filename="case.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# =========================
# COMPANY: Template upload + Case data import + Request turnos
# =========================

@router.post("/companies/{company_id}/template")
async def ui_company_upload_template(
    company_id: uuid.UUID,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Sube el template base del solver (por empresa). Se guarda como case_template.xlsx
    y se usa para vocab (catálogos) + build_case_from_db.
    """
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    if not file.filename.lower().endswith(".xlsx"):
        return _redirect(f"/ui/companies/{company_id}?err=Template+debe+ser+.xlsx")

    template_path = _company_template_path(company_id)
    contents = await file.read()
    template_path.write_bytes(contents)

    # Seed vocab desde template (si aplica)
    try:
        seed_company_vocab(db, company_id, template_path)
    except Exception:
        pass

    return _redirect(f"/ui/companies/{company_id}?ok=Template+cargado")


def _xlsx_sheet_rows(xlsx_path: Path, sheet_name: str) -> list[dict]:
    """Lee una hoja como lista de dict usando fila 1 como headers. Ignora filas vacías."""
    if not xlsx_path.exists():
        return []
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    if ws.max_row < 2:
        return []

    headers: list[str] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        headers.append(str(v).strip() if v is not None else "")

    out: list[dict] = []
    for r in range(2, ws.max_row + 1):
        row: dict = {}
        empty = True
        for i, h in enumerate(headers, start=1):
            if not h:
                continue
            v = ws.cell(row=r, column=i).value
            if v is not None and str(v).strip() != "":
                empty = False
            row[h] = v
        if not empty:
            out.append(row)
    return out


@router.post("/companies/{company_id}/case-data-import")
async def ui_company_case_data_import(
    company_id: uuid.UUID,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Importa carga masiva (onboarding) desde un xlsx:
    Dotacion, DemandaUnidad, PoolTurnos, RestriccionesEmpleado, AusentismoEmpleado.
    """
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    if not file.filename.lower().endswith(".xlsx"):
        return _redirect(f"/ui/companies/{company_id}?err=Carga+masiva+debe+ser+.xlsx")

    tmp_path = Path("/tmp") / f"case_data_{company_id}.xlsx"
    tmp_path.write_bytes(await file.read())

    # Map org_unit_key -> OrgUnit (de esta empresa)
    ous = (
        db.query(OrgUnit)
        .join(Branch, OrgUnit.branch_id == Branch.id)
        .filter(Branch.company_id == company_id)
        .all()
    )
    ou_by_key = {o.org_unit_key: o for o in ous}

    created_emp = 0
    updated_emp = 0

    # ---------- Dotacion -> Employee ----------
    # Esperado: employee_id (o employee_key), nombre, org_unit_id (org_unit_key), cargo_id, jornada_id
    dot_rows = _xlsx_sheet_rows(tmp_path, "Dotacion")
    for r in dot_rows:
        emp_key = str(r.get("employee_id") or r.get("employee_key") or "").strip()
        nombre = str(r.get("nombre") or r.get("name") or "").strip()
        ou_key = str(r.get("org_unit_id") or r.get("org_unit_key") or "").strip()
        cargo_id = str(r.get("cargo_id") or "").strip()
        jornada_id = str(r.get("jornada_id") or "").strip()

        if not emp_key or not nombre or not ou_key:
            continue

        ou = ou_by_key.get(ou_key)
        if not ou:
            continue

        emp = db.query(Employee).filter(Employee.org_unit_id == ou.id, Employee.employee_key == emp_key).first()
        if emp:
            # OJO: tu modelo usa "nombre" en varios lugares; si tu columna real es "name", ajusta aquí.
            if hasattr(emp, "nombre"):
                emp.nombre = nombre
            elif hasattr(emp, "name"):
                emp.name = nombre

            if cargo_id:
                emp.cargo_id = cargo_id
            if jornada_id:
                emp.jornada_id = jornada_id
            updated_emp += 1
        else:
            kwargs = dict(
                org_unit_id=ou.id,
                employee_key=emp_key,
                cargo_id=cargo_id or None,
                jornada_id=jornada_id or None,
            )
            if "nombre" in Employee.__table__.columns:
                kwargs["nombre"] = nombre
            else:
                kwargs["name"] = nombre

            # IMPORTANTÍSIMO: NO pasar is_active (tu modelo no lo tiene)
            db.add(Employee(**kwargs))
            created_emp += 1

    # ---------- DemandaUnidad -> DemandUnit (reemplaza por OU) ----------
    dem_rows = _xlsx_sheet_rows(tmp_path, "DemandaUnidad")
    dem_by_ou: dict[str, list[dict]] = {}
    for r in dem_rows:
        ou_key = str(r.get("org_unit_id") or "").strip()
        if ou_key:
            dem_by_ou.setdefault(ou_key, []).append(r)

    for ou_key, rows in dem_by_ou.items():
        ou = ou_by_key.get(ou_key)
        if not ou:
            continue
        db.query(DemandUnit).filter(DemandUnit.org_unit_id == ou.id).delete()
        for r in rows:
            dia = _normalize_dow(str(r.get("dia_semana") or "").strip())
            inicio = str(r.get("inicio") or "").strip()
            fin = str(r.get("fin") or "").strip()
            try:
                req_i = int(r.get("requeridos"))
            except Exception:
                continue
            db.add(DemandUnit(org_unit_id=ou.id, dia_semana=dia, inicio=inicio, fin=fin, requeridos=req_i, active=True))

    # ---------- PoolTurnos -> PoolTurno (reemplaza por OU) ----------
    pool_rows = _xlsx_sheet_rows(tmp_path, "PoolTurnos")
    pool_by_ou: dict[str, list[dict]] = {}
    for r in pool_rows:
        ou_key = str(r.get("org_unit_id") or "").strip()
        if ou_key:
            pool_by_ou.setdefault(ou_key, []).append(r)

    for ou_key, rows in pool_by_ou.items():
        ou = ou_by_key.get(ou_key)
        if not ou:
            continue
        db.query(PoolTurno).filter(PoolTurno.org_unit_id == ou.id).delete()
        for r in rows:
            cargo_id = str(r.get("cargo_id") or "").strip()
            cargo = str(r.get("cargo") or cargo_id).strip()
            dia = _normalize_dow(str(r.get("dia_semana") or "").strip())
            shift_id = str(r.get("shift_id") or "").strip()
            try:
                hab_i = int(r.get("habilitado") or 1)
            except Exception:
                hab_i = 1
            if not cargo_id or not dia or not shift_id:
                continue
            db.add(PoolTurno(org_unit_id=ou.id, cargo_id=cargo_id, cargo=cargo, dia_semana=dia, shift_id=shift_id, habilitado=hab_i))

    # ---------- RestriccionesEmpleado ----------
    restr_rows = _xlsx_sheet_rows(tmp_path, "RestriccionesEmpleado")
    if restr_rows:
        db.query(RestriccionEmpleado).filter(RestriccionEmpleado.company_id == company_id).delete()
        for r in restr_rows:
            tipo = str(r.get("tipo") or "").strip()
            if not tipo:
                continue
            employee_id = str(r.get("employee_id") or "").strip() or None
            db.add(
                RestriccionEmpleado(
                    company_id=company_id,
                    employee_id=employee_id,
                    tipo=tipo,
                    valor1=str(r.get("valor1") or "").strip() or None,
                    valor2=str(r.get("valor2") or "").strip() or None,
                    dia_semana=_normalize_dow(str(r.get("dia_semana") or "").strip()) or None,
                    fecha=str(r.get("fecha") or "").strip() or None,
                    hard=int(r.get("hard") or 1),
                    penalizacion=int(r.get("penalizacion") or 100000000),
                    detalle=str(r.get("detalle") or "").strip() or None,
                )
            )

    # ---------- AusentismoEmpleado ----------
    aus_rows = _xlsx_sheet_rows(tmp_path, "AusentismoEmpleado")
    if aus_rows:
        db.query(AusentismoEmpleado).filter(AusentismoEmpleado.company_id == company_id).delete()
        for r in aus_rows:
            employee_id = str(r.get("employee_id") or "").strip()
            if not employee_id:
                continue
            db.add(
                AusentismoEmpleado(
                    company_id=company_id,
                    employee_id=employee_id,
                    fecha_inicio=str(r.get("fecha_inicio") or "").strip(),
                    fecha_fin=str(r.get("fecha_fin") or "").strip(),
                    ausentismo=str(r.get("ausentismo") or "").strip(),
                    detalle=str(r.get("detalle") or "").strip() or None,
                    hard=int(r.get("hard") or 1),
                    penalizacion=int(r.get("penalizacion") or 100000000),
                )
            )

    db.commit()
    return _redirect(f"/ui/companies/{company_id}?ok=Carga+masiva+OK+(emp+nuevos:{created_emp},+actualizados:{updated_emp})")


@router.post("/companies/{company_id}/request-turnos")
def ui_company_request_turnos(
    company_id: uuid.UUID,
    request: Request,
    month: str = Form(...),  # YYYY-MM
    db: Session = Depends(get_db),
):
    """
    Genera case.xlsx desde la BD (NO subes case), usando template base del solver.
    Semanas = número de domingos del mes (tu regla).
    """
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    # Validar mes
    try:
        y, m = month.strip().split("-")
        year = int(y)
        mon = int(m)
        if mon < 1 or mon > 12:
            raise ValueError()
    except Exception:
        return _redirect(f"/ui/companies/{company_id}?err=Mes+inválido+(usa+YYYY-MM)")

    template_path = _company_template_path(company_id)
    if not template_path.exists():
        return _redirect(f"/ui/companies/{company_id}?err=Primero+sube+el+template+base+del+solver")

    semanas = _count_sundays_in_month(year, mon)

    run_id = uuid.uuid4()
    paths = get_run_paths(run_id)
    ensure_dirs(paths)

    # OJO: tu case_builder.py REAL usa esta firma:
    # build_case_from_db(db, company_id, month, template_path, out_case_path)
    try:
        build_case_from_db(
            db=db,
            company_id=company_id,
            month=f"{year:04d}-{mon:02d}",
            template_path=template_path,
            out_case_path=paths.case_path,
        )
    except Exception as e:
        return _redirect(f"/ui/companies/{company_id}?err=No+se+pudo+generar+case:+{type(e).__name__}")

    run = Run(
        id=run_id,
        status="queued",
        original_filename=f"generated_case_{year:04d}-{mon:02d}.xlsx",
        case_path=str(paths.case_path),
        out_dir=str(paths.out_dir),
        log_path=str(paths.log_path),
    )
    db.add(run)
    db.commit()

    celery_app.send_task('execute_run', args=[str(run_id)])
    return _redirect(f"/ui/runs/{run_id}")

# =====================================================
# EXPORTS (PDF / Excel)
# =====================================================

def _parse_month_yyyymm(month: str) -> tuple[int, int]:
    s = (month or "").strip()
    try:
        y_s, m_s = s.split("-", 1)
        y = int(y_s)
        m = int(m_s)
        if m < 1 or m > 12:
            raise ValueError
        return y, m
    except Exception:
        raise ValueError("month must be YYYY-MM (e.g. 2026-03)")


def _month_dates(year: int, month: int) -> list[date]:
    last = calendar.monthrange(year, month)[1]
    return [date(year, month, d) for d in range(1, last + 1)]


def _short_shift_id(shift_id: str) -> str:
    """
    Convierte el ID técnico del turno en label legible.
    S_0730_1530_60 -> 07:30-15:30 · 1h col.
    S_0730_1530_30 -> 07:30-15:30 · 30m col.
    S_0730_1530    -> 07:30-15:30
    """
    s = (shift_id or "").strip()
    if not s or s.upper() == "LIBRE":
        return s or "LIBRE"
    if s.startswith("S_"):
        parts = s.split("_")
        if len(parts) >= 3 and len(parts[1]) == 4 and len(parts[2]) == 4:
            a, b = parts[1], parts[2]
            base = f"{a[:2]}:{a[2:]}-{b[:2]}:{b[2:]}"
            if len(parts) >= 4 and parts[3].isdigit():
                col = int(parts[3])
                if col > 0:
                    col_str = f"{col // 60}h col." if col % 60 == 0 else f"{col}m col."
                    return f"{base} · {col_str}"
            return base
    return s


def _shift_labels(shift_options: list) -> dict:
    """Devuelve {shift_id: label_legible} para el template."""
    return {sid: _short_shift_id(sid) for sid in shift_options}


def _plan_map_for_run(db: Session, run: Run) -> dict[tuple[str, str], str]:
    """Mergea plan_mensual.csv + overrides."""
    plan_rows = _load_plan_rows_from_run(run)
    base_map: dict[tuple[str, str], str] = {}
    for r in plan_rows:
        eid = (r.get("employee_id") or "").strip()
        f = (r.get("fecha") or "").strip()[:10]
        sid = (r.get("shift_id") or "").strip()
        # Nota: NO inventamos LIBRE. Si el solver escribió "LIBRE", vendrá explícito.
        if eid and f and sid:
            base_map[(eid, f)] = sid

    ov_rows = db.query(PlanOverride).filter(PlanOverride.run_id == run.id).all()
    ov_map = {(o.employee_id, o.fecha[:10]): o.shift_id for o in ov_rows}

    merged = dict(base_map)
    for k, v in ov_map.items():
        merged[k] = v
    return merged



@router.get("/companies/{company_id}/vocab", response_class=HTMLResponse)
def ui_company_vocab(company_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    """Administrar catálogos que viven en el template por empresa (Jornadas + CatalogoTurnos)."""
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")
    if not can_see_company(db, current_user, company_id):
        return _redirect("/ui/companies?err=Sin+acceso+a+empresa")
    if not getattr(current_user, "can_manage_catalogs", False):
        return _redirect(f"/ui/companies/{company_id}?err=Sin+permiso+para+catálogos")

    company = db.get(Company, company_id)
    if not company:
        return _redirect("/ui/companies?err=Empresa+no+existe")

    template_path = _company_template_path(company_id)
    template_ok = template_path.exists()

    jornadas_rows = _template_read_rows(template_path, "Jornadas") if template_ok else []
    turnos_rows = _template_read_rows(template_path, "CatalogoTurnos") if template_ok else []

    jornadas_ok = len(jornadas_rows) > 0
    turnos_ok = len(turnos_rows) > 0

    return TEMPLATES.TemplateResponse(
        _company_vocab_template_name(),
        _enrich(
            db,
            request,
            {
                "request": request,
                "company": company,
                "template_ok": template_ok,
                "jornadas_ok": jornadas_ok,
                "turnos_ok": turnos_ok,
                "jornadas_rows": jornadas_rows,
                "turnos_rows": turnos_rows,
                "err": request.query_params.get("err"),
                "ok": request.query_params.get("ok"),
            },
        ),
    )


@router.post("/companies/{company_id}/catalog/jornadas/add")
def ui_company_catalog_add_jornada(
    company_id: uuid.UUID,
    request: Request,
    jornada_id: str = Form(...),
    cap_min_semana: int = Form(...),
    dias_trabajo_obj_sem: int = Form(...),
    domingos_libre_mes: int = Form(...),
    min_descanso_horas: int = Form(...),
    dias_permitidos_semana: str = Form(...),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")
    if not can_see_company(db, current_user, company_id):
        return _redirect("/ui/companies?err=Sin+acceso+a+empresa")
    if not getattr(current_user, "can_manage_catalogs", False):
        return _redirect(f"/ui/companies/{company_id}/vocab?err=Sin+permiso+para+catálogos")

    template_path = _company_template_path(company_id)
    if not template_path.exists():
        return _redirect(f"/ui/companies/{company_id}/vocab?err=Primero+sube+el+template+base+(case_template.xlsx)")

    payload = {
        "jornada_id": _norm_str(jornada_id),
        "cap_min_semana": int(cap_min_semana),
        "dias_trabajo_obj_sem": int(dias_trabajo_obj_sem),
        "domingos_libre_mes": int(domingos_libre_mes),
        "min_descanso_horas": int(min_descanso_horas),
        "dias_permitidos_semana": _norm_str(dias_permitidos_semana),
    }

    try:
        _template_append_row(template_path, "Jornadas", "jornada_id", payload)
    except Exception as e:
        return _redirect(f"/ui/companies/{company_id}/vocab?err=No+se+pudo+agregar+jornada:+{quote_plus(str(e))}")

    return _redirect(f"/ui/companies/{company_id}/vocab?ok=Jornada+agregada")


@router.post("/companies/{company_id}/catalog/turnos/add")
def ui_company_catalog_add_turno(
    company_id: uuid.UUID,
    request: Request,
    shift_id: str = Form(...),
    nombre: str = Form(""),
    inicio: str = Form(...),  # HH:MM
    fin: str = Form(...),     # HH:MM
    cruza_medianoche: str = Form("NO"),
    colacion_min: int = Form(0),
    minutos_efectivos: int = Form(...),
    tipo: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")
    if not can_see_company(db, current_user, company_id):
        return _redirect("/ui/companies?err=Sin+acceso+a+empresa")
    if not getattr(current_user, "can_manage_catalogs", False):
        return _redirect(f"/ui/companies/{company_id}/vocab?err=Sin+permiso+para+catálogos")

    template_path = _company_template_path(company_id)
    if not template_path.exists():
        return _redirect(f"/ui/companies/{company_id}/vocab?err=Primero+sube+el+template+base+(case_template.xlsx)")

    # normalizaciones simples
    def hhmm(s: str) -> str:
        s = _norm_str(s)
        return s[:5] if len(s) >= 5 else s

    cruza = _norm_str(cruza_medianoche).upper()
    if cruza in ("SI", "SÍ", "TRUE", "1"):
        cruza = "SI"
    elif cruza in ("NO", "FALSE", "0", ""):
        cruza = "NO"

    payload = {
        "shift_id": _norm_str(shift_id),
        "nombre": _norm_str(nombre),
        "inicio": hhmm(inicio),
        "fin": hhmm(fin),
        "cruza_medianoche": cruza,
        "colacion_min": int(colacion_min or 0),
        "minutos_efectivos": int(minutos_efectivos),
        "tipo": _norm_str(tipo).upper(),
    }

    try:
        _template_append_row(template_path, "CatalogoTurnos", "shift_id", payload)
    except Exception as e:
        return _redirect(f"/ui/companies/{company_id}/vocab?err=No+se+pudo+agregar+turno:+{quote_plus(str(e))}")

    return _redirect(f"/ui/companies/{company_id}/vocab?ok=Turno+agregado")

@router.get("/exports", response_class=HTMLResponse)
def ui_exports(request: Request, db: Session = Depends(get_db)):
    """Pantalla para exportar (por persona / por OU)."""
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    q_company_id = _parse_uuid(request.query_params.get("company_id"))
    q_run_id = _parse_uuid(request.query_params.get("run_id"))
    q_month = (request.query_params.get("month") or "").strip()
    q_org_unit_id = _parse_uuid(request.query_params.get("org_unit_id"))

    companies = filter_companies(db, current_user)

    runs: list[Run] = (
        db.query(Run)
        .filter(Run.status == "success")
        .order_by(Run.created_at.desc())
        .limit(50)
        .all()
    )

    # Defaults
    if not q_company_id and ctx.company_id:
        q_company_id = ctx.company_id

    # Month default: current month
    if not q_month:
        today = date.today()
        q_month = f"{today.year:04d}-{today.month:02d}"

    org_units: list[OrgUnit] = []
    employees: list[Employee] = []

    if q_company_id:
        # OrgUnits por empresa
        org_units = (
            db.query(OrgUnit)
            .join(Branch, Branch.id == OrgUnit.branch_id)
            .filter(Branch.company_id == q_company_id)
            .order_by(OrgUnit.org_unit_key.asc())
            .all()
        )

        emp_q = (
            db.query(Employee)
            .join(OrgUnit, OrgUnit.id == Employee.org_unit_id)
            .join(Branch, Branch.id == OrgUnit.branch_id)
            .filter(Branch.company_id == q_company_id)
        )
        if q_org_unit_id:
            emp_q = emp_q.filter(Employee.org_unit_id == q_org_unit_id)
        employees = emp_q.order_by(Employee.nombre.asc()).limit(500).all()

    # Run selected
    run = db.get(Run, q_run_id) if q_run_id else (runs[0] if runs else None)

    return TEMPLATES.TemplateResponse(
        "exports.html",
        _enrich(
            db,
            request,
            {
                "request": request,
                "companies": companies,
                "runs": runs,
                "run": run,
                "selected_company_id": str(q_company_id) if q_company_id else "",
                "selected_run_id": str(run.id) if run else "",
                "month": q_month,
                "org_units": org_units,
                "employees": employees,
                "selected_org_unit_id": str(q_org_unit_id) if q_org_unit_id else "",
                "err": request.query_params.get("err"),
                "ok": request.query_params.get("ok"),
            },
        ),
    )


def _check_internal_token(request: Request) -> bool:
    """Verifica si la request viene del worker usando el token interno."""
    internal_token = os.getenv("INTERNAL_API_TOKEN", "")
    if not internal_token:
        return False
    req_token = request.headers.get("X-Internal-Token", "")
    return bool(internal_token) and req_token == internal_token


@router.get("/exports/employee.pdf")
def ui_export_employee_pdf(
    request: Request,
    company_id: str,
    run_id: str,
    month: str,
    employee_id: str,
    db: Session = Depends(get_db),
):
    """PDF tipo calendario (mes completo) para una persona.
    Acepta autenticación por sesión de usuario O por X-Internal-Token (para el worker).
    """
    is_internal = _check_internal_token(request)
    if not is_internal:
        ctx = _load_ctx(request)
        current_user = get_current_user(db, ctx.user_id)
        if not current_user:
            raise HTTPException(status_code=401, detail="Not authenticated")
    else:
        current_user = None

    cid = _parse_uuid(company_id)
    rid = _parse_uuid(run_id)
    if not cid or not rid:
        raise HTTPException(status_code=400, detail="company_id/run_id inválido")

    if not is_internal and not can_see_company(db, current_user, cid):
        raise HTTPException(status_code=403, detail="Forbidden")

    run = db.get(Run, rid)
    if not run:
        raise HTTPException(status_code=404, detail="Run no existe")

    emp = (
        db.query(Employee)
        .join(OrgUnit, Employee.org_unit_id == OrgUnit.id)
        .join(Branch, OrgUnit.branch_id == Branch.id)
        .filter(Employee.employee_key == employee_id, Branch.company_id == cid)
        .first()
    )
    if not emp:
        raise HTTPException(status_code=404, detail="Empleado no existe")

    try:
        y, m = _parse_month_yyyymm(month)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    plan_map = _plan_map_for_run(db, run)

    # --- build calendar matrix
    cal = calendar.Calendar(firstweekday=0)  # Monday
    weeks = cal.monthdatescalendar(y, m)
    # Horizon: sólo mostramos celdas para fechas que existen en el plan (resultado del solver).
    emp_dates = {dt for (eid, dt) in plan_map.keys() if eid == emp.employee_key}
    # reportlab (PDF)
    from io import BytesIO
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail="PDF no disponible: falta dependencia 'reportlab' en el contenedor. Agrega 'reportlab' a requirements y rebuild. Detalle: %s" % (e,),
        )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"Turnos {month} - {emp.nombre}",
    )

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=16, leading=18, spaceAfter=6)
    meta = ParagraphStyle("meta", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#475569"))
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, leading=9)
    cell_muted = ParagraphStyle("cellm", parent=cell, textColor=colors.HexColor("#94A3B8"))

    month_name = [
        "Enero",
        "Febrero",
        "Marzo",
        "Abril",
        "Mayo",
        "Junio",
        "Julio",
        "Agosto",
        "Septiembre",
        "Octubre",
        "Noviembre",
        "Diciembre",
    ][m - 1]

    elems = [
        Paragraph(f"Calendario de turnos - {month_name} {y}", h1),
        Paragraph(f"Colaborador: <b>{emp.nombre}</b> ({emp.employee_key})", meta),
        Spacer(1, 8),
    ]

    # Header days
    day_names = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    data: list[list[Any]] = [day_names]

    for w in weeks:
        row = []
        for d in w:
            # Mostrar exactamente el resultado del solver:
            # - Si la fecha NO aparece en el plan para este colaborador => celda en blanco (sin número).
            # - Si aparece => mostramos día + shift_id (o vacío si viene vacío).
            dt = d.isoformat()
            if dt not in emp_dates:
                row.append(Paragraph("&nbsp;", cell_muted))
                continue
            sid = plan_map.get((emp.employee_key, dt))
            sid_s = _short_shift_id(sid) if sid else ""
            if sid_s:
                row.append(Paragraph(f"<b>{d.day}</b><br/>{sid_s}", cell))
            else:
                row.append(Paragraph(f"<b>{d.day}</b><br/>", cell))
        data.append(row)

    tbl = Table(data, colWidths=[(A4[0] - 28 * mm) / 7.0] * 7)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F1F5F9")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("LEFTPADDING", (0, 1), (-1, -1), 6),
                ("RIGHTPADDING", (0, 1), (-1, -1), 6),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 8),
            ]
        )
    )

    elems.append(tbl)

    doc.build(elems)
    buf.seek(0)

    from fastapi.responses import StreamingResponse

    filename = f"turnos_{month}_{emp.employee_key}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store, max-age=0",
            "Pragma": "no-cache",
        },
    )


@router.get("/exports/org-unit.pdf")
def ui_export_org_unit_pdf(
    request: Request,
    company_id: str,
    run_id: str,
    month: str,
    org_unit_id: str,
    db: Session = Depends(get_db),
):
    """PDF mensual por OU (horizontal)."""
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    cid = _parse_uuid(company_id)
    rid = _parse_uuid(run_id)
    ouid = _parse_uuid(org_unit_id)
    if not cid or not rid or not ouid:
        raise HTTPException(status_code=400, detail="company_id/run_id/org_unit_id inválido")

    if not can_see_company(db, current_user, cid):
        raise HTTPException(status_code=403, detail="Forbidden")

    run = db.get(Run, rid)
    if not run:
        raise HTTPException(status_code=404, detail="Run no existe")

    ou = db.get(OrgUnit, ouid)
    if not ou:
        raise HTTPException(status_code=404, detail="OU no existe")

    br = db.get(Branch, ou.branch_id)
    if not br or br.company_id != cid:
        raise HTTPException(status_code=403, detail="OU fuera de la empresa")

    try:
        y, m = _parse_month_yyyymm(month)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    employees = db.query(Employee).filter(Employee.org_unit_id == ouid).order_by(Employee.nombre.asc()).all()

    plan_map = _plan_map_for_run(db, run)
    dates = _month_dates(y, m)

    # Split columns into 2 pages max (1-16, 17-end)
    chunks: list[list[date]] = []
    if len(dates) <= 16:
        chunks = [dates]
    else:
        chunks = [dates[:16], dates[16:]]
    # reportlab (PDF)
    from io import BytesIO
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail="PDF no disponible: falta dependencia 'reportlab' en el contenedor. Agrega 'reportlab' a requirements y rebuild. Detalle: %s" % (e,),
        )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"Plan OU {ou.org_unit_key} {month}",
    )

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading2"], fontSize=14, leading=16, spaceAfter=6)
    meta = ParagraphStyle("meta", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#475569"))
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=6.2, leading=7.2)

    month_name = [
        "Enero",
        "Febrero",
        "Marzo",
        "Abril",
        "Mayo",
        "Junio",
        "Julio",
        "Agosto",
        "Septiembre",
        "Octubre",
        "Noviembre",
        "Diciembre",
    ][m - 1]

    elems = [
        Paragraph(f"Planificación mensual - {ou.org_unit_key} / {ou.name}", h1),
        Paragraph(f"Mes: <b>{month_name} {y}</b> | Run: {run.id}", meta),
        Spacer(1, 6),
    ]

    page_w, page_h = landscape(A4)
    usable_w = page_w - 20 * mm

    for ci, chunk in enumerate(chunks):
        header = ["Colaborador"] + [str(d.day) for d in chunk]
        data: list[list[Any]] = [header]

        for e in employees:
            row: list[Any] = [Paragraph(f"<b>{e.nombre}</b><br/>{e.employee_key}", cell)]
            for d in chunk:
                sid = plan_map.get((e.employee_key, d.isoformat()), "LIBRE")
                row.append(Paragraph(_short_shift_id(sid), cell))
            data.append(row)

        # column widths
        name_w = 70 * mm
        day_w = max(10.5 * mm, (usable_w - name_w) / max(1, len(chunk)))
        col_widths = [name_w] + [day_w] * len(chunk)

        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F1F5F9")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 7),
            ("ALIGN", (1, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#E2E8F0")),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]

        # weekend shading
        for j, d in enumerate(chunk, start=1):
            if d.weekday() >= 5:
                style_cmds.append(("BACKGROUND", (j, 1), (j, -1), colors.HexColor("#FBFDFF")))

        tbl.setStyle(TableStyle(style_cmds))

        if ci > 0:
            elems.append(PageBreak())
        elems.append(tbl)

    doc.build(elems)
    buf.seek(0)

    from fastapi.responses import StreamingResponse

    filename = f"plan_ou_{ou.org_unit_key}_{month}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store, max-age=0",
            "Pragma": "no-cache",
        },
    )


@router.get("/exports/org-unit.xlsx")
def ui_export_org_unit_xlsx(
    request: Request,
    company_id: str,
    run_id: str,
    month: str,
    org_unit_id: str,
    db: Session = Depends(get_db),
):
    """Excel mensual por OU (mes completo)."""
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    cid = _parse_uuid(company_id)
    rid = _parse_uuid(run_id)
    ouid = _parse_uuid(org_unit_id)
    if not cid or not rid or not ouid:
        raise HTTPException(status_code=400, detail="company_id/run_id/org_unit_id inválido")

    if not can_see_company(db, current_user, cid):
        raise HTTPException(status_code=403, detail="Forbidden")

    run = db.get(Run, rid)
    if not run:
        raise HTTPException(status_code=404, detail="Run no existe")

    ou = db.get(OrgUnit, ouid)
    if not ou:
        raise HTTPException(status_code=404, detail="OU no existe")

    br = db.get(Branch, ou.branch_id)
    if not br or br.company_id != cid:
        raise HTTPException(status_code=403, detail="OU fuera de la empresa")

    try:
        y, m = _parse_month_yyyymm(month)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    employees = db.query(Employee).filter(Employee.org_unit_id == ouid).order_by(Employee.nombre.asc()).all()
    plan_map = _plan_map_for_run(db, run)
    dates = _month_dates(y, m)

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Plan"

    # Header
    ws["A1"].value = "OU"
    ws["B1"].value = f"{ou.org_unit_key} - {ou.name}"
    ws["A2"].value = "Mes"
    ws["B2"].value = month
    ws["A3"].value = "Run"
    ws["B3"].value = str(run.id)

    header_row = 5
    ws.cell(row=header_row, column=1, value="Nombre")
    ws.cell(row=header_row, column=2, value="Employee Key")

    # Days headers
    for j, d in enumerate(dates, start=3):
        ws.cell(row=header_row, column=j, value=d.day)

    # Styles
    thin = Side(style="thin", color="E6ECF5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="F1F5F9")
    header_font = Font(bold=True, color="0F172A")

    for c in range(1, 3 + len(dates)):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Category fills (suave)
    fill_livre = PatternFill("solid", fgColor="F3F6FA")
    fill_aus = PatternFill("solid", fgColor="FFF7ED")
    fill_am = PatternFill("solid", fgColor="EFF6FF")
    fill_pm = PatternFill("solid", fgColor="ECFDF5")
    fill_night = PatternFill("solid", fgColor="F5F3FF")

    def classify_for_fill(sid: str):
        s = (sid or "").upper()
        if s == "LIBRE":
            return fill_livre
        if s in {"LM", "VAC", "PERM", "LIC", "AUS", "INAS", "MED"}:
            return fill_aus
        if s.startswith("S_") and len(s.split("_")) >= 3:
            try:
                hhmm = s.split("_")[1]
                hh = int(hhmm[:2])
                mm = int(hhmm[2:])
                minutes = hh * 60 + mm
                if 5 * 60 <= minutes < 12 * 60:
                    return fill_am
                if 12 * 60 <= minutes < 19 * 60:
                    return fill_pm
                return fill_night
            except Exception:
                return None
        return None

    ws.freeze_panes = ws["C6"]

    # Data rows
    r0 = header_row + 1
    for i, e in enumerate(employees):
        r = r0 + i
        ws.cell(row=r, column=1, value=e.nombre)
        ws.cell(row=r, column=2, value=e.employee_key)
        ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
        ws.cell(row=r, column=2).alignment = Alignment(vertical="center")

        for j, d in enumerate(dates, start=3):
            sid = plan_map.get((e.employee_key, d.isoformat()), "LIBRE")
            val = _short_shift_id(sid)
            c = ws.cell(row=r, column=j, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
            fill = classify_for_fill(sid)
            if fill:
                c.fill = fill

        # borders left cells
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2).border = border

    # column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    for j in range(3, 3 + len(dates)):
        ws.column_dimensions[chr(64 + j) if j <= 26 else "A" + chr(64 + (j - 26))].width = 10

    # row heights
    ws.row_dimensions[header_row].height = 18
    for r in range(r0, r0 + len(employees)):
        ws.row_dimensions[r].height = 18

    # Print setup: landscape, fit to 2 pages wide (best effort)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 2
    ws.page_setup.fitToHeight = 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    from fastapi.responses import StreamingResponse

    filename = f"plan_ou_{ou.org_unit_key}_{month}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )



# ============================================================
# MÓDULO DE INFORMES GERENCIALES
# Genera un Excel con 4 hojas: Domingos, Horas, Cobertura, Cambios
# ============================================================

@router.get("/exports/informes.xlsx")
def ui_export_informes(
    request: Request,
    company_id: str,
    run_id: str,
    month: str,
    db: Session = Depends(get_db),
):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    cid = _parse_uuid(company_id)
    rid = _parse_uuid(run_id)
    if not cid or not rid:
        raise HTTPException(status_code=400, detail="Parámetros inválidos")
    if not can_see_company(db, current_user, cid):
        raise HTTPException(status_code=403, detail="Forbidden")

    run = db.get(Run, rid)
    if not run:
        raise HTTPException(status_code=404, detail="Run no existe")

    try:
        y, m = _parse_month_yyyymm(month)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Estilos comunes
    HDR_FILL  = PatternFill("solid", start_color="1E293B")
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    SUBHDR_FILL = PatternFill("solid", start_color="E2E8F0")
    SUBHDR_FONT = Font(bold=True, name="Arial", size=10)
    BODY_FONT = Font(name="Arial", size=10)
    OK_FILL   = PatternFill("solid", start_color="DCFCE7")
    WARN_FILL = PatternFill("solid", start_color="FEF9C3")
    ERR_FILL  = PatternFill("solid", start_color="FEE2E2")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def style_body(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = BODY_FONT
            cell.border = border

    # Cargar datos base
    plan_map  = _plan_map_for_run(db, run)
    employees = (
        db.query(Employee, OrgUnit)
        .join(OrgUnit, Employee.org_unit_id == OrgUnit.id)
        .join(Branch, OrgUnit.branch_id == Branch.id)
        .filter(Branch.company_id == cid)
        .order_by(OrgUnit.name.asc(), Employee.nombre.asc())
        .all()
    )
    dates   = _month_dates(y, m)
    template_path = _company_template_path(cid)
    cat = _load_catalogo_turnos(template_path)

    def shift_mins(sid):
        if not sid or sid in ("LIBRE", "SALIENTE"):
            return 0
        v = cat.get(sid)
        if not v or not v.get("inicio") or not v.get("fin"):
            # Intentar extraer desde nombre S_HHMM_HHMM_PAUSA
            parts = sid.split("_")
            if len(parts) >= 3 and parts[0] == "S":
                try:
                    ini = int(parts[1][:2]) * 60 + int(parts[1][2:])
                    fin = int(parts[2][:2]) * 60 + int(parts[2][2:])
                    pausa = int(parts[3]) if len(parts) >= 4 and parts[3].isdigit() else 0
                    return max(0, fin - ini - pausa)
                except Exception:
                    return 0
            return 0
        ini_m = int(v["inicio"][:2]) * 60 + int(v["inicio"][3:5])
        fin_m = int(v["fin"][:2]) * 60 + int(v["fin"][3:5])
        cruza = 1440 if v.get("cruza") == "SI" else 0
        pausa = int(sid.rsplit("_", 1)[-1]) if sid.startswith("S_") and sid.rsplit("_", 1)[-1].isdigit() else 0
        return max(0, fin_m - ini_m + cruza - pausa)

    wb = Workbook()

    # ─────────────────────────────────────────────
    # HOJA 1: INFORME DE DOMINGOS
    # ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Domingos"
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 20
    ws1.column_dimensions["F"].width = 20

    ws1["A1"] = f"Informe de Domingos — {month}"
    ws1["A1"].font = Font(bold=True, size=14, name="Arial")

    domingos_mes = [d for d in dates if d.weekday() == 6]
    sabados_mes  = [d for d in dates if d.weekday() == 5]

    headers1 = ["Colaborador", "OU", "Domingos libres", "Cumple (≥2)", "Domingos trabajados", "Sábados con dom libre"]
    for j, h in enumerate(headers1, 1):
        ws1.cell(row=3, column=j, value=h)
    style_header(ws1, 3, len(headers1))

    row = 4
    for emp, ou in employees:
        dom_libres    = sum(1 for d in domingos_mes if plan_map.get((emp.employee_key, d.isoformat()), "LIBRE") in ("LIBRE", "SALIENTE"))
        dom_trabajados = sum(1 for d in domingos_mes if plan_map.get((emp.employee_key, d.isoformat()), "LIBRE") not in ("LIBRE", "SALIENTE"))
        cumple = dom_libres >= 2

        # Sábados con domingo libre siguiente (para compensación comercio)
        sab_con_dom_libre = 0
        for sab in sabados_mes:
            dom_siguiente = sab + timedelta(days=1)
            if dom_siguiente in domingos_mes:
                sab_sid = plan_map.get((emp.employee_key, sab.isoformat()), "LIBRE")
                dom_sid = plan_map.get((emp.employee_key, dom_siguiente.isoformat()), "LIBRE")
                if sab_sid not in ("LIBRE", "SALIENTE") and dom_sid in ("LIBRE", "SALIENTE"):
                    sab_con_dom_libre += 1

        ws1.cell(row=row, column=1, value=emp.nombre)
        ws1.cell(row=row, column=2, value=ou.name)
        ws1.cell(row=row, column=3, value=dom_libres)
        ws1.cell(row=row, column=4, value="✅ Sí" if cumple else "❌ No")
        ws1.cell(row=row, column=5, value=dom_trabajados)
        ws1.cell(row=row, column=6, value=sab_con_dom_libre)
        style_body(ws1, row, len(headers1))
        fill = OK_FILL if cumple else ERR_FILL
        ws1.cell(row=row, column=4).fill = fill
        row += 1

    ws1.freeze_panes = "A4"

    # ─────────────────────────────────────────────
    # HOJA 2: HORAS ASIGNADAS VS CONTRATADAS
    # ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Horas por Semana")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18

    ws2["A1"] = f"Horas Asignadas vs Contratadas — {month}"
    ws2["A1"].font = Font(bold=True, size=14, name="Arial")

    # Calcular semanas con la misma lógica del solver:
    # semanas = número de domingos del mes, horizonte desde el lunes anterior al día 1
    n_semanas = _count_sundays_in_month(y, m)
    first_day = date(y, m, 1)
    horizon_start = first_day - timedelta(days=first_day.weekday())
    weeks = []
    for w in range(n_semanas):
        week_start = horizon_start + timedelta(weeks=w)
        week_days = [week_start + timedelta(days=i) for i in range(7)]
        weeks.append((week_start, week_days))

    headers2 = ["Colaborador", "OU", "Contrato (min/sem)"]
    for w, _ in weeks:
        headers2.append(f"Sem {w.strftime('%d/%m')}")
    headers2 += ["Total mes (min)", "Total mes (HH:MM)", "Desviación"]

    for j, h in enumerate(headers2, 1):
        ws2.cell(row=3, column=j, value=h)
        ws2.column_dimensions[get_column_letter(j)].width = max(14, len(h) + 2)
    style_header(ws2, 3, len(headers2))

    row = 4
    for emp, ou in employees:
        contrato = getattr(emp, "contrato_max_min_semana", 0) or 0
        sem_mins = []
        for w, week_days in weeks:
            sem_total = sum(
                shift_mins(plan_map.get((emp.employee_key, d.isoformat()), "LIBRE"))
                for d in week_days
            )
            sem_mins.append(sem_total)

        total_mes = sum(sem_mins)
        total_hhmm = f"{total_mes // 60:02d}:{total_mes % 60:02d}"
        contrato_mes = contrato * n_semanas
        desv = total_mes - contrato_mes

        ws2.cell(row=row, column=1, value=emp.nombre)
        ws2.cell(row=row, column=2, value=ou.name)
        ws2.cell(row=row, column=3, value=contrato)
        for i, sm in enumerate(sem_mins):
            ws2.cell(row=row, column=4 + i, value=sm)
        ws2.cell(row=row, column=4 + len(weeks),     value=total_mes)
        ws2.cell(row=row, column=4 + len(weeks) + 1, value=total_hhmm)
        ws2.cell(row=row, column=4 + len(weeks) + 2, value=desv)
        style_body(ws2, row, len(headers2))

        # Color desviación
        desv_cell = ws2.cell(row=row, column=4 + len(weeks) + 2)
        if abs(desv) < 60:
            desv_cell.fill = OK_FILL
        elif abs(desv) < 180:
            desv_cell.fill = WARN_FILL
        else:
            desv_cell.fill = ERR_FILL
        row += 1

    ws2.freeze_panes = "A4"

    # ─────────────────────────────────────────────
    # HOJA 3: COBERTURA VS DEMANDA
    # ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Cobertura")
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 18
    ws3.column_dimensions["E"].width = 14
    ws3.column_dimensions["F"].width = 14
    ws3.column_dimensions["G"].width = 14

    ws3["A1"] = f"Cobertura vs Demanda — {month}"
    ws3["A1"].font = Font(bold=True, size=14, name="Arial")

    headers3 = ["Fecha", "Día", "OU", "Cargo", "Requeridos", "Cubiertos", "Brecha"]
    for j, h in enumerate(headers3, 1):
        ws3.cell(row=3, column=j, value=h)
    style_header(ws3, 3, len(headers3))

    # Leer reporte_brechas.csv si existe
    brechas_path = Path(run.out_dir) / "reporte_brechas.csv"
    row = 4
    chart_data = {}  # fecha -> {req, cub}

    if brechas_path.exists():
        import csv as csv_mod
        with brechas_path.open("r", encoding="utf-8-sig", newline="") as f:
            for r in csv_mod.DictReader(f):
                fecha  = str(r.get("fecha", ""))[:10]
                dia    = str(r.get("dia_semana", ""))
                ou_id  = str(r.get("org_unit_id", ""))
                cargo  = str(r.get("cargo", ""))
                req    = int(float(r.get("requeridos_personas", 0) or 0))
                cub    = int(float(r.get("cubiertos_personas", 0) or 0))
                brecha = cub - req

                ws3.cell(row=row, column=1, value=fecha)
                ws3.cell(row=row, column=2, value=dia)
                ws3.cell(row=row, column=3, value=ou_id)
                ws3.cell(row=row, column=4, value=cargo)
                ws3.cell(row=row, column=5, value=req)
                ws3.cell(row=row, column=6, value=cub)
                ws3.cell(row=row, column=7, value=brecha)
                style_body(ws3, row, 7)

                brecha_cell = ws3.cell(row=row, column=7)
                if brecha == 0:
                    brecha_cell.fill = OK_FILL
                elif brecha > 0:
                    brecha_cell.fill = WARN_FILL
                else:
                    brecha_cell.fill = ERR_FILL

                # Acumular por fecha para gráfico
                if fecha not in chart_data:
                    chart_data[fecha] = {"req": 0, "cub": 0}
                chart_data[fecha]["req"] += req
                chart_data[fecha]["cub"] += cub
                row += 1
    else:
        ws3["A4"] = "No se encontró reporte_brechas.csv para este run"

    # Hoja auxiliar para datos del gráfico
    ws3g = wb.create_sheet("_chart_data")
    ws3g.sheet_state = "hidden"
    ws3g["A1"] = "Fecha"
    ws3g["B1"] = "Requeridos"
    ws3g["C1"] = "Cubiertos"
    for i, (fecha, vals) in enumerate(sorted(chart_data.items()), start=2):
        ws3g.cell(row=i, column=1, value=fecha)
        ws3g.cell(row=i, column=2, value=vals["req"])
        ws3g.cell(row=i, column=3, value=vals["cub"])

    if chart_data:
        n = len(chart_data) + 1
        chart = BarChart()
        chart.type = "col"
        chart.title = "Cobertura vs Demanda por Día"
        chart.y_axis.title = "Personas"
        chart.x_axis.title = "Fecha"
        chart.width  = 20
        chart.height = 12

        data_ref = Reference(ws3g, min_col=2, max_col=3, min_row=1, max_row=n)
        cats_ref = Reference(ws3g, min_col=1, min_row=2, max_row=n)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = "3B82F6"
        chart.series[1].graphicalProperties.solidFill = "22C55E"
        ws3.add_chart(chart, f"A{row + 2}")

    ws3.freeze_panes = "A4"

    # ─────────────────────────────────────────────
    # HOJA 4: HISTORIAL DE CAMBIOS DE TURNO
    # ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Cambios de Turno")
    ws4.column_dimensions["A"].width = 14
    ws4.column_dimensions["B"].width = 30
    ws4.column_dimensions["C"].width = 18
    ws4.column_dimensions["D"].width = 18
    ws4.column_dimensions["E"].width = 22
    ws4.column_dimensions["F"].width = 22

    ws4["A1"] = f"Historial de Cambios de Turno — {month}"
    ws4["A1"].font = Font(bold=True, size=14, name="Arial")

    headers4 = ["Fecha cambio", "Colaborador", "Fecha turno", "OU", "Turno original", "Turno nuevo"]
    for j, h in enumerate(headers4, 1):
        ws4.cell(row=3, column=j, value=h)
    style_header(ws4, 3, len(headers4))

    overrides = (
        db.query(PlanOverride)
        .filter(PlanOverride.run_id == rid)
        .order_by(PlanOverride.fecha.asc())
        .all()
    )

    # Construir base_map para saber el turno original
    plan_rows_base = _load_plan_rows_from_run(run)
    base_map = {}
    for r in plan_rows_base:
        eid = (r.get("employee_id") or "").strip()
        f   = (r.get("fecha") or "").strip()[:10]
        sid = (r.get("shift_id") or "LIBRE").strip()
        if eid and f:
            base_map[(eid, f)] = sid

    emp_by_key = {e.employee_key: e for e, _ in employees}
    ou_by_emp  = {e.employee_key: ou for e, ou in employees}

    row = 4
    for ov in overrides:
        # Filtrar solo el mes solicitado
        try:
            ov_date = date.fromisoformat(ov.fecha[:10])
            if ov_date.year != y or ov_date.month != m:
                continue
        except Exception:
            continue

        emp = emp_by_key.get(ov.employee_id)
        ou  = ou_by_emp.get(ov.employee_id)
        original = base_map.get((ov.employee_id, ov.fecha[:10]), "LIBRE")

        created_str = ""
        if hasattr(ov, "created_at") and ov.created_at:
            created_str = ov.created_at.strftime("%Y-%m-%d %H:%M") if hasattr(ov.created_at, "strftime") else str(ov.created_at)[:16]

        ws4.cell(row=row, column=1, value=created_str)
        ws4.cell(row=row, column=2, value=emp.nombre if emp else ov.employee_id)
        ws4.cell(row=row, column=3, value=ov.fecha[:10])
        ws4.cell(row=row, column=4, value=ou.name if ou else "")
        ws4.cell(row=row, column=5, value=original)
        ws4.cell(row=row, column=6, value=ov.shift_id)
        style_body(ws4, row, len(headers4))
        ws4.cell(row=row, column=6).fill = WARN_FILL
        row += 1

    if row == 4:
        ws4["A4"] = "Sin cambios manuales para este mes"

    ws4.freeze_panes = "A4"

    # Guardar y retornar
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    company = db.get(Company, cid)
    comp_name = company.name.replace(" ", "_") if company else "empresa"
    filename = f"informes_{comp_name}_{month}.xlsx"

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================================================
# RESPUESTA PÚBLICA A CAMBIO DE TURNO (sin login)
# ============================================================

@router.get("/override-response/{token}", response_class=HTMLResponse)
def ui_override_response(token: str, action: str = "accept", db: Session = Depends(get_db)):
    """Endpoint público — el colaborador acepta o rechaza desde el email."""
    from sqlalchemy import text as _text

    row = db.execute(
        _text("SELECT * FROM plan_override_responses WHERE token=:t"),
        {"t": token},
    ).mappings().first()

    if not row:
        return HTMLResponse("<h2>❌ Enlace inválido o expirado.</h2>", status_code=404)

    if row["estado"] != "pending":
        estado_label = "aceptado ✅" if row["estado"] == "accepted" else "rechazado ❌"
        return HTMLResponse(f"""
        <html><body style="font-family:Arial;text-align:center;padding:60px;">
        <h2>Este cambio ya fue {estado_label}</h2>
        <p>No es necesario hacer nada más.</p>
        </body></html>""")

    from datetime import datetime, timezone as tz
    _fl = row["fecha_limite"]
    if hasattr(_fl, "tzinfo") and _fl.tzinfo is None:
        _fl = _fl.replace(tzinfo=tz.utc)
    # Solo marcar vencido si pasaron las 24h + 1h de gracia
    _ahora = datetime.now(tz.utc)
    if _ahora > _fl + __import__("datetime").timedelta(hours=1):
        return HTMLResponse("""
        <html><body style="font-family:Arial;text-align:center;padding:60px;">
        <h2>⏰ El plazo para responder ha vencido</h2>
        <p>El cambio fue aceptado automáticamente.</p>
        </body></html>""")

    nuevo_estado = "accepted" if action == "accept" else "rejected"
    emoji = "✅" if nuevo_estado == "accepted" else "❌"
    label = "aceptado" if nuevo_estado == "accepted" else "rechazado"

    db.execute(
        _text("""
            UPDATE plan_override_responses
            SET estado=:estado, fecha_respuesta=NOW()
            WHERE token=:t
        """),
        {"estado": nuevo_estado, "t": token},
    )
    db.commit()

    # Notificar al supervisor
    try:
        celery_app.send_task("notify_supervisor_override_response", args=[token])
    except Exception:
        pass

    return HTMLResponse(f"""
    <html>
    <head><meta charset="utf-8"></head>
    <body style="font-family:Arial,sans-serif;text-align:center;padding:60px;background:#f8fafc;">
      <div style="max-width:480px;margin:0 auto;background:#fff;padding:40px;border-radius:12px;box-shadow:0 4px 20px rgba(0,0,0,.1);">
        <div style="font-size:56px;margin-bottom:16px;">{emoji}</div>
        <h2 style="margin:0 0 8px 0;">Cambio {label}</h2>
        <p style="color:#64748b;">Turno del día <b>{row["fecha"]}</b>: <code>{row["shift_id_new"]}</code></p>
        <p style="color:#94a3b8;font-size:13px;margin-top:24px;">Puedes cerrar esta ventana.</p>
      </div>
    </body>
    </html>""")


# ============================================================
# ALERTAS DE OVERRIDES PENDIENTES (para la plataforma)
# ============================================================

@router.get("/override-alerts", response_class=HTMLResponse)
def ui_override_alerts(request: Request, db: Session = Depends(get_db)):
    """Vista de alertas de cambios de turno pendientes de respuesta."""
    from sqlalchemy import text as _text

    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    # Buscar overrides pendientes de empresas que puede ver el usuario
    companies = filter_companies(db, current_user)
    company_ids = [str(c.id) for c in companies]

    if not company_ids:
        alerts = []
    else:
        placeholders = ",".join(f"'{cid}'" for cid in company_ids)
        alerts = db.execute(
            _text(f"""
                SELECT r.*, e.nombre as emp_nombre
                FROM plan_override_responses r
                LEFT JOIN employees e ON e.employee_key = r.employee_id
                WHERE r.company_id IN ({placeholders})
                  AND r.created_at > NOW() - INTERVAL '30 days'
                ORDER BY r.created_at DESC
                LIMIT 100
            """),
        ).mappings().fetchall()

    return TEMPLATES.TemplateResponse(
        "override_alerts.html",
        _enrich(db, request, {
            "request": request,
            "alerts": alerts,
            "ok":  request.query_params.get("ok"),
            "err": request.query_params.get("err"),
        }),
    )

# ============================================================
# PUBLICAR TURNOS (Run + Mes + OU) via Resend + Celery
# - No elimina rutas existentes: solo agrega nuevas rutas /ui/publish...
# - Requiere tablas: run_publications, run_publication_recipients, y employees.email
# ============================================================

from sqlalchemy import text as _sql_text
import uuid as _uuid

def build_employee_calendar_pdf_bytes(*, db: Session, company_id: str, run_id: str, month: str, employee_key: str) -> bytes:
    """Genera PDF mensual para un employee_key, mostrando SOLO fechas que existan en el output del solver para ese employee_key.
    No inventa LIBRE. Si el solver no trae la fecha, la celda queda en blanco.
    """
    plan_map = _plan_map_for_run(db, _parse_uuid(run_id))

    emp_dates = sorted([d for (ek, d) in plan_map.keys() if ek == employee_key])
    emp_dates_set = set(emp_dates)

    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from io import BytesIO

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=16, leading=18, spaceAfter=6)
    base = ParagraphStyle("base", parent=styles["BodyText"], fontSize=9, leading=11)
    muted = ParagraphStyle("muted", parent=styles["BodyText"], fontSize=8, leading=10, textColor=colors.grey)

    story = []
    story.append(Paragraph(f"Turnos — {month} — {employee_key}", h1))
    story.append(Spacer(1, 8))

    if not emp_dates:
        story.append(Paragraph("No hay turnos para este colaborador en este run/mes.", muted))
        doc.build(story)
        return buf.getvalue()

    import calendar
    y, m = [int(x) for x in month.split("-")]
    cal = calendar.Calendar(firstweekday=0)  # lunes
    weeks = cal.monthdatescalendar(y, m)

    header = ["Lun","Mar","Mié","Jue","Vie","Sáb","Dom"]
    data = [header]

    for w in weeks:
        row = []
        for d in w:
            iso = d.isoformat()
            if iso not in emp_dates_set:
                row.append(Paragraph("&nbsp;", muted))
            else:
                sid = plan_map.get((employee_key, iso), "") or ""
                txt = f"<b>{d.day}</b><br/>{sid}" if sid else f"<b>{d.day}</b>"
                row.append(Paragraph(txt, base))
        data.append(row)

    t = Table(data, colWidths=[(A4[0]-36)/7]*7)
    t.setStyle(TableStyle([
        ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#E6ECF5")),
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#F8FAFF")),
        ("ALIGN",(0,0),(-1,0),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,0),9),
    ]))

    story.append(t)
    doc.build(story)
    return buf.getvalue()


@router.get("/publish", response_class=HTMLResponse)
def ui_publish(request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    q_company_id = request.query_params.get("company_id") or (str(ctx.company_id) if ctx.company_id else "")
    q_run_id = request.query_params.get("run_id") or ""
    month = request.query_params.get("month") or ""
    q_ou_id = request.query_params.get("org_unit_id") or ""
    mode = request.query_params.get("mode") or "all"
    test_email = request.query_params.get("test_email") or ""

    companies = filter_companies(db, current_user)

    runs = []
    org_units = []
    preview_count = 0
    missing_email_count = 0

    if q_company_id:
        cid = _parse_uuid(q_company_id)
        if cid and can_see_company(db, current_user, cid):
            runs = db.query(Run).order_by(Run.created_at.desc()).limit(50).all()
            org_units = (
                db.query(OrgUnit)
                .join(Branch, OrgUnit.branch_id == Branch.id)
                .filter(Branch.company_id == cid)
                .order_by(OrgUnit.org_unit_key.asc())
                .all()
            )
            if q_ou_id:
                ouid = _parse_uuid(q_ou_id)
                if ouid:
                    emps = db.query(Employee).filter(Employee.org_unit_id == ouid).all()
                    with_email = [e for e in emps if getattr(e, "email", None)]
                    preview_count = len(with_email) if mode != "test" else (1 if test_email else 0)
                    missing_email_count = len(emps) - len(with_email)

    # Si las tablas no existen, mostrar lista vacía (evita 500)
    try:
        recent_publications = db.execute(_sql_text("SELECT * FROM run_publications ORDER BY created_at DESC LIMIT 20")).mappings().all()
    except Exception:
        recent_publications = []

    return TEMPLATES.TemplateResponse(
        "publish.html",
        _enrich(
            db,
            request,
            {
                "request": request,
                "companies": companies,
                "runs": runs,
                "org_units": org_units,
                "selected_company_id": q_company_id,
                "selected_run_id": q_run_id,
                "selected_org_unit_id": q_ou_id,
                "month": month,
                "mode": mode,
                "test_email": test_email,
                "preview_count": preview_count,
                "missing_email_count": missing_email_count,
                "recent_publications": recent_publications,
                "ok": request.query_params.get("ok"),
                "err": request.query_params.get("err"),
            },
        ),
    )


@router.post("/publish")
def ui_publish_create(
    request: Request,
    company_id: str = Form(...),
    run_id: str = Form(...),
    month: str = Form(...),
    org_unit_id: str = Form(...),
    mode: str = Form("all"),
    test_email: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    cid = _parse_uuid(company_id)
    rid = _parse_uuid(run_id)
    ouid = _parse_uuid(org_unit_id)

    if not cid or not rid or not ouid:
        return _redirect("/ui/publish?err=Datos+inválidos")

    if not can_see_company(db, current_user, cid):
        raise HTTPException(403, "Forbidden")

    if mode == "test" and not test_email:
        return _redirect(f"/ui/publish?company_id={company_id}&run_id={run_id}&month={month}&org_unit_id={org_unit_id}&mode=test&err=Falta+correo+de+prueba")

    pub_id = str(_uuid.uuid4())

    # INSERT incluye org_unit_id (NOT NULL en tabla)
    db.execute(
        _sql_text("""
            INSERT INTO run_publications (id, run_id, company_id, month, org_unit_id, mode, test_email, status, created_by)
            VALUES (:id, :run_id, :company_id, :month, :org_unit_id, :mode, :test_email, 'queued', :created_by)
        """),
        {
            "id": pub_id,
            "run_id": str(rid),
            "company_id": str(cid),
            "month": month,
            "org_unit_id": str(ouid),
            "mode": mode,
            "test_email": test_email or None,
            "created_by": str(current_user.id) if getattr(current_user, "id", None) else None,
        },
    )

    emps = db.query(Employee).filter(Employee.org_unit_id == ouid).all()

    rows = []
    if mode == "test":
        # ✅ En modo test enviamos SOLO a test_email, aunque los empleados no tengan email.
        # Elegimos el primer colaborador con employee_key dentro de la OU (solo para "calzar" el PDF / calendario).
        chosen = None
        for e in emps:
            ek = getattr(e, "employee_key", None)
            if ek:
                chosen = ek
                break
        if not chosen:
            db.rollback()
            return _redirect(f"/ui/publish?company_id={company_id}&run_id={run_id}&month={month}&org_unit_id={org_unit_id}&mode=test&err=OU+sin+colaboradores")
        rows.append((str(_uuid.uuid4()), pub_id, chosen, test_email, "queued"))
    else:
        for e in emps:
            ek = getattr(e, "employee_key", None)
            email = getattr(e, "email", None)
            if not ek:
                continue
            status = "queued" if email else "no_email"
            rows.append((str(_uuid.uuid4()), pub_id, ek, email, status))

    for (rid2, pid, ek, email, status) in rows:
        db.execute(
            _sql_text("""
                INSERT INTO run_publication_recipients (id, publication_id, employee_key, email, status)
                VALUES (:id, :pid, :ek, :email, :status)
            """),
            {"id": rid2, "pid": pid, "ek": ek, "email": email, "status": status},
        )

    db.commit()

    # Encolar (solo queued). En modo test, solo 1 envío.
    from worker.tasks import publish_send_employee_pdf
    for (_, _, ek, email, status) in rows:
        if status != "queued":
            continue
        publish_send_employee_pdf.delay(pub_id, ek)
        if mode == "test":
            break

    return _redirect(f"/ui/publish/{pub_id}")


@router.get("/publish/{publication_id}", response_class=HTMLResponse)
def ui_publish_detail(publication_id: str, request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    pub = db.execute(_sql_text("SELECT * FROM run_publications WHERE id=:id"), {"id": publication_id}).mappings().first()
    if not pub:
        raise HTTPException(404, "Publicación no encontrada")

    cid = _parse_uuid(str(pub["company_id"]))
    if cid and not can_see_company(db, current_user, cid):
        raise HTTPException(403, "Forbidden")

    recipients = db.execute(
        _sql_text("""
            SELECT employee_key, email, status, sent_at, last_error
            FROM run_publication_recipients
            WHERE publication_id=:id
            ORDER BY status, employee_key
        """),
        {"id": publication_id},
    ).mappings().all()

    def _count(st): return sum(1 for r in recipients if r["status"] == st)

    counts = {"sent": _count("sent"), "queued": _count("queued"), "failed": _count("failed"), "no_email": _count("no_email")}

    return TEMPLATES.TemplateResponse(
        "publish_detail.html",
        _enrich(db, request, {"request": request, "pub": pub, "recipients": recipients, "counts": counts}),
    )


@router.post("/publish/{publication_id}/retry")
def ui_publish_retry(publication_id: str, request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    pub = db.execute(_sql_text("SELECT * FROM run_publications WHERE id=:id"), {"id": publication_id}).mappings().first()
    if not pub:
        return _redirect("/ui/publish?err=Publicación+no+existe")

    cid = _parse_uuid(str(pub["company_id"]))
    if cid and not can_see_company(db, current_user, cid):
        raise HTTPException(403, "Forbidden")

    failed = db.execute(
        _sql_text("""
            SELECT employee_key
            FROM run_publication_recipients
            WHERE publication_id=:id AND status='failed'
        """),
        {"id": publication_id},
    ).mappings().all()

    from worker.tasks import publish_send_employee_pdf
    for r in failed:
        publish_send_employee_pdf.delay(publication_id, r["employee_key"])

    return _redirect(f"/ui/publish/{publication_id}")



@router.get("/holidays", response_class=HTMLResponse)
def ui_holidays(request: Request, db: Session = Depends(get_db)):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")
    companies = filter_companies(db, current_user)
    q_company_id = request.query_params.get("company_id") or (str(ctx.company_id) if ctx.company_id else "")
    q_branch_id  = request.query_params.get("branch_id") or ""
    try:
        selected_year = int(request.query_params.get("year") or datetime.now(timezone.utc).year)
    except ValueError:
        selected_year = datetime.now(timezone.utc).year
    years = list(range(selected_year - 1, selected_year + 3))
    branches = []
    holidays_out = []
    if q_company_id:
        cid = _parse_uuid(q_company_id)
        if cid and can_see_company(db, current_user, cid):
            branches = (
                db.query(Branch)
                .filter(Branch.company_id == cid)
                .order_by(Branch.code.asc())
                .all()
            )
    if q_branch_id:
        bid = _parse_uuid(q_branch_id)
        if bid:
            raw = (
                db.query(Holiday)
                .filter(
                    Holiday.branch_id == bid,
                    Holiday.fecha >= date(selected_year, 1, 1),
                    Holiday.fecha <= date(selected_year, 12, 31),
                )
                .order_by(Holiday.fecha.asc())
                .all()
            )
            for h in raw:
                d = h.fecha if isinstance(h.fecha, date) else date.fromisoformat(str(h.fecha))
                holidays_out.append({
                    "id": h.id,
                    "fecha": d.isoformat(),
                    "dow": _DOW_ES[d.weekday()],
                    "nombre": h.nombre,
                    "irrenunciable": h.irrenunciable,
                })
    return TEMPLATES.TemplateResponse(
        "holidays.html",
        _enrich(db, request, {
            "request": request,
            "companies": companies,
            "branches": branches,
            "holidays": holidays_out,
            "selected_company_id": q_company_id,
            "selected_branch_id": q_branch_id,
            "selected_year": selected_year,
            "years": years,
            "err": request.query_params.get("err"),
            "ok":  request.query_params.get("ok"),
        }),
    )


@router.post("/holidays")
def ui_holidays_create(
    request: Request,
    company_id: str = Form(...),
    branch_id: str = Form(...),
    fecha: str = Form(...),
    nombre: str = Form(...),
    irrenunciable: str = Form(""),
    year: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")
    cid = _parse_uuid(company_id)
    bid = _parse_uuid(branch_id)
    if not cid or not bid:
        return _redirect(f"/ui/holidays?company_id={company_id}&branch_id={branch_id}&year={year}&err=Datos+invalidos")
    if not can_see_company(db, current_user, cid):
        raise HTTPException(403, "Forbidden")
    try:
        d = date.fromisoformat(fecha.strip())
    except ValueError:
        return _redirect(f"/ui/holidays?company_id={company_id}&branch_id={branch_id}&year={year}&err=Fecha+invalida")
    exists = db.query(Holiday).filter(Holiday.branch_id == bid, Holiday.fecha == d).first()
    if exists:
        return _redirect(
            f"/ui/holidays?company_id={company_id}&branch_id={branch_id}&year={year}"
            f"&err=Ya+existe+un+feriado+el+{d.isoformat()}+para+esta+sucursal"
        )
    db.add(Holiday(branch_id=bid, fecha=d, nombre=nombre.strip(), irrenunciable=bool(irrenunciable)))
    db.commit()
    y = year or str(d.year)
    return _redirect(f"/ui/holidays?company_id={company_id}&branch_id={branch_id}&year={y}&ok=Feriado+agregado")


@router.post("/holidays/{holiday_id}/delete")
def ui_holidays_delete(
    holiday_id: str,
    request: Request,
    company_id: str = Form(""),
    branch_id: str = Form(""),
    year: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")
    hid = _parse_uuid(holiday_id)
    if not hid:
        return _redirect(f"/ui/holidays?company_id={company_id}&branch_id={branch_id}&year={year}&err=ID+invalido")
    h = db.get(Holiday, hid)
    if not h:
        return _redirect(f"/ui/holidays?company_id={company_id}&branch_id={branch_id}&year={year}&err=Feriado+no+existe")
    branch = db.get(Branch, h.branch_id)
    if branch:
        cid = _parse_uuid(company_id) or branch.company_id
        if not can_see_company(db, current_user, cid):
            raise HTTPException(403, "Forbidden")
    db.delete(h)
    db.commit()
    return _redirect(f"/ui/holidays?company_id={company_id}&branch_id={branch_id}&year={year}&ok=Feriado+eliminado")


# =========================
# IMPORTAR FERIADOS CHILE
# =========================

@router.post("/holidays/sync-cl")
def ui_holidays_sync_cl(
    request: Request,
    year: str = Form(...),
    db: Session = Depends(get_db),
):
    """Descarga feriados de Chile desde apis.digital.gob.cl y los guarda en holidays_cl."""
    import urllib.request as urlreq
    import json as _json

    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")

    try:
        yr = int(year)
    except ValueError:
        return _redirect(f"/ui/holidays?year={year}&err=Año+invalido")

    IRRENUNCIABLES = {
        (1, 1), (1, 5), (9, 18), (9, 19), (12, 25),
    }

    try:
        url = f"https://apis.digital.gob.cl/fl/feriados/{yr}"
        req = urlreq.Request(url, headers={"User-Agent": "vigatec-runner/1.0"})
        with urlreq.urlopen(req, timeout=10) as resp:
            data = _json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        return _redirect(f"/ui/holidays?year={year}&err=Error+descargando+feriados:+{str(e)[:80]}")

    added = 0
    skipped = 0
    for item in data:
        try:
            fecha_str = item.get("fecha") or item.get("date") or ""
            nombre = item.get("nombre") or item.get("name") or ""
            d = date.fromisoformat(fecha_str[:10])
            irrenunciable = (d.month, d.day) in IRRENUNCIABLES
            exists = db.query(HolidayCl).filter(
                HolidayCl.fecha == d,
                HolidayCl.region == None,
            ).first()
            if exists:
                skipped += 1
                continue
            db.add(HolidayCl(
                fecha=d,
                nombre=nombre,
                irrenunciable=irrenunciable,
                nacional=True,
                region=None,
            ))
            added += 1
        except Exception:
            continue
    db.commit()
    msg = f"Catálogo+actualizado:+{added}+feriados+descargados"
    if skipped:
        msg += f"+({skipped}+ya+existían)"
    return _redirect(f"/ui/holidays?year={year}&ok={msg}")


@router.post("/holidays/import-cl")
def ui_holidays_import_cl(
    request: Request,
    company_id: str = Form(...),
    branch_id: str = Form(...),
    year: str = Form(...),
    db: Session = Depends(get_db),
):
    ctx = _load_ctx(request)
    current_user = get_current_user(db, ctx.user_id)
    if not current_user:
        return _redirect("/ui/system/users")
    cid = _parse_uuid(company_id)
    bid = _parse_uuid(branch_id)
    if not cid or not bid:
        return _redirect(f"/ui/holidays?company_id={company_id}&branch_id={branch_id}&year={year}&err=Datos+invalidos")
    if not can_see_company(db, current_user, cid):
        raise HTTPException(403, "Forbidden")
    try:
        yr = int(year)
    except ValueError:
        return _redirect(f"/ui/holidays?company_id={company_id}&branch_id={branch_id}&year={year}&err=Año+invalido")

    # Cargar desde catalogo nacional
    nacionales = (
        db.query(HolidayCl)
        .filter(
            HolidayCl.nacional == True,
            HolidayCl.fecha >= date(yr, 1, 1),
            HolidayCl.fecha <= date(yr, 12, 31),
        )
        .all()
    )
    added = 0
    skipped = 0
    for h in nacionales:
        exists = db.query(Holiday).filter(Holiday.branch_id == bid, Holiday.fecha == h.fecha).first()
        if exists:
            skipped += 1
            continue
        db.add(Holiday(
            branch_id=bid,
            fecha=h.fecha,
            nombre=h.nombre,
            irrenunciable=h.irrenunciable,
        ))
        added += 1
    db.commit()
    msg = f"Importados+{added}+feriados"
    if skipped:
        msg += f"+({skipped}+ya+existian)"
    return _redirect(f"/ui/holidays?company_id={company_id}&branch_id={branch_id}&year={year}&ok={msg}")