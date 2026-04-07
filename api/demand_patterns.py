from __future__ import annotations

import json
from datetime import datetime, timezone, time
from pathlib import Path
from typing import Any, Optional, List, Dict

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field, field_validator
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from api.db import get_db
from api.models import Run
from api.storage import safe_resolve_under


router = APIRouter(tags=["planning"])


DOW_ALLOWED = {"LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"}


def _iso_utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _as_str(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip()


def _normalize_dow(x: Any) -> str:
    s = _as_str(x).upper()
    s = s.replace("MIÉ", "MIE").replace("MIÉ.", "MIE")
    s = s.replace("SÁB", "SAB").replace("SÁB.", "SAB")
    s = s.replace("DOM.", "DOM").replace("LUN.", "LUN").replace("MAR.", "MAR").replace("MIE.", "MIE").replace("JUE.", "JUE").replace("VIE.", "VIE").replace("SAB.", "SAB")
    if s not in DOW_ALLOWED:
        raise ValueError(f"dia_semana inválido: {s} (usa LUN..DOM)")
    return s


def _normalize_time(x: Any) -> str:
    """
    Normaliza a HH:MM:SS.
    Acepta:
      - '07:30' o '07:30:00'
      - datetime.time
      - pandas Timestamp/datetime
    """
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    if isinstance(x, time):
        return x.strftime("%H:%M:%S")
    if isinstance(x, datetime):
        return x.time().strftime("%H:%M:%S")
    # pandas Timestamp
    if hasattr(x, "to_pydatetime"):
        try:
            dt = x.to_pydatetime()
            if isinstance(dt, datetime):
                return dt.time().strftime("%H:%M:%S")
        except Exception:
            pass

    s = _as_str(x)
    if not s:
        return ""
    # permitir "07:30" y convertir a "07:30:00"
    if len(s) == 5 and s[2] == ":":
        return f"{s}:00"
    # si ya viene HH:MM:SS lo devolvemos
    if len(s) == 8 and s[2] == ":" and s[5] == ":":
        return s
    # algunos casos vienen "07:30:00.000"
    if len(s) > 8 and s[2] == ":":
        return s[:8]
    raise ValueError(f"hora inválida: {s} (esperado HH:MM o HH:MM:SS)")


class DemandPatternRow(BaseModel):
    org_unit_id: str
    cargo_id: str
    dia_semana: str
    inicio: str
    fin: str
    requeridos: int = Field(ge=0)

    # opcional (en tu sheet existe "cargo")
    cargo: Optional[str] = None

    @field_validator("org_unit_id", "cargo_id")
    @classmethod
    def _not_empty(cls, v: str) -> str:
        v = _as_str(v)
        if not v:
            raise ValueError("no puede ser vacío")
        return v

    @field_validator("dia_semana")
    @classmethod
    def _valid_dow(cls, v: str) -> str:
        return _normalize_dow(v)

    @field_validator("inicio", "fin")
    @classmethod
    def _valid_time(cls, v: str) -> str:
        return _normalize_time(v)


class DemandPatternGetResponse(BaseModel):
    run_id: str
    source: str  # "json" | "case"
    count: int
    generated_at_utc: str
    rows: List[DemandPatternRow]


class DemandPatternPutRequest(BaseModel):
    rows: List[DemandPatternRow]
    write_back_case: bool = True
    rebuild_demanda_unidad: bool = True


class DemandPatternPutResponse(BaseModel):
    run_id: str
    saved_rows: int
    wrote_case: bool
    rebuilt_demanda_unidad: bool
    updated_at_utc: str
    overlay_path: str


def _run_base_dirs(run: Run) -> Dict[str, Path]:
    """
    Estructura esperada:
      /app/storage/runs/<run_id>/
        input/case.xlsx
        out/...
        logs/...
    """
    out_dir = Path(run.out_dir).resolve()
    run_base = out_dir.parent.resolve()
    input_dir = (run_base / "input").resolve()
    return {"run_base": run_base, "input_dir": input_dir, "out_dir": out_dir}


def _load_run(db: Session, run_id: str) -> Run:
    try:
        rid = run_id
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid run_id")

    run: Run | None = db.get(Run, rid)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    return run


def _read_patterns_from_case(case_path: Path) -> List[DemandPatternRow]:
    if not case_path.exists():
        raise HTTPException(status_code=404, detail="case.xlsx not found for this run")

    try:
        df = pd.read_excel(case_path, sheet_name="NecesidadMinimos")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cannot read NecesidadMinimos: {e}")

    # normalizar columnas (ignorar columnas vacías extra)
    cols = {c: c.strip() if isinstance(c, str) else c for c in df.columns}
    df = df.rename(columns=cols)

    needed = ["org_unit_id", "cargo_id", "dia_semana", "inicio", "fin", "requeridos"]
    for c in needed:
        if c not in df.columns:
            raise HTTPException(status_code=500, detail=f"NecesidadMinimos missing column '{c}'")

    # Filtrar filas realmente llenas (evitar filas en blanco)
    df2 = df.copy()
    df2["org_unit_id"] = df2["org_unit_id"].astype(str).where(df2["org_unit_id"].notna(), "")
    df2["cargo_id"] = df2["cargo_id"].astype(str).where(df2["cargo_id"].notna(), "")
    df2["dia_semana"] = df2["dia_semana"].astype(str).where(df2["dia_semana"].notna(), "")
    df2 = df2[(df2["org_unit_id"].str.strip() != "") & (df2["cargo_id"].str.strip() != "") & (df2["dia_semana"].str.strip() != "")]

    rows: List[DemandPatternRow] = []
    for _, r in df2.iterrows():
        payload = {
            "org_unit_id": _as_str(r.get("org_unit_id")),
            "cargo_id": _as_str(r.get("cargo_id")),
            "dia_semana": _normalize_dow(r.get("dia_semana")),
            "inicio": _normalize_time(r.get("inicio")),
            "fin": _normalize_time(r.get("fin")),
            "requeridos": int(r.get("requeridos")) if not pd.isna(r.get("requeridos")) else 0,
            "cargo": _as_str(r.get("cargo")) if "cargo" in df2.columns else None,
        }
        rows.append(DemandPatternRow(**payload))

    return rows


def _ensure_sheet(wb, name: str) -> Worksheet:
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(title=name)


def _clear_worksheet(ws: Worksheet):
    # Borra todo y deja hoja limpia
    ws.delete_rows(1, ws.max_row if ws.max_row > 0 else 1)


def _write_table(ws: Worksheet, headers: List[str], data_rows: List[List[Any]]):
    _clear_worksheet(ws)
    ws.append(headers)
    for row in data_rows:
        ws.append(row)


def _write_patterns_to_case(case_path: Path, rows: List[DemandPatternRow], rebuild_demanda_unidad: bool):
    wb = load_workbook(case_path)

    # NecesidadMinimos
    ws_nm = _ensure_sheet(wb, "NecesidadMinimos")
    headers_nm = ["org_unit_id", "cargo_id", "cargo", "dia_semana", "inicio", "fin", "requeridos"]
    data_nm: List[List[Any]] = []
    for r in rows:
        data_nm.append([
            r.org_unit_id,
            r.cargo_id,
            r.cargo or "",
            r.dia_semana,
            r.inicio,
            r.fin,
            int(r.requeridos),
        ])
    _write_table(ws_nm, headers_nm, data_nm)

    # DemandaUnidad (opcional): agregamos sumando por unidad+día+tramo
    if rebuild_demanda_unidad:
        ws_du = _ensure_sheet(wb, "DemandaUnidad")
        headers_du = ["org_unit_id", "dia_semana", "inicio", "fin", "requeridos"]
        # agrupar
        tmp = pd.DataFrame([r.model_dump() for r in rows])
        if not tmp.empty:
            grp = (
                tmp.groupby(["org_unit_id", "dia_semana", "inicio", "fin"], as_index=False)["requeridos"]
                .sum()
                .sort_values(["org_unit_id", "dia_semana", "inicio", "fin"])
            )
            data_du = grp[headers_du].values.tolist()
        else:
            data_du = []
        _write_table(ws_du, headers_du, data_du)

    wb.save(case_path)


@router.get("/runs/{run_id}/demand-patterns", response_model=DemandPatternGetResponse)
def get_demand_patterns(run_id: str, db: Session = Depends(get_db)):
    run = _load_run(db, run_id)
    dirs = _run_base_dirs(run)
    input_dir = dirs["input_dir"]

    overlay_candidate = (input_dir / "demand_patterns.json")
    overlay_path = safe_resolve_under(input_dir, overlay_candidate)

    if overlay_path.exists():
        try:
            data = json.loads(overlay_path.read_text(encoding="utf-8"))
            raw_rows = data.get("rows", [])
            rows = [DemandPatternRow(**rr) for rr in raw_rows]
            return DemandPatternGetResponse(
                run_id=str(run.id),
                source="json",
                count=len(rows),
                generated_at_utc=_iso_utc_now(),
                rows=rows,
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Invalid overlay demand_patterns.json: {e}")

    # fallback: leer desde case.xlsx
    case_path = Path(run.case_path)
    rows = _read_patterns_from_case(case_path)

    return DemandPatternGetResponse(
        run_id=str(run.id),
        source="case",
        count=len(rows),
        generated_at_utc=_iso_utc_now(),
        rows=rows,
    )


@router.put("/runs/{run_id}/demand-patterns", response_model=DemandPatternPutResponse)
def put_demand_patterns(run_id: str, req: DemandPatternPutRequest, db: Session = Depends(get_db)):
    run = _load_run(db, run_id)
    dirs = _run_base_dirs(run)
    input_dir = dirs["input_dir"]

    # Guardar overlay JSON
    overlay_candidate = (input_dir / "demand_patterns.json")
    overlay_path = safe_resolve_under(input_dir, overlay_candidate)

    payload = {
        "run_id": str(run.id),
        "updated_at_utc": _iso_utc_now(),
        "rows": [r.model_dump() for r in req.rows],
    }
    overlay_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    wrote_case = False
    if req.write_back_case:
        try:
            case_path = Path(run.case_path)
            _write_patterns_to_case(case_path, req.rows, rebuild_demanda_unidad=req.rebuild_demanda_unidad)
            wrote_case = True
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed writing back to case.xlsx: {e}")

    return DemandPatternPutResponse(
        run_id=str(run.id),
        saved_rows=len(req.rows),
        wrote_case=wrote_case,
        rebuilt_demanda_unidad=req.rebuild_demanda_unidad and wrote_case,
        updated_at_utc=_iso_utc_now(),
        overlay_path=str(overlay_path),
    )


@router.post("/runs/{run_id}/demand-patterns/reset", response_model=DemandPatternPutResponse)
def reset_demand_patterns(run_id: str, db: Session = Depends(get_db)):
    """
    Borra el overlay JSON para volver a leer desde el case.
    (No toca el case.xlsx)
    """
    run = _load_run(db, run_id)
    dirs = _run_base_dirs(run)
    input_dir = dirs["input_dir"]

    overlay_candidate = (input_dir / "demand_patterns.json")
    overlay_path = safe_resolve_under(input_dir, overlay_candidate)

    if overlay_path.exists():
        overlay_path.unlink()

    # devolver el estado actual desde case
    rows = _read_patterns_from_case(Path(run.case_path))

    # (Opcional) dejar un overlay "limpio" igual, para que la UI tenga fuente estable
    payload = {
        "run_id": str(run.id),
        "updated_at_utc": _iso_utc_now(),
        "rows": [r.model_dump() for r in rows],
    }
    overlay_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    return DemandPatternPutResponse(
        run_id=str(run.id),
        saved_rows=len(rows),
        wrote_case=False,
        rebuilt_demanda_unidad=False,
        updated_at_utc=_iso_utc_now(),
        overlay_path=str(overlay_path),
    )
