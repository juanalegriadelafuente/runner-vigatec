from __future__ import annotations

import uuid
from pathlib import Path
from typing import Dict, List

from sqlalchemy.orm import Session

from api.vocab_models import SolverVocabItem

CAT_AUSENTISMO = "AUSENTISMO_CODE"
CAT_RESTR_TIPO = "RESTRICCION_TIPO"
CAT_JORNADA = "JORNADA_ID"
CAT_TURNO = "TURNO_SHIFT_ID"  # opcional para dropdowns

DEFAULT_RESTR_TIPOS = [
    ("DIA_LIBRE_FIJO", "Día libre fijo"),
    ("REGLA_APERTURA_CARGO", "Regla apertura cargo"),
    ("REGLA_CIERRE_CARGO", "Regla cierre cargo"),
]


def _upsert(db: Session, company_id: uuid.UUID, category: str, value: str, label: str | None = None) -> None:
    value = (value or "").strip()
    if not value:
        return

    row = (
        db.query(SolverVocabItem)
        .filter(
            SolverVocabItem.company_id == company_id,
            SolverVocabItem.category == category,
            SolverVocabItem.value == value,
        )
        .first()
    )

    if row:
        if label is not None and label != "":
            row.label = label
        row.active = True
        db.add(row)
        return

    db.add(SolverVocabItem(company_id=company_id, category=category, value=value, label=label, active=True))


def seed_company_vocab(db: Session, company_id: uuid.UUID, template_path: Path | None) -> None:
    """
    Si hay template:
      - Jornadas: lee hoja Jornadas/jornada_id
      - Ausentismos: shift_id tipo NO_TRABAJADO de CatalogoTurnos (excluye LIBRE)
      - Turnos: shift_id tipo TRABAJADO de CatalogoTurnos (opcional)
    Siempre:
      - Tipos de restricción base
      - Ausentismos mínimos LM/VAC
    """
    # Tipos de restricción por defecto
    for value, label in DEFAULT_RESTR_TIPOS:
        _upsert(db, company_id, CAT_RESTR_TIPO, value, label)

    # Ausentismos mínimos
    _upsert(db, company_id, CAT_AUSENTISMO, "LM", "Licencia médica")
    _upsert(db, company_id, CAT_AUSENTISMO, "VAC", "Vacaciones")

    if not template_path or not template_path.exists():
        db.commit()
        return

    from openpyxl import load_workbook

    wb = load_workbook(template_path, data_only=True)

    # Jornadas
    if "Jornadas" in wb.sheetnames:
        ws = wb["Jornadas"]
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        col = headers.index("jornada_id") if "jornada_id" in headers else 0

        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            v = r[col] if col < len(r) else None
            if v is None:
                continue
            jid = str(v).strip()
            if jid:
                _upsert(db, company_id, CAT_JORNADA, jid, jid)

    # CatalogoTurnos
    if "CatalogoTurnos" in wb.sheetnames:
        ws = wb["CatalogoTurnos"]
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        idx_shift = headers.index("shift_id") if "shift_id" in headers else 0
        idx_tipo = headers.index("tipo") if "tipo" in headers else None
        idx_nombre = headers.index("nombre") if "nombre" in headers else None

        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r:
                continue

            shift_id = r[idx_shift] if idx_shift < len(r) else None
            if not shift_id:
                continue

            shift_id = str(shift_id).strip()
            if not shift_id or shift_id == "LIBRE":
                continue

            tipo = ""
            if idx_tipo is not None and idx_tipo < len(r) and r[idx_tipo]:
                tipo = str(r[idx_tipo]).strip().upper()

            label = None
            if idx_nombre is not None and idx_nombre < len(r) and r[idx_nombre]:
                label = str(r[idx_nombre]).strip()

            if tipo == "NO_TRABAJADO":
                _upsert(db, company_id, CAT_AUSENTISMO, shift_id, label or shift_id)

            if tipo == "TRABAJADO":
                _upsert(db, company_id, CAT_TURNO, shift_id, label or shift_id)

    db.commit()


def list_vocab(db: Session, company_id: uuid.UUID) -> Dict[str, List[SolverVocabItem]]:
    rows = (
        db.query(SolverVocabItem)
        .filter(SolverVocabItem.company_id == company_id, SolverVocabItem.active == True)  # noqa: E712
        .order_by(SolverVocabItem.category.asc(), SolverVocabItem.value.asc())
        .all()
    )
    out: Dict[str, List[SolverVocabItem]] = {}
    for r in rows:
        out.setdefault(r.category, []).append(r)
    return out


def list_vocab_values(db: Session, company_id: uuid.UUID, category: str) -> List[str]:
    rows = (
        db.query(SolverVocabItem)
        .filter(
            SolverVocabItem.company_id == company_id,
            SolverVocabItem.category == category,
            SolverVocabItem.active == True,  # noqa: E712
        )
        .order_by(SolverVocabItem.value.asc())
        .all()
    )
    return [r.value for r in rows]