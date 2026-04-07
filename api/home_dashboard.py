"""
api/ui/home.py — Dashboard de inicio
Vigatec Runner · FastAPI + Jinja2
Actualizado: 2026-03-18 — Cobertura v3 (Mínimo vs Ideal)
"""
from __future__ import annotations

import glob
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, Request
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func

from api.db import get_db
from api.masterdata_models import Company, Branch, OrgUnit, Employee
from api.case_data_models import AusentismoEmpleado, RestriccionEmpleado
from api.models import Run
from api.rbac_models import User
from api.rbac import get_current_user, filter_companies, filter_branches, filter_org_units

TEMPLATES = Jinja2Templates(directory=str(Path(__file__).parent.parent / "templates"))
router = APIRouter(prefix="/ui", tags=["ui-home"])


# ─── Contexto activo (cookies, igual que el resto de ui.py) ─────────────────

def _parse_uuid(v):
    if not v:
        return None
    try:
        import uuid
        return uuid.UUID(v)
    except Exception:
        return None

def _get_context(request: Request, db: Session):
    company_id  = _parse_uuid(request.cookies.get("q_company_id"))
    branch_id   = _parse_uuid(request.cookies.get("q_branch_id"))
    ou_id       = _parse_uuid(request.cookies.get("q_org_unit_id"))

    active_company = db.get(Company, company_id) if company_id else None
    active_branch  = db.get(Branch,  branch_id)  if branch_id  else None
    active_ou      = db.get(OrgUnit, ou_id)       if ou_id      else None
    return active_company, active_branch, active_ou


# ─── KPI stats ───────────────────────────────────────────────────────────────

def _build_stats(db: Session, active_ou: Optional[OrgUnit]) -> dict:
    today      = date.today()
    week_start = today - timedelta(days=today.weekday())

    emp_q = db.query(func.count(Employee.id)).filter(Employee.active == True)
    if active_ou:
        emp_q = emp_q.filter(Employee.org_unit_id == active_ou.id)
    total_employees = emp_q.scalar() or 0

    new_this_week = (
        db.query(func.count(Employee.id))
        .filter(Employee.active == True, Employee.created_at >= week_start)
        .scalar() or 0
    )

    total_ous      = db.query(func.count(OrgUnit.id)).scalar() or 0
    total_branches = db.query(func.count(Branch.id)).scalar() or 0

    # Ausentismos activos (fecha_inicio/fin guardadas como string YYYY-MM-DD)
    today_str = today.isoformat()
    ausencias = (
        db.query(AusentismoEmpleado)
        .filter(
            AusentismoEmpleado.fecha_inicio <= today_str,
            AusentismoEmpleado.fecha_fin    >= today_str,
        ).all()
    )
    active_absences = len(ausencias)
    # "Pendientes" = sin detalle (campo detalle vacío)
    pending_absences = sum(1 for a in ausencias if not a.detalle)

    # Alertas de restricción: restricciones hard activas
    restriction_alerts = (
        db.query(func.count(RestriccionEmpleado.id))
        .filter(RestriccionEmpleado.hard == True)
        .scalar() or 0
    )

    # Cobertura: empleados activos con turno en runs completados este mes
    # (aproximación: % de runs exitosos sobre total de runs del mes)
    month_start = today.replace(day=1).isoformat()
    total_runs  = db.query(func.count(Run.id)).filter(Run.created_at >= month_start).scalar() or 1
    ok_runs     = db.query(func.count(Run.id)).filter(
        Run.created_at >= month_start,
        Run.status == "success",
    ).scalar() or 0
    coverage_pct = round((ok_runs / total_runs) * 100)

    return {
        "total_employees":    total_employees,
        "new_employees_week": new_this_week,
        "total_ous":          total_ous,
        "total_branches":     total_branches,
        "active_absences":    active_absences,
        "justified_absences": active_absences - pending_absences,
        "pending_absences":   pending_absences,
        "restriction_alerts": restriction_alerts,
        "coverage_pct":       coverage_pct,
    }


# ─── Cobertura por OU ────────────────────────────────────────────────────────

def _build_ou_coverage(db: Session) -> list[dict]:
    ous = (
        db.query(OrgUnit)
        .join(Branch)
        .join(Company)
        .all()
    )
    result = []
    for ou in ous:
        emp_count = (
            db.query(func.count(Employee.id))
            .filter(Employee.org_unit_id == ou.id, Employee.active == True)
            .scalar() or 0
        )
        # Cobertura: si hay empleados activos asumimos 100%,
        # baja si hay ausentismos activos hoy
        today_str = date.today().isoformat()
        emp_keys = [
            e.employee_key for e in
            db.query(Employee).filter(Employee.org_unit_id == ou.id, Employee.active == True).all()
        ]
        ausentes = 0
        if emp_keys:
            ausentes = (
                db.query(func.count(AusentismoEmpleado.id))
                .filter(
                    AusentismoEmpleado.employee_id.in_(emp_keys),
                    AusentismoEmpleado.fecha_inicio <= today_str,
                    AusentismoEmpleado.fecha_fin    >= today_str,
                ).scalar() or 0
            )
        coverage = round(((emp_count - ausentes) / emp_count) * 100) if emp_count else 0
        result.append({
            "name":           ou.name,
            "branch_name":    ou.branch.name,
            "company_name":   ou.branch.company.name,
            "employee_count": emp_count,
            "coverage":       max(coverage, 0),
        })
    return sorted(result, key=lambda x: x["coverage"])


# ─── Alertas ─────────────────────────────────────────────────────────────────

def _time_ago(dt) -> str:
    """Humaniza una fecha sin dependencias externas."""
    if dt is None:
        return "—"
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt)
        except Exception:
            return dt
    now  = datetime.now(dt.tzinfo)
    diff = now - dt
    s    = int(diff.total_seconds())
    if s < 60:       return "Ahora"
    if s < 3600:     return f"Hace {s // 60} min"
    if s < 86400:    return f"Hace {s // 3600} h"
    if s < 172800:   return "Ayer"
    return dt.strftime("%-d %b")

def _build_alerts(db: Session) -> list[dict]:
    alerts = []

    # Restricciones hard
    hard = (
        db.query(RestriccionEmpleado)
        .filter(RestriccionEmpleado.hard == True)
        .order_by(RestriccionEmpleado.updated_at.desc())
        .limit(2).all()
    )
    for r in hard:
        who = f"Empleado {r.employee_id}" if r.employee_id else "Global"
        alerts.append({
            "type":        "danger",
            "icon":        "⚠️",
            "title":       f"Restricción hard — {who}",
            "description": r.detalle or f"Tipo: {r.tipo}",
            "time_ago":    _time_ago(r.updated_at),
        })

    # Ausentismos sin detalle (pendientes)
    today_str = date.today().isoformat()
    pending = (
        db.query(AusentismoEmpleado)
        .filter(
            AusentismoEmpleado.fecha_inicio <= today_str,
            AusentismoEmpleado.fecha_fin    >= today_str,
            AusentismoEmpleado.detalle      == None,
        )
        .limit(2).all()
    )
    for a in pending:
        alerts.append({
            "type":        "warning",
            "icon":        "🏥",
            "title":       "Ausentismo sin detalle",
            "description": f"Empleado {a.employee_id} · desde {a.fecha_inicio}",
            "time_ago":    _time_ago(a.created_at),
        })

    # Último run completado
    last_ok = (
        db.query(Run)
        .filter(Run.status == "success")
        .order_by(Run.finished_at.desc())
        .first()
    )
    if last_ok:
        alerts.append({
            "type":        "success",
            "icon":        "✅",
            "title":       f"Run completado",
            "description": f"{last_ok.original_filename or 'sin nombre'} · sin errores",
            "time_ago":    _time_ago(last_ok.finished_at),
        })

    # Runs fallidos recientes
    last_fail = (
        db.query(Run)
        .filter(Run.status == "failed")
        .order_by(Run.finished_at.desc())
        .first()
    )
    if last_fail:
        alerts.append({
            "type":        "danger",
            "icon":        "❌",
            "title":       "Run fallido",
            "description": last_fail.error_message or "Revisa el log del run.",
            "time_ago":    _time_ago(last_fail.finished_at),
        })

    return alerts[:6]


# ─── Vista semanal mini ───────────────────────────────────────────────────────

def _build_weekly(db: Session, active_ou: Optional[OrgUnit]):
    today      = date.today()
    week_start = today - timedelta(days=today.weekday())
    days       = [week_start + timedelta(days=i) for i in range(7)]
    DAY_LABELS = ["Lu", "Ma", "Mi", "Ju", "Vi", "Sa", "Do"]
    week_days  = [{"label": DAY_LABELS[i], "is_today": d == today} for i, d in enumerate(days)]

    if not active_ou:
        return [], week_days

    employees = (
        db.query(Employee)
        .filter(Employee.org_unit_id == active_ou.id, Employee.active == True)
        .order_by(Employee.nombre)
        .limit(6).all()
    )
    if not employees:
        return [], week_days

    # Ausentismos de la semana
    emp_keys   = [e.employee_key for e in employees]
    week_end   = days[-1].isoformat()
    week_start_s = days[0].isoformat()
    ausencias  = (
        db.query(AusentismoEmpleado)
        .filter(
            AusentismoEmpleado.employee_id.in_(emp_keys),
            AusentismoEmpleado.fecha_inicio <= week_end,
            AusentismoEmpleado.fecha_fin    >= week_start_s,
        ).all()
    )
    # Indexar por (employee_key, date_str)
    aus_set = set()
    for a in ausencias:
        try:
            fi = date.fromisoformat(a.fecha_inicio)
            ff = date.fromisoformat(a.fecha_fin)
            for d in days:
                if fi <= d <= ff:
                    aus_set.add((a.employee_id, d.isoformat()))
        except Exception:
            pass

    schedule = []
    for emp in employees:
        shifts = []
        for d in days:
            is_today = d == today
            if (emp.employee_key, d.isoformat()) in aus_set:
                shifts.append({"code": "a", "label": "Ausentismo", "is_today": is_today})
            else:
                # Sin plan_mensual.csv disponible en BD, mostramos celda vacía
                shifts.append({"code": "empty", "label": "—", "is_today": is_today})
        schedule.append({"name": emp.nombre, "shifts": shifts})

    return schedule, week_days


# ─── Runs recientes ───────────────────────────────────────────────────────────

def _build_runs(db: Session) -> list[dict]:
    runs = db.query(Run).order_by(Run.created_at.desc()).limit(4).all()
    result = []
    for r in runs:
        if r.status == "success":
            status, label = "ok",    f"Run completado"
            desc = r.original_filename or "—"
        elif r.status in ("queued", "running"):
            status, label = "queue", f"{'En cola' if r.status == 'queued' else 'Ejecutando'}"
            desc = r.original_filename or "—"
        else:
            status, label = "error", "Run fallido"
            desc = r.error_message or "Revisa el log."
        result.append({
            "status":      status,
            "label":       label,
            "description": desc,
            "time_ago":    _time_ago(r.created_at),
        })
    return result


# ─── Colaboradores de hoy ─────────────────────────────────────────────────────

def _build_todays(db: Session, active_ou: Optional[OrgUnit]) -> list[dict]:
    """
    Sin plan_mensual en BD, mostramos los primeros colaboradores activos
    de la OU seleccionada como referencia.
    """
    q = db.query(Employee).filter(Employee.active == True)
    if active_ou:
        q = q.filter(Employee.org_unit_id == active_ou.id)
    employees = q.order_by(Employee.nombre).limit(6).all()

    result = []
    for emp in employees:
        parts    = emp.nombre.split()
        initials = (parts[0][0] + parts[-1][0]).upper() if len(parts) >= 2 else emp.nombre[:2].upper()
        ou_name  = emp.org_unit.name if emp.org_unit else "—"
        result.append({
            "name":       emp.nombre,
            "initials":   initials,
            "ou":         ou_name,
            "time_range": emp.jornada_id or "—",
            "type":       "m",
            "type_label": emp.cargo_id or "—",
        })
    return result


# ─── Cobertura v3: Mínimo vs Ideal ────────────────────────────────────────────

def _build_coverage_summary_v3(db: Session) -> dict:
    """
    Construye resumen de cobertura Mínimo vs Ideal para el Dashboard.
    Lee desde reporte_brechas del último run exitoso.
    
    Retorna dict con: total_slots, sobre_ideal, bajo_ideal, bajo_minimo,
    porcentajes, run_id, run_fecha, tiene_datos
    """
    best_path = None
    best_mtime = 0
    
    # Buscar el reporte más reciente (xlsx o csv)
    for pattern in ["/app/storage/runs/*/out/reporte_brechas.xlsx", 
                    "/app/storage/runs/*/out/reporte_brechas.csv"]:
        for path in glob.glob(pattern):
            try:
                mtime = Path(path).stat().st_mtime
                if mtime > best_mtime:
                    best_mtime = mtime
                    best_path = path
            except Exception:
                continue
    
    if not best_path:
        return {"tiene_datos": False}
    
    # Leer archivo
    rows = []
    try:
        if best_path.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(best_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            wb.close()
        else:
            import csv as csv_mod
            with open(best_path, encoding="utf-8-sig") as f:
                rows = list(csv_mod.DictReader(f))
    except Exception as e:
        print(f"[Dashboard v3] Error leyendo {best_path}: {e}")
        return {"tiene_datos": False}
    
    if not rows:
        return {"tiene_datos": False}
    
    # Contar por diagnóstico
    sobre_ideal = 0
    bajo_ideal = 0
    bajo_minimo = 0
    total_slots = 0
    
    for r in rows:
        r_lower = {str(k).lower().strip(): v for k, v in r.items()}
        
        req_min = int(float(r_lower.get("requeridos_min_personas", 0) or 
                           r_lower.get("requeridos_personas", 0) or 0))
        if req_min == 0:
            continue
            
        total_slots += 1
        
        diag = str(r_lower.get("diagnostic_ideal", "")).upper().strip()
        
        if diag == "BAJO_MINIMO":
            bajo_minimo += 1
        elif diag == "BAJO_IDEAL":
            bajo_ideal += 1
        elif diag == "SOBRE_IDEAL":
            sobre_ideal += 1
        else:
            # Fallback legacy
            falt_min = int(float(r_lower.get("faltantes_vs_min_personas", 
                                              r_lower.get("faltantes_personas", 0)) or 0))
            falt_ideal = int(float(r_lower.get("faltantes_vs_ideal_personas", 0) or 0))
            
            if falt_min > 0:
                bajo_minimo += 1
            elif falt_ideal > 0:
                bajo_ideal += 1
            else:
                sobre_ideal += 1
    
    if total_slots == 0:
        return {"tiene_datos": False}
    
    pct_sobre = round(sobre_ideal / total_slots * 100, 1)
    pct_bajo_i = round(bajo_ideal / total_slots * 100, 1)
    pct_bajo_m = round(bajo_minimo / total_slots * 100, 1)
    
    run_id = Path(best_path).parent.parent.name
    run_fecha = datetime.fromtimestamp(best_mtime).strftime("%-d %b %Y")
    
    return {
        "tiene_datos": True,
        "total_slots": total_slots,
        "sobre_ideal": sobre_ideal,
        "bajo_ideal": bajo_ideal,
        "bajo_minimo": bajo_minimo,
        "pct_sobre_ideal": pct_sobre,
        "pct_bajo_ideal": pct_bajo_i,
        "pct_bajo_minimo": pct_bajo_m,
        "run_id": run_id,
        "run_fecha": run_fecha,
    }


# ─── ROUTE ────────────────────────────────────────────────────────────────────

@router.get("/home", response_class=HTMLResponse)
async def ui_home(
    request:      Request,
    db:           Session = Depends(get_db),
    current_user: User    = Depends(get_current_user),
):
    active_company, active_branch, active_ou = _get_context(request, db)
    weekly_schedule, week_days = _build_weekly(db, active_ou)
    alerts = _build_alerts(db)

    return TEMPLATES.TemplateResponse("home.html", {
        "request":         request,
        "active_nav":      "home",
        "current_user":    current_user,
        "active_company":  active_company,
        "active_branch":   active_branch,
        "active_ou":       active_ou,

        # Dashboard
        "stats":           _build_stats(db, active_ou),
        "ou_coverage":     _build_ou_coverage(db),
        "alerts":          alerts,
        "alerts_count":    len(alerts),
        "weekly_schedule": weekly_schedule,
        "week_days":       week_days,
        "recent_runs":     _build_runs(db),
        "todays_shifts":   _build_todays(db, active_ou),

        # NUEVO: Cobertura v3 (Mínimo vs Ideal)
        "coverage_v3":     _build_coverage_summary_v3(db),

        "now": datetime.now(),
    })