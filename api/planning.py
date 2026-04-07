from __future__ import annotations

import calendar
import json
import shutil
import uuid
from dataclasses import dataclass, asdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session

from api.db import get_db
from api.models import Run
from api.schemas import RunCreateResponse
from api.storage import get_run_paths, ensure_dirs
from worker.tasks import execute_run


DOW_MAP = {
    0: "LUN",
    1: "MAR",
    2: "MIE",
    3: "JUE",
    4: "VIE",
    5: "SAB",
    6: "DOM",
}

router = APIRouter(prefix="/planning", tags=["planning"])


# -------------------------
# Helpers: ciclos mensuales
# -------------------------
def _parse_month(month_str: str) -> Tuple[int, int]:
    # month_str: "YYYY-MM"
    try:
        y_s, m_s = month_str.split("-", 1)
        y = int(y_s)
        m = int(m_s)
        if m < 1 or m > 12:
            raise ValueError
        return y, m
    except Exception:
        raise HTTPException(status_code=400, detail="month must be YYYY-MM (e.g. 2026-03)")


def _month_first_day(year: int, month: int) -> date:
    return date(year, month, 1)


def _monday_of_week(d: date) -> date:
    # Monday=0
    return d - timedelta(days=d.weekday())


def _sundays_in_month(year: int, month: int) -> List[date]:
    last_day = calendar.monthrange(year, month)[1]
    sundays: List[date] = []
    for day in range(1, last_day + 1):
        d = date(year, month, day)
        if d.weekday() == 6:  # Sunday
            sundays.append(d)
    return sundays


@dataclass
class PlanningCycle:
    month: str
    start_date: str  # YYYY-MM-DD
    weeks: int
    sundays: List[str]
    end_date: str  # YYYY-MM-DD (inclusive)

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


def compute_cycle(month: str) -> PlanningCycle:
    y, m = _parse_month(month)
    first = _month_first_day(y, m)
    start = _monday_of_week(first)
    sundays = _sundays_in_month(y, m)
    weeks = len(sundays) if len(sundays) > 0 else 4  # fallback, pero en práctica siempre hay
    end = start + timedelta(days=weeks * 7 - 1)
    return PlanningCycle(
        month=month,
        start_date=start.isoformat(),
        weeks=weeks,
        sundays=[d.isoformat() for d in sundays],
        end_date=end.isoformat(),
    )


def _company_cycles_path(empresa_id: str) -> Path:
    # Guardamos por empresa en storage (sin DB por ahora)
    base = Path("/app/storage/companies") / empresa_id
    base.mkdir(parents=True, exist_ok=True)
    return base / "planning_cycles.json"


def _load_company_cycles(empresa_id: str) -> Dict[str, Any]:
    p = _company_cycles_path(empresa_id)
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_company_cycles(empresa_id: str, data: Dict[str, Any]) -> None:
    p = _company_cycles_path(empresa_id)
    p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


@router.get("/cycles/{month}")
def get_cycle(
    month: str,
    empresa_id: Optional[str] = Query(default=None, description="Si se entrega, intenta devolver el ciclo cacheado por empresa"),
):
    if empresa_id:
        cached = _load_company_cycles(empresa_id).get(month)
        if cached:
            return cached
    return compute_cycle(month).to_dict()


@router.post("/cycles/precompute")
def precompute_cycles(
    empresa_id: str = Query(..., description="ID empresa (clave de almacenamiento)"),
    start_month: str = Query(..., description="YYYY-MM, ejemplo 2026-03"),
    months: int = Query(24, ge=1, le=120),
):
    y0, m0 = _parse_month(start_month)

    def add_months(y: int, m: int, k: int) -> Tuple[int, int]:
        mm = (y * 12 + (m - 1)) + k
        yy = mm // 12
        mo = (mm % 12) + 1
        return yy, mo

    data = _load_company_cycles(empresa_id)
    for k in range(months):
        yy, mo = add_months(y0, m0, k)
        month = f"{yy:04d}-{mo:02d}"
        data[month] = compute_cycle(month).to_dict()

    _save_company_cycles(empresa_id, data)

    return {
        "empresa_id": empresa_id,
        "start_month": start_month,
        "months": months,
        "saved": months,
        "path": str(_company_cycles_path(empresa_id)),
    }


# -------------------------
# Helpers: construir PlanPrevio + actualizar Parametros
# -------------------------
def _parse_date_any(x: Any) -> Optional[date]:
    if x is None:
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    if isinstance(x, str):
        s = x.strip()
        if not s:
            return None
        s = s.replace("Z", "")
        if "T" in s:
            s = s.split("T", 1)[0]
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None
    return None


def _as_iso_z(d: date) -> str:
    # Consistente con tu template: "YYYY-MM-DDT00:00:00Z"
    return f"{d.isoformat()}T00:00:00Z"


def _read_catalogo_cruza_map(case_wb) -> Dict[str, bool]:
    # Mapea shift_id -> cruza_medianoche
    mp: Dict[str, bool] = {}
    if "CatalogoTurnos" not in case_wb.sheetnames:
        return mp
    ws = case_wb["CatalogoTurnos"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

    # Esperamos una columna "shift_id" y "cruza_medianoche"
    if "shift_id" not in idx or "cruza_medianoche" not in idx:
        return mp

    for row in ws.iter_rows(min_row=2):
        sid = row[idx["shift_id"]].value
        cm = row[idx["cruza_medianoche"]].value
        if sid is None:
            continue
        s = str(sid).strip()
        v = "" if cm is None else str(cm).strip().upper()
        mp[s] = v in ("SI", "SÍ", "YES", "TRUE", "1")
    return mp


def _read_plan_rows(plan_path: Path) -> List[Dict[str, Any]]:
    # Soporta plan_mensual.xlsx (preferido) o plan_mensual.csv (fallback)
    if not plan_path.exists():
        return []

    if plan_path.suffix.lower() == ".csv":
        import csv

        rows: List[Dict[str, Any]] = []
        with plan_path.open("r", encoding="utf-8-sig", newline="") as f:
            rdr = csv.DictReader(f)
            for r in rdr:
                rows.append(dict(r))
        return rows

    # xlsx
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"openpyxl required to read xlsx plans: {e}")

    wb = load_workbook(plan_path, data_only=True)
    # tomamos hoja más probable
    sheet = None
    for name in wb.sheetnames:
        if "plan" in name.lower():
            sheet = name
            break
    if sheet is None:
        sheet = wb.sheetnames[0]
    ws = wb[sheet]

    # header row
    header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [str(h).strip() if h is not None else "" for h in header_cells]
    idx = {h: i for i, h in enumerate(headers) if h}

    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        d: Dict[str, Any] = {}
        for h, i in idx.items():
            if i < len(r):
                d[h] = r[i]
        rows.append(d)
    return rows


def _ensure_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(title=name)


def _update_parametros(case_wb, start_date: date, weeks: int) -> None:
    ws = _ensure_sheet(case_wb, "Parametros")

    # Si la hoja está vacía, ponemos headers estándar
    if ws.max_row < 1 or (ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None):
        ws["A1"].value = "parametro"
        ws["B1"].value = "valor"
        ws["C1"].value = "detalle"

    def set_param(key: str, value: Any, detail: str = ""):
        found = False
        for row in ws.iter_rows(min_row=2):
            k = row[0].value
            if k is None:
                continue
            if str(k).strip() == key:
                row[1].value = value
                if len(row) >= 3:
                    row[2].value = detail
                found = True
                break
        if not found:
            ws.append([key, value, detail])

    set_param("fecha_inicio_mes", _as_iso_z(start_date), "Lunes de inicio de planificación (según ciclo)")
    set_param("semanas", int(weeks), "Horizonte de planificación (= # domingos del mes)")


def _build_plan_previo(
    case_wb,
    base_plan_rows: List[Dict[str, Any]],
    start_date: date,
) -> None:
    # PlanPrevio = semana anterior a start_date (LUN..DOM)
    last_week_start = start_date - timedelta(days=7)
    last_week_end = start_date - timedelta(days=1)

    ws = _ensure_sheet(case_wb, "PlanPrevio")

    # Limpia todo y deja header exacto (como tu template)
    ws.delete_rows(1, ws.max_row)
    headers = ["employee_id", "fecha", "dia_semana", "org_unit_id", "cargo", "shift_id", "es_saliente", "nota"]
    ws.append(headers)

    # Columnas posibles del plan_mensual
    # aceptamos: employee_id, fecha, org_unit_id, cargo (o cargo_id), shift_id, nota
    # además algunos planes pueden tener dia_semana ya, pero lo recalculamos.
    def get_any(r: Dict[str, Any], keys: List[str]) -> Any:
        for k in keys:
            if k in r:
                return r.get(k)
        return None

    # Para es_saliente necesitamos mirar el día anterior también, así que tomamos un rango extendido
    extend_start = last_week_start - timedelta(days=1)
    extend_end = last_week_end

    # Parseamos solo lo necesario y construimos mapa por (emp, fecha)->shift
    parsed: List[Dict[str, Any]] = []
    for r in base_plan_rows:
        emp = get_any(r, ["employee_id", "EmployeeId", "empleado", "id_empleado"])
        dt = _parse_date_any(get_any(r, ["fecha", "date", "Fecha"]))
        if emp is None or dt is None:
            continue
        if dt < extend_start or dt > extend_end:
            continue

        org = get_any(r, ["org_unit_id", "unidad", "ou", "OrgUnitId"])
        cargo = get_any(r, ["cargo", "cargo_id", "Cargo", "CargoId"])
        shift = get_any(r, ["shift_id", "turno", "ShiftId", "TurnoId"])
        nota = get_any(r, ["nota", "notes", "Nota"])

        parsed.append(
            {
                "employee_id": str(emp).strip(),
                "fecha": dt,
                "org_unit_id": "" if org is None else str(org).strip(),
                "cargo": "" if cargo is None else str(cargo).strip(),
                "shift_id": "" if shift is None else str(shift).strip(),
                "nota": "" if nota is None else str(nota),
            }
        )

    shift_by_emp_date: Dict[Tuple[str, date], str] = {(p["employee_id"], p["fecha"]): p["shift_id"] for p in parsed}

    cruza_map = _read_catalogo_cruza_map(case_wb)

    def crosses(sid: str) -> bool:
        if not sid:
            return False
        return cruza_map.get(sid, False)

    # Ahora escribimos SOLO la semana previa
    rows_week = [p for p in parsed if last_week_start <= p["fecha"] <= last_week_end]
    rows_week.sort(key=lambda x: (x["employee_id"], x["fecha"]))

    for p in rows_week:
        emp = p["employee_id"]
        d = p["fecha"]
        prev_d = d - timedelta(days=1)
        prev_shift = shift_by_emp_date.get((emp, prev_d), "")
        es_saliente = 1 if crosses(prev_shift) else 0

        ws.append(
            [
                emp,
                _as_iso_z(d),
                DOW_MAP[d.weekday()],
                p["org_unit_id"],
                p["cargo"],
                p["shift_id"],
                es_saliente,
                p["nota"],
            ]
        )


def build_case_for_month_from_run(
    base_run: Run,
    target_month: str,
    out_case_path: Path,
) -> Dict[str, Any]:
    # Copiamos case base
    base_case = Path(base_run.case_path)
    if not base_case.exists():
        raise HTTPException(status_code=409, detail="Base run case.xlsx not found on disk")
    shutil.copyfile(base_case, out_case_path)

    # Calcula ciclo
    cycle = compute_cycle(target_month)
    start_date = datetime.strptime(cycle.start_date, "%Y-%m-%d").date()
    weeks = int(cycle.weeks)

    # Encuentra plan mensual del run base
    out_dir = Path(base_run.out_dir)
    plan_xlsx = out_dir / "plan_mensual.xlsx"
    plan_csv = out_dir / "plan_mensual.csv"
    plan_path = plan_xlsx if plan_xlsx.exists() else plan_csv
    if not plan_path.exists():
        raise HTTPException(status_code=409, detail="Base run has no plan_mensual.xlsx/csv in out_dir")

    base_plan_rows = _read_plan_rows(plan_path)

    # Editamos el case copiado (Parametros + PlanPrevio)
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"openpyxl required to edit case.xlsx: {e}")

    wb = load_workbook(out_case_path)
    _update_parametros(wb, start_date=start_date, weeks=weeks)
    _build_plan_previo(wb, base_plan_rows=base_plan_rows, start_date=start_date)

    wb.save(out_case_path)

    return cycle.to_dict()


# -------------------------
# Endpoint: solicitar turnos (1 click)
# -------------------------
@router.post("/request_turnos", response_model=RunCreateResponse)
def request_turnos(
    base_run_id: uuid.UUID = Query(..., description="Run exitoso del mes actual (base)"),
    target_month: str = Query(..., description="YYYY-MM del mes que quieres planificar (ej. 2026-04)"),
    db: Session = Depends(get_db),
):
    base_run: Run | None = db.get(Run, base_run_id)
    if not base_run:
        raise HTTPException(status_code=404, detail="Base run not found")
    if base_run.status != "success":
        raise HTTPException(status_code=409, detail="Base run must be success")

    new_run_id = uuid.uuid4()
    paths = get_run_paths(new_run_id)
    ensure_dirs(paths)

    cycle = build_case_for_month_from_run(base_run, target_month=target_month, out_case_path=paths.case_path)

    run = Run(
        id=new_run_id,
        status="queued",
        original_filename=f"generated_case_{target_month}.xlsx",
        case_path=str(paths.case_path),
        out_dir=str(paths.out_dir),
        log_path=str(paths.log_path),
    )
    db.add(run)
    db.commit()

    execute_run.delay(str(new_run_id))

    return RunCreateResponse(id=new_run_id, status="queued")