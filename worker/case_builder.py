from __future__ import annotations

import calendar
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable

import pandas as pd
import openpyxl


DOW_ES = {
    0: "LUN",
    1: "MAR",
    2: "MIE",
    3: "JUE",
    4: "VIE",
    5: "SAB",
    6: "DOM",
}


@dataclass(frozen=True)
class PlanningCycle:
    month: str  # "YYYY-MM"
    weeks: int
    cycle_start: date  # Monday
    cycle_end: date  # Sunday
    sundays: list[date]


def parse_month(s: str) -> tuple[int, int]:
    # s: "YYYY-MM"
    if len(s) != 7 or s[4] != "-":
        raise ValueError("target_month must be in YYYY-MM format")
    y = int(s[:4])
    m = int(s[5:])
    if not (1 <= m <= 12):
        raise ValueError("month must be 1..12")
    return y, m


def compute_cycle(target_month: str) -> PlanningCycle:
    """
    Rule: weeks to plan = number of Sundays in the month.
    cycle_end = last Sunday of the month.
    cycle_start = cycle_end - (7*weeks - 1) days  (so it starts on a Monday).
    """
    y, m = parse_month(target_month)
    cal = calendar.monthcalendar(y, m)

    sundays: list[date] = []
    # monthcalendar: weeks rows, columns Mon..Sun (0..6)
    for wk in cal:
        d = wk[calendar.SUNDAY]
        if d != 0:
            sundays.append(date(y, m, d))

    if not sundays:
        # should never happen
        raise ValueError("No Sundays found; invalid calendar?")

    weeks = len(sundays)
    cycle_end = max(sundays)
    cycle_start = cycle_end - timedelta(days=(7 * weeks - 1))

    if cycle_start.weekday() != 0:
        # safety: enforce Monday
        cycle_start = cycle_start - timedelta(days=cycle_start.weekday())

    return PlanningCycle(
        month=target_month,
        weeks=weeks,
        cycle_start=cycle_start,
        cycle_end=cycle_end,
        sundays=sundays,
    )


def _read_prev_plan_df(prev_plan_xlsx: Path) -> pd.DataFrame:
    if not prev_plan_xlsx.exists():
        raise FileNotFoundError(f"Previous plan not found: {prev_plan_xlsx}")

    df = pd.read_excel(prev_plan_xlsx, sheet_name=0)
    # normalize expected columns
    df.columns = [str(c).strip() for c in df.columns]
    # coerce date
    if "fecha" in df.columns:
        df["fecha"] = pd.to_datetime(df["fecha"]).dt.date
    return df


def _read_dotacion_df(case_path: Path) -> pd.DataFrame:
    df = pd.read_excel(case_path, sheet_name="Dotacion")
    df.columns = [str(c).strip() for c in df.columns]
    # ensure expected columns exist
    for c in ["employee_id", "org_unit_id", "cargo"]:
        if c not in df.columns:
            raise ValueError(f"Dotacion sheet missing column: {c}")
    df["employee_id"] = df["employee_id"].astype(str)
    return df


def _daterange(d1: date, d2: date) -> Iterable[date]:
    cur = d1
    while cur <= d2:
        yield cur
        cur += timedelta(days=1)


def build_plan_previo(
    *,
    case_path: Path,
    prev_plan_xlsx: Path,
    cycle_start: date,
) -> pd.DataFrame:
    """
    PlanPrevio = 7 days before cycle_start (Mon..Sun).
    We take rows from prev plan_mensual.xlsx for that week.
    If any employee-date is missing, we fill it as LIBRE using Dotacion org_unit_id/cargo.
    Columns required:
      employee_id, fecha, dia_semana, org_unit_id, cargo, shift_id, es_saliente, nota
    """
    prev_week_start = cycle_start - timedelta(days=7)
    prev_week_end = cycle_start - timedelta(days=1)

    df_plan = _read_prev_plan_df(prev_plan_xlsx)
    df_dot = _read_dotacion_df(case_path)

    # Ensure plan columns
    for c in ["employee_id", "fecha", "org_unit_id", "cargo", "shift_id"]:
        if c not in df_plan.columns:
            raise ValueError(f"plan_mensual.xlsx missing column: {c}")

    df_plan["employee_id"] = df_plan["employee_id"].astype(str)

    # Filter to the prev week window
    df_week = df_plan[(df_plan["fecha"] >= prev_week_start) & (df_plan["fecha"] <= prev_week_end)].copy()

    # Add missing columns if needed
    if "dia_semana" not in df_week.columns:
        df_week["dia_semana"] = df_week["fecha"].apply(lambda d: DOW_ES[d.weekday()])
    if "es_saliente" not in df_week.columns:
        df_week["es_saliente"] = 0
    if "nota" not in df_week.columns:
        df_week["nota"] = "PlanPrevio"

    # Reduce to expected columns
    df_week = df_week[["employee_id", "fecha", "dia_semana", "org_unit_id", "cargo", "shift_id", "es_saliente", "nota"]]

    # Fill missing employee-date combinations as LIBRE
    dot_map = df_dot.set_index("employee_id")[["org_unit_id", "cargo"]].to_dict(orient="index")

    existing = set(zip(df_week["employee_id"], df_week["fecha"]))
    rows_fill = []
    for emp_id in df_dot["employee_id"].astype(str).tolist():
        info = dot_map.get(emp_id, {"org_unit_id": "", "cargo": ""})
        for d in _daterange(prev_week_start, prev_week_end):
            key = (emp_id, d)
            if key in existing:
                continue
            rows_fill.append(
                {
                    "employee_id": emp_id,
                    "fecha": d,
                    "dia_semana": DOW_ES[d.weekday()],
                    "org_unit_id": info.get("org_unit_id", ""),
                    "cargo": info.get("cargo", ""),
                    "shift_id": "LIBRE",
                    "es_saliente": 0,
                    "nota": "PlanPrevio_filled",
                }
            )

    if rows_fill:
        df_week = pd.concat([df_week, pd.DataFrame(rows_fill)], ignore_index=True)

    # Sort for readability
    df_week = df_week.sort_values(["org_unit_id", "cargo", "employee_id", "fecha"]).reset_index(drop=True)

    # Cast es_saliente to int
    df_week["es_saliente"] = df_week["es_saliente"].fillna(0).astype(int)

    return df_week


def apply_parametros_and_planprevio(
    *,
    case_path: Path,
    target_month: str,
    prev_plan_xlsx: Path,
) -> dict:
    """
    Opens case.xlsx, updates Parametros (fecha_inicio_mes, semanas),
    and overwrites PlanPrevio with computed data from previous plan.

    Returns a dict with summary info for logs.
    """
    cycle = compute_cycle(target_month)

    wb = openpyxl.load_workbook(case_path)

    # --- Update Parametros sheet ---
    if "Parametros" not in wb.sheetnames:
        raise ValueError("case.xlsx missing sheet: Parametros")

    ws = wb["Parametros"]

    # Find columns: parametro / valor
    # Expect header row in row 1
    headers = {str(ws.cell(1, j).value).strip().lower(): j for j in range(1, ws.max_column + 1)}
    if "parametro" not in headers or "valor" not in headers:
        raise ValueError("Parametros sheet must have columns: parametro, valor")

    col_param = headers["parametro"]
    col_val = headers["valor"]

    def _set_param(param_name: str, new_val):
        for i in range(2, ws.max_row + 1):
            p = ws.cell(i, col_param).value
            if p is None:
                continue
            if str(p).strip() == param_name:
                ws.cell(i, col_val).value = new_val
                return True
        return False

    ok1 = _set_param("fecha_inicio_mes", datetime(cycle.cycle_start.year, cycle.cycle_start.month, cycle.cycle_start.day))
    ok2 = _set_param("semanas", int(cycle.weeks))
    if not ok1 or not ok2:
        raise ValueError("Parametros missing required rows: fecha_inicio_mes and/or semanas")

    # --- Build PlanPrevio ---
    df_previo = build_plan_previo(case_path=case_path, prev_plan_xlsx=prev_plan_xlsx, cycle_start=cycle.cycle_start)

    # --- Write PlanPrevio sheet ---
    if "PlanPrevio" in wb.sheetnames:
        ws_prev = wb["PlanPrevio"]
        # Clear existing
        ws_prev.delete_rows(1, ws_prev.max_row)
    else:
        ws_prev = wb.create_sheet("PlanPrevio")

    cols = ["employee_id", "fecha", "dia_semana", "org_unit_id", "cargo", "shift_id", "es_saliente", "nota"]
    ws_prev.append(cols)
    for _, r in df_previo.iterrows():
        ws_prev.append(
            [
                str(r.get("employee_id", "")),
                r.get("fecha"),
                r.get("dia_semana", ""),
                r.get("org_unit_id", ""),
                r.get("cargo", ""),
                r.get("shift_id", ""),
                int(r.get("es_saliente", 0)),
                r.get("nota", ""),
            ]
        )

    wb.save(case_path)

    return {
        "target_month": target_month,
        "cycle_start": str(cycle.cycle_start),
        "cycle_end": str(cycle.cycle_end),
        "weeks": cycle.weeks,
        "planprevio_rows": int(df_previo.shape[0]),
    }
