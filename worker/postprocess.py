from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, Tuple, List, Dict, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


DOW_MAP = {
    0: "LUN",
    1: "MAR",
    2: "MIE",
    3: "JUE",
    4: "VIE",
    5: "SAB",
    6: "DOM",
}

SHIFT_RE = re.compile(r"^S[_-]?(?P<a>[^_]+)_(?P<b>[^_]+)_(?P<rest>.+)$", re.IGNORECASE)


# -------------------------
# Helpers generales
# -------------------------
def _utcnow_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _norm_empid(x) -> str:
    """Normaliza employee_id (rut) para join: string, sin espacios, upper."""
    if x is None:
        return ""
    s = str(x).strip().upper()
    if s.lower() == "nan":
        return ""
    s = s.replace(" ", "")
    return s


def _pick_col(cols: List[str], candidates: List[str]) -> Optional[str]:
    """Encuentra columna por contains (case-insensitive)."""
    cols_norm = [str(c).strip() for c in cols]
    lower = {c.lower(): c for c in cols_norm}
    for cand in candidates:
        cand_l = cand.lower()
        if cand_l in lower:
            return lower[cand_l]
        for c in cols_norm:
            if cand_l in c.lower():
                return c
    return None


def _find_case_xlsx(out_dir: Path) -> Optional[Path]:
    """
    Busca el case.xlsx asociado al run.
    Asume estructura típica:
      runs/<id>/out  (out_dir)
      runs/<id>/in/case.xlsx  o runs/<id>/input/case.xlsx  o runs/<id>/case.xlsx
    """
    run_dir = out_dir.parent  # .../runs/<id>
    candidates = [
        run_dir / "in" / "case.xlsx",
        run_dir / "input" / "case.xlsx",
        run_dir / "inputs" / "case.xlsx",
        run_dir / "case.xlsx",
        out_dir / "case.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    for folder in [run_dir / "in", run_dir / "input", run_dir / "inputs"]:
        if folder.exists():
            for p in folder.glob("*.xlsx"):
                return p
    return None


# -------------------------
# Parsing de turnos
# -------------------------
def _norm_time(token: str) -> Optional[str]:
    if token is None:
        return None
    s = str(token).strip()
    if not s:
        return None
    s = s.replace(".", ":")
    if ":" in s:
        parts = s.split(":")
        if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
            h = int(parts[0])
            m = int(parts[1])
            return f"{h:02d}:{m:02d}"
        return None
    if s.isdigit():
        if len(s) == 4:
            h = int(s[:2])
            m = int(s[2:])
            return f"{h:02d}:{m:02d}"
        if len(s) == 3:
            h = int(s[0])
            m = int(s[1:])
            return f"{h:02d}:{m:02d}"
    return None


def _parse_shift_id(shift_id: str) -> Tuple[str, Optional[str], Optional[str], Optional[int]]:
    if shift_id is None:
        return "", None, None, None

    s = str(shift_id).strip().upper()
    if s in {"", "N/A", "-", "0"}:
        return "", None, None, None
    if s == "LIBRE":
        return "LIBRE", None, None, None

    m = SHIFT_RE.match(s)
    if not m:
        return s, None, None, None

    a = _norm_time(m.group("a"))
    b = _norm_time(m.group("b"))

    colacion = None
    rest = str(m.group("rest"))
    for tok in reversed(rest.split("_")):
        tok2 = tok.strip()
        if tok2.isdigit():
            colacion = int(tok2)
            break

    if a and b:
        return f"{a}–{b}", a, b, colacion

    return s, None, None, colacion


# -------------------------
# Cruce de nombres desde case.xlsx
# -------------------------
def _build_name_from_parts(df: pd.DataFrame) -> Optional[pd.Series]:
    cols = list(df.columns)
    nombres_col = _pick_col(cols, ["nombres", "nombre(s)", "given", "first"])
    apellidos_col = _pick_col(cols, ["apellidos", "apellido(s)", "last", "surname"])
    if nombres_col and apellidos_col:
        nombres = df[nombres_col].fillna("").astype(str).str.strip()
        apellidos = df[apellidos_col].fillna("").astype(str).str.strip()
        full = (nombres + " " + apellidos).str.strip()
        full = full.replace({"nan": ""})
        return full
    return None


def _load_name_map_from_case(case_path: Path) -> Tuple[Optional[pd.DataFrame], Dict[str, Any]]:
    """
    Busca en todas las hojas del case un mapping employee_id/rut -> nombre.
    Devuelve (DF_map, info).
    DF_map: employee_id, nombre_case
    """
    info: Dict[str, Any] = {"found": False}
    try:
        xls = pd.ExcelFile(case_path)
    except Exception:
        return None, info

    id_candidates = [
        "employee_id", "rut", "id_colaborador", "id_trabajador",
        "colaborador", "trabajador", "run"
    ]
    name_candidates = [
        "nombre", "name", "full_name", "nombre_completo",
        "trabajador_nombre", "colaborador_nombre"
    ]

    best_map = None
    best_score = 0
    best_info: Dict[str, Any] = {"found": False}

    for sheet in xls.sheet_names:
        try:
            df = pd.read_excel(case_path, sheet_name=sheet)
        except Exception:
            continue
        if df is None or df.empty:
            continue

        df.columns = [str(c).strip() for c in df.columns]
        cols = list(df.columns)

        id_col = _pick_col(cols, id_candidates)
        if not id_col:
            continue

        name_col = _pick_col(cols, name_candidates)
        name_series = None
        name_method = None

        if name_col:
            name_series = df[name_col].fillna("").astype(str).str.strip()
            name_method = f"col:{name_col}"
        else:
            name_series = _build_name_from_parts(df)
            if name_series is not None:
                name_method = "parts:nombres+apellidos"

        if name_series is None:
            continue

        tmp = pd.DataFrame(
            {
                "employee_id": df[id_col].apply(_norm_empid),
                "nombre_case": name_series.fillna("").astype(str).str.strip(),
            }
        )
        tmp = tmp[(tmp["employee_id"] != "") & (tmp["nombre_case"] != "")]
        if tmp.empty:
            continue

        score = int(tmp["employee_id"].nunique())
        if score > best_score:
            tmp = tmp.drop_duplicates(subset=["employee_id"], keep="first")
            best_map = tmp
            best_score = score
            best_info = {
                "found": True,
                "sheet": sheet,
                "id_col": id_col,
                "name_method": name_method,
                "unique_ids": score,
            }

    return best_map, best_info


# -------------------------
# QA mínimo (sin perdernos)
# -------------------------
def _build_qa(df_long: pd.DataFrame) -> Tuple[Dict[str, Any], pd.DataFrame]:
    """
    QA mínimo:
      - duplicados employee_id+fecha
      - nombre faltante (por empleado)
      - turno faltante (shift_id vacío o turno_display vacío)
      - libre por día
    Devuelve (qa_dict, issues_df)
    """
    issues = []

    # Duplicados por llave
    key_cols = ["employee_id", "fecha"]
    dup_mask = df_long.duplicated(subset=key_cols, keep=False)
    dup_rows = df_long[dup_mask].copy()
    dup_count = int(dup_rows.shape[0])

    if dup_count > 0:
        tmp = dup_rows[["org_unit_id", "cargo", "employee_id", "nombre", "fecha", "shift_id", "turno_display"]].copy()
        tmp.insert(0, "issue_type", "DUPLICATE_EMPLOYEE_FECHA")
        issues.append(tmp)

    # Nombre faltante (por empleado)
    # (ojo: nombre podría venir vacío aunque el plan esté ok, pero lo reportamos)
    no_name_mask = df_long["nombre"].fillna("").astype(str).str.strip() == ""
    no_name_emps = df_long.loc[no_name_mask, "employee_id"].dropna().unique().tolist()
    no_name_emp_count = int(len(no_name_emps))

    if no_name_emp_count > 0:
        tmp = df_long[df_long["employee_id"].isin(no_name_emps)][
            ["org_unit_id", "cargo", "employee_id", "nombre", "fecha", "shift_id", "turno_display"]
        ].copy()
        tmp = tmp.drop_duplicates(subset=["employee_id"]).head(200)  # muestra acotada
        tmp.insert(0, "issue_type", "MISSING_NAME_EMPLOYEE")
        issues.append(tmp)

    # Turno faltante
    shift_empty = df_long["shift_id"].fillna("").astype(str).str.strip() == ""
    display_empty = df_long["turno_display"].fillna("").astype(str).str.strip() == ""
    missing_turno_mask = shift_empty | display_empty
    missing_turno_rows = df_long[missing_turno_mask].copy()
    missing_turno_count = int(missing_turno_rows.shape[0])

    if missing_turno_count > 0:
        tmp = missing_turno_rows[
            ["org_unit_id", "cargo", "employee_id", "nombre", "fecha", "shift_id", "turno_display"]
        ].copy()
        tmp.insert(0, "issue_type", "MISSING_TURNO")
        issues.append(tmp)

    # Libres por día
    libre_mask = df_long["turno_display"].fillna("").astype(str).str.upper().str.strip() == "LIBRE"
    libres_por_dia = (
        df_long.assign(is_libre=libre_mask)
        .groupby("fecha")["is_libre"]
        .sum()
        .sort_index()
        .astype(int)
        .to_dict()
    )

    # Totales por día (filas)
    filas_por_dia = df_long.groupby("fecha")["employee_id"].count().sort_index().astype(int).to_dict()

    # Resumen general
    qa = {
        "rows_total": int(df_long.shape[0]),
        "employees_unique": int(df_long["employee_id"].nunique()),
        "date_min": str(df_long["fecha"].min()) if not df_long.empty else "",
        "date_max": str(df_long["fecha"].max()) if not df_long.empty else "",
        "duplicates_employee_fecha_rows": dup_count,
        "missing_name_employees": no_name_emp_count,
        "missing_turno_rows": missing_turno_count,
        "libres_por_dia": {str(k): int(v) for k, v in libres_por_dia.items()},
        "filas_por_dia": {str(k): int(v) for k, v in filas_por_dia.items()},
    }

    issues_df = pd.concat(issues, ignore_index=True) if issues else pd.DataFrame(
        columns=["issue_type", "org_unit_id", "cargo", "employee_id", "nombre", "fecha", "shift_id", "turno_display"]
    )
    return qa, issues_df


# -------------------------
# Main: genera artifacts visuales
# -------------------------
def build_visual_artifacts(out_dir: str, run_id: str) -> None:
    outp = Path(out_dir)
    plan_xlsx = outp / "plan_mensual.xlsx"
    plan_csv = outp / "plan_mensual.csv"

    if plan_xlsx.exists():
        df = pd.read_excel(plan_xlsx, sheet_name=0)
        source = "plan_mensual.xlsx"
    elif plan_csv.exists():
        df = pd.read_csv(plan_csv)
        source = "plan_mensual.csv"
    else:
        raise FileNotFoundError("No plan_mensual.xlsx/csv found in out_dir")

    df.columns = [str(c).strip() for c in df.columns]

    required = ["employee_id", "fecha", "org_unit_id", "cargo", "shift_id"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"plan_mensual missing columns: {missing}. Found: {list(df.columns)}")

    df["employee_id"] = df["employee_id"].apply(_norm_empid)
    df["fecha"] = pd.to_datetime(df["fecha"]).dt.date
    df["run_id"] = run_id

    if "nombre" not in df.columns:
        df["nombre"] = ""
    df["nombre"] = df["nombre"].fillna("").astype(str).str.strip()

    # Cruce nombres desde case
    case_path = _find_case_xlsx(outp)
    name_info: Dict[str, Any] = {"found": False}
    if case_path:
        name_map, name_info = _load_name_map_from_case(case_path)
        if name_map is not None and not name_map.empty:
            df = df.merge(name_map, on="employee_id", how="left")
            df["nombre_case"] = df["nombre_case"].fillna("").astype(str).str.strip()
            df["nombre"] = df["nombre"].where(df["nombre"] != "", df["nombre_case"])
            df = df.drop(columns=["nombre_case"])

    # Parse shift_id
    parsed = df["shift_id"].apply(_parse_shift_id)
    df["turno_display"] = [p[0] for p in parsed]
    df["inicio"] = [p[1] for p in parsed]
    df["fin"] = [p[2] for p in parsed]
    df["colacion_min"] = [p[3] for p in parsed]

    # Canonical long
    canon_cols = [
        "run_id",
        "org_unit_id",
        "cargo",
        "employee_id",
        "nombre",
        "fecha",
        "shift_id",
        "turno_display",
        "inicio",
        "fin",
        "colacion_min",
    ]
    for optional in ["dia_semana", "es_saliente", "nota"]:
        if optional in df.columns:
            canon_cols.append(optional)

    df_long = df[canon_cols].copy()
    df_long["dow"] = pd.to_datetime(df_long["fecha"]).dt.dayofweek.map(DOW_MAP)

    # reorder safely: place dow right after fecha
    cols = [c for c in df_long.columns if c != "dow"]
    if "fecha" in cols:
        idx = cols.index("fecha") + 1
        cols.insert(idx, "dow")
    else:
        cols.insert(0, "dow")
    df_long = df_long[cols]
    df_long = df_long.loc[:, ~df_long.columns.duplicated()]

    # Save canónicos
    (outp / "turnos_long.csv").write_text(df_long.to_csv(index=False), encoding="utf-8")

    df_legend = (
        df_long[["shift_id", "turno_display", "inicio", "fin", "colacion_min"]]
        .drop_duplicates()
        .sort_values(["shift_id"])
    )
    (outp / "leyenda_turnos.csv").write_text(df_legend.to_csv(index=False), encoding="utf-8")

    # QA artifacts
    qa, issues_df = _build_qa(df_long)
    qa["run_id"] = run_id
    qa["generated_at_utc"] = _utcnow_iso()
    qa["source_plan"] = source
    qa["case_path_found"] = str(case_path) if case_path else ""
    qa["name_map_info"] = name_info

    (outp / "qa_plan.json").write_text(json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8")
    (outp / "qa_issues.csv").write_text(issues_df.to_csv(index=False), encoding="utf-8")

    # CALENDARIO pivot
    fechas = sorted(df["fecha"].unique())
    index_cols = ["org_unit_id", "cargo", "employee_id", "nombre"]

    df_cal = df[index_cols + ["fecha", "turno_display"]].copy()
    df_pivot = df_cal.pivot_table(
        index=index_cols,
        columns="fecha",
        values="turno_display",
        aggfunc="first",
        fill_value="",
    ).reset_index()

    # Excel formatting
    header_fill = PatternFill("solid", fgColor="FF1F2937")
    subheader_fill = PatternFill("solid", fgColor="FF374151")
    libre_fill = PatternFill("solid", fgColor="FFE5E7EB")
    weekend_fill = PatternFill("solid", fgColor="FFF3F4F6")

    header_font = Font(color="FFFFFF", bold=True)
    subheader_font = Font(color="FFFFFF", bold=True)

    wb = Workbook()

    # META
    ws_meta = wb.active
    ws_meta.title = "META"
    ws_meta.append(["key", "value"])

    names_total = int(df_long["employee_id"].nunique())
    names_filled = int(df_long[df_long["nombre"].fillna("").astype(str).str.strip() != ""]["employee_id"].nunique())
    fill_rate = (names_filled / names_total * 100.0) if names_total > 0 else 0.0

    meta_rows = [
        ("run_id", run_id),
        ("generated_at_utc", qa["generated_at_utc"]),
        ("source_plan", source),
        ("case_path_found", str(case_path) if case_path else ""),
        ("name_map_found", str(bool(name_info.get("found", False)))),
        ("name_map_sheet", str(name_info.get("sheet", ""))),
        ("names_total", str(names_total)),
        ("names_filled", str(names_filled)),
        ("names_fill_rate_pct", f"{fill_rate:.2f}"),
        ("duplicates_employee_fecha_rows", str(qa["duplicates_employee_fecha_rows"])),
        ("missing_turno_rows", str(qa["missing_turno_rows"])),
        ("missing_name_employees", str(qa["missing_name_employees"])),
        ("qa_plan", "qa_plan.json"),
        ("qa_issues", "qa_issues.csv"),
        ("notes", "CALENDARIO=vista humana. TURNOS_LONG=tabla canónica (BI/UI)."),
    ]
    for k, v in meta_rows:
        ws_meta.append([k, v])

    ws_meta["A1"].font = Font(bold=True)
    ws_meta["B1"].font = Font(bold=True)

    # CALENDARIO
    ws = wb.create_sheet("CALENDARIO")

    fixed_headers = ["org_unit_id", "cargo", "employee_id", "nombre"]
    date_headers = [d.strftime("%Y-%m-%d") for d in fechas]
    dow_headers = [DOW_MAP[pd.to_datetime(d).dayofweek] for d in fechas]

    ws.append(fixed_headers + date_headers)
    ws.append(["", "", "", ""] + dow_headers)

    last_col = len(fixed_headers) + len(fechas)

    for col_idx in range(1, last_col + 1):
        c1 = ws.cell(row=1, column=col_idx)
        c1.fill = header_fill
        c1.font = header_font
        c1.alignment = Alignment(horizontal="center", vertical="center")

        c2 = ws.cell(row=2, column=col_idx)
        c2.fill = subheader_fill
        c2.font = subheader_font
        c2.alignment = Alignment(horizontal="center", vertical="center")

    for _, row in df_pivot.iterrows():
        values = [row[h] for h in fixed_headers] + [row.get(d, "") for d in fechas]
        ws.append(values)

    ws.freeze_panes = "E3"
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{ws.max_row}"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 26
    for i in range(5, last_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11

    for j, d in enumerate(fechas, start=5):
        dow = pd.to_datetime(d).dayofweek
        if dow in (5, 6):
            for r in range(3, ws.max_row + 1):
                ws.cell(row=r, column=j).fill = weekend_fill

    for r in range(3, ws.max_row + 1):
        for c in range(5, last_col + 1):
            cell = ws.cell(row=r, column=c)
            val = str(cell.value or "").strip().upper()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if val == "LIBRE":
                cell.fill = libre_fill

    # TURNOS_LONG
    ws_long = wb.create_sheet("TURNOS_LONG")
    ws_long.append(list(df_long.columns))
    for cell in ws_long[1]:
        cell.font = Font(bold=True)
    for _, r in df_long.iterrows():
        ws_long.append([r.get(c, "") for c in df_long.columns])

    # LEYENDA_TURNOS
    ws_leg = wb.create_sheet("LEYENDA_TURNOS")
    ws_leg.append(list(df_legend.columns))
    for cell in ws_leg[1]:
        cell.font = Font(bold=True)
    for _, r in df_legend.iterrows():
        ws_leg.append([r.get(c, "") for c in df_legend.columns])

    out_xlsx = outp / "plan_visual.xlsx"
    wb.save(out_xlsx)
